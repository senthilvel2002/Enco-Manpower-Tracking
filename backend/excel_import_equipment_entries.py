"""
Import equipment entries from Excel (.xlsx) into MongoDB `equipment_entries`.

Supports common layouts:
- Dedicated equipment sheets (one row = one equipment entry)
- Mixed sheets with Man/Equ column (imports only rows where Man/Equ=Equipment)
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime
from typing import Any

from bson import ObjectId
from openpyxl import load_workbook

from location_normalize import normalize_work_location_name


COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "work_date": ("date", "work date", "entry date"),
    "equipment_id": ("equipment id", "equipment_id", "eq id"),
    "equipment_name": ("equipment", "equipment name", "name", "name surname", "machine"),
    "plate_number": ("plate", "plate no", "plate number", "serial", "serial no"),
    "location": ("location", "area", "site"),
    "equipment_status": ("equipment status", "eq status", "status"),
    "hours": ("hours", "total hours", "working hours", "manhours"),
    "activity": ("activity", "main activity", "today activity", "description"),
    "operator_name": ("operator", "operator name", "entered by"),
    "civil_id": ("civil id", "operator civil id", "civil_id"),
    "company_name": ("company", "company name"),
    "rental_amount": ("rental amount", "rental amt", "rental"),
    "time_from": ("from", "time from", "start time"),
    "time_to": ("to", "time to", "end time"),
    "man_equ": ("man/equ", "man or equ", "type"),
}

MAN_EQU_EQUIPMENT = {"equipment", "equip", "eq"}
EQUIP_LOCATION_MAP: dict[str, str] = {
    "yard 2": "YARD 2",
    "yard 3": "YARD 3",
    "parking area": "CONTRACTOR PARKING AREA",
    "car parking": "CONTRACTOR PARKING AREA",
    "laydown": "TSF",
    "tsf": "TSF",
    "site office": "ENCO - INDUSTRIAL AREA",
    "common area": "YARD 3",
    "reclaimer substation": "YARD 3",
}


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _parse_work_date(v: Any) -> str | None:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v or "").strip()
    if not s:
        return None
    try:
        return date.fromisoformat(s[:10]).isoformat()
    except Exception:
        pass
    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            continue
    return None


def _parse_float(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    try:
        return float(s.replace(",", ""))
    except Exception:
        return None


def _alias_to_field(text: str) -> str | None:
    t = _norm(text)
    for field, aliases in COLUMN_ALIASES.items():
        if t in (_norm(a) for a in aliases):
            return field
    return None


def _build_header_map(rows: list[tuple[Any, ...]]) -> tuple[int | None, dict[int, str]]:
    best_row_idx = None
    best_map: dict[int, str] = {}
    best_score = -1
    scan_upto = min(len(rows), 12)
    for ridx in range(scan_upto):
        r = rows[ridx]
        hmap: dict[int, str] = {}
        for cidx, cell in enumerate(r):
            f = _alias_to_field(str(cell or ""))
            if f:
                hmap[cidx] = f
        score = len(set(hmap.values()))
        if score > best_score:
            best_score = score
            best_row_idx = ridx
            best_map = hmap
    if best_score < 3:
        return None, {}
    return best_row_idx, best_map


def _looks_like_mnp_data_row(rv: tuple[Any, ...]) -> bool:
    if len(rv) < 14:
        return False
    # Common MNP layout: date at index 2, Man/Equ at 7 or 8, location around 9, hours around 13.
    has_date = _parse_work_date(rv[2]) is not None
    man_equ_7 = _norm(str(rv[7] or ""))
    man_equ_8 = _norm(str(rv[8] or ""))
    has_type = (
        man_equ_7 in MAN_EQU_EQUIPMENT
        or man_equ_8 in MAN_EQU_EQUIPMENT
        or man_equ_7 == "manpower"
        or man_equ_8 == "manpower"
    )
    return has_date and has_type


def _mnp_row_to_dict(rv: tuple[Any, ...]) -> dict[str, Any]:
    # Positional mapping for monthly MNP-like exports without headers.
    # [2]=Date, [4]=Name/Equipment, [7]/[8]=Man/Equ, [9]=Area, [12]=Shift/Status, [13]=Hours, [17]=Note
    return {
        "work_date": rv[2] if len(rv) > 2 else None,
        "company_name": str(rv[3] or "").strip() if len(rv) > 3 else "",
        "equipment_name": str(rv[4] or "").strip() if len(rv) > 4 else "",
        "plate_number": "",
        "man_equ": str(rv[7] or rv[8] or "").strip() if len(rv) > 8 else "",
        "location": str(rv[9] or "").strip() if len(rv) > 9 else "",
        "equipment_status": str(rv[12] or "").strip() if len(rv) > 12 else "",
        "hours": rv[13] if len(rv) > 13 else None,
        "activity": str(rv[17] or rv[16] or rv[15] or rv[14] or "").strip() if len(rv) > 17 else "",
    }


def _token_jaccard(a: str, b: str) -> float:
    ta = set(re.findall(r"[A-Z0-9]+", (a or "").upper()))
    tb = set(re.findall(r"[A-Z0-9]+", (b or "").upper()))
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)


def _normalize_eq_label(s: str) -> str:
    x = (s or "").upper().strip()
    x = re.sub(r"\s*\(RENTAL\)\s*", " ", x)
    x = re.sub(r"\s*-\s*(ZL|LIEBHERR|SANY|ZOOMLION|HITACHI)\s*$", "", x)
    x = re.sub(r"\s*-\s*(\d{1,2})\s*$", "", x)  # -01, -02
    x = re.sub(r"\s+", " ", x).strip()
    return x


def _extract_plate_candidates(raw_name: str) -> list[str]:
    txt = (raw_name or "").upper()
    matches = re.findall(r"\b([0-9]{2,5}[A-Z]{1,4}|[A-Z]{1,4}[0-9]{2,5}[A-Z]{0,4})\b", txt)
    out: list[str] = []
    seen: set[str] = set()
    for m in matches:
        if len(m) < 4:
            continue
        if m in seen:
            continue
        seen.add(m)
        out.append(m)
    return out


def _candidate_location_values(raw_location: str) -> list[str]:
    loc = normalize_work_location_name((raw_location or "").strip())
    nloc = _norm(loc)
    vals: list[str] = []
    mapped = EQUIP_LOCATION_MAP.get(nloc)
    if mapped:
        vals.append(mapped)
    if loc:
        vals.append(loc)
    # unique preserve order
    seen: set[str] = set()
    result: list[str] = []
    for v in vals:
        if v not in seen:
            seen.add(v)
            result.append(v)
    return result


def _resolve_equipment(db, row: dict[str, Any]) -> dict | None:
    equipment_id = (row.get("equipment_id") or "").strip()
    if equipment_id:
        oid = None
        try:
            oid = ObjectId(equipment_id)
        except Exception:
            oid = None
        eq = db.equipment_details.find_one({"_id": oid, "is_active": True}) if oid else None
        if eq:
            return eq

    raw_location = (row.get("location") or "").strip()
    plate = (row.get("plate_number") or "").strip().upper()
    raw_name = (row.get("equipment_name") or "").strip()
    name_label = _normalize_eq_label(raw_name)
    plate_hints = ([plate] if plate else []) + _extract_plate_candidates(raw_name)

    location_candidates = _candidate_location_values(raw_location)

    # Search scope: preferred mapped/exact location first, then all locations.
    search_scopes: list[list[dict[str, Any]]] = []
    for loc in location_candidates:
        cands = list(
            db.equipment_details.find(
                {"is_active": True, "location": {"$regex": "^" + re.escape(loc) + "$", "$options": "i"}}
            )
        )
        if cands:
            search_scopes.append(cands)
    search_scopes.append(list(db.equipment_details.find({"is_active": True})))

    for cands in search_scopes:
        if not cands:
            continue

        # 1) Plate-based match (strongest)
        if plate_hints:
            for ph in plate_hints:
                for eq in cands:
                    eq_plate = (eq.get("plate_number") or "").upper().strip()
                    if eq_plate and eq_plate == ph:
                        return eq

        # 2) Exact cleaned-name match
        for eq in cands:
            eq_name = _normalize_eq_label(eq.get("name") or "")
            if eq_name and eq_name == name_label:
                return eq

        # 3) Token similarity fallback
        best = None
        best_score = 0.0
        for eq in cands:
            eq_name = _normalize_eq_label(eq.get("name") or "")
            score = _token_jaccard(name_label, eq_name)
            if score > best_score:
                best_score = score
                best = eq
        if best and best_score >= 0.28:
            return best

    return None


def import_equipment_excel_rows(
    db,
    content: bytes,
    *,
    sheet_index: int = 0,
    sheet_name: str | None = None,
    dry_run: bool = False,
    default_operator_civil_id: str = "141228922",
) -> dict[str, Any]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        idx = max(0, min(sheet_index, len(wb.worksheets) - 1))
        ws = wb.worksheets[idx]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return {"ok": False, "error": "sheet is empty"}

    header_row_idx, header_map = _build_header_map(rows)
    use_mnp_positional = False
    if header_row_idx is None:
        # Fallback: monthly MNP-like files with no headers (data starts row 1).
        sample_n = min(len(rows), 40)
        sample_hits = sum(1 for rv in rows[:sample_n] if _looks_like_mnp_data_row(rv))
        if sample_hits >= 3:
            use_mnp_positional = True
        else:
            wb.close()
            return {
                "ok": False,
                "error": "Unable to detect headers. Add columns like Date, Equipment, Area/Location, Hours.",
            }

    inserted = 0
    merged = 0
    skipped = 0
    errors: list[dict[str, Any]] = []
    equipment_rows = 0

    start_idx = (header_row_idx + 1) if (header_row_idx is not None) else 0
    for ridx in range(start_idx, len(rows)):
        rv = rows[ridx]
        if use_mnp_positional:
            row = _mnp_row_to_dict(rv)
        else:
            row: dict[str, Any] = {}
            for cidx, field in header_map.items():
                if cidx < len(rv):
                    row[field] = rv[cidx]

        # Mixed file support.
        man_equ = _norm(str(row.get("man_equ") or ""))
        if man_equ and man_equ not in MAN_EQU_EQUIPMENT:
            continue

        work_date = _parse_work_date(row.get("work_date"))
        if not work_date:
            if any(str(x or "").strip() for x in rv):
                errors.append({"row": ridx + 1, "error": "missing/invalid work date"})
            continue

        equipment_rows += 1
        canonical_location = normalize_work_location_name(str(row.get("location") or "").strip())
        row["location"] = canonical_location

        eq = _resolve_equipment(db, row)
        if not eq:
            errors.append(
                {
                    "row": ridx + 1,
                    "error": f"equipment not found for name={str(row.get('equipment_name') or '').strip()!r}, plate={str(row.get('plate_number') or '').strip()!r}, location={canonical_location!r}",
                }
            )
            continue

        eid = str(eq["_id"])
        hours_value = _parse_float(row.get("hours"))
        rental_amount = _parse_float(row.get("rental_amount"))
        operator_name = (str(row.get("operator_name") or "").strip() or "Excel Import")
        civil_id = (str(row.get("civil_id") or "").strip() or default_operator_civil_id)
        company_name = (str(row.get("company_name") or "").strip())
        activity = (str(row.get("activity") or "").strip())
        equipment_status = (str(row.get("equipment_status") or "").strip() or "Working")
        time_from = (str(row.get("time_from") or "").strip() or None)
        time_to = (str(row.get("time_to") or "").strip() or None)

        doc = {
            "civil_id": civil_id,
            "company_name": company_name,
            "operator_name": operator_name,
            "equipment_id": eid,
            "equipment_name": eq.get("name", ""),
            "plate_number": eq.get("plate_number", ""),
            "equipment_type": eq.get("equipment_type", ""),
            "equipment_location": eq.get("location", ""),
            "ownership": eq.get("ownership", ""),
            "work_date": work_date,
            "location": canonical_location or eq.get("location", ""),
            "equipment_status": equipment_status,
            "time_from": time_from,
            "time_to": time_to,
            "hours": hours_value,
            "activity": activity,
            "supply_rate": eq.get("supply_rate", {}),
            "contract_rate": eq.get("contract_rate", {}),
            "rental_amount": rental_amount,
            "status": "submitted",
            "created_at": datetime.utcnow().isoformat(),
        }

        existing = db.equipment_entries.find_one({"equipment_id": eid, "work_date": work_date})
        if existing:
            if not dry_run:
                db.equipment_entries.update_one(
                    {"_id": existing["_id"]},
                    {
                        "$set": {
                            "location": doc["location"],
                            "equipment_status": doc["equipment_status"],
                            "time_from": doc["time_from"],
                            "time_to": doc["time_to"],
                            "hours": doc["hours"],
                            "activity": doc["activity"],
                            "rental_amount": doc["rental_amount"],
                            "operator_name": doc["operator_name"],
                            "civil_id": doc["civil_id"],
                            "company_name": doc["company_name"],
                            "status": "import_merged",
                        }
                    },
                )
            merged += 1
            continue

        if dry_run:
            skipped += 1
            continue

        db.equipment_entries.insert_one(doc)
        inserted += 1

    wb.close()
    return {
        "ok": True,
        "sheet": ws.title,
        "rows_read": max(0, len(rows) - start_idx),
        "equipment_rows": equipment_rows,
        "inserted": inserted,
        "merged": merged,
        "dry_run_skipped": skipped,
        "mode": "mnp-positional" if use_mnp_positional else "header",
        "errors": errors[:200],
        "error_count": len(errors),
    }

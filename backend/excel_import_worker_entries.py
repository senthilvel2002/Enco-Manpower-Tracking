"""
Import worker work entries from Excel (.xlsx) into MongoDB `work_entries`.

Expected: row 1 = headers. Column names are matched flexibly (see COLUMN_ALIASES).

Usage (from backend folder, with MONGODB_URI in .env):
  python scripts/import_worker_entries_excel.py path/to/file.xlsx
  python scripts/import_worker_entries_excel.py path/to/file.xlsx --dry-run

HTTP (management only):
  POST /api/management/work-entries/import-excel
  multipart: file=<xlsx>, management_password=<password>
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook

from excel_import_progress import progress_done, progress_line, progress_message, progress_step
from location_normalize import normalize_work_location_name

DEVELOPER_MASTER_CIVIL_ID = "141228922"

# Excel header (any of these labels, case-insensitive, extra spaces OK) -> internal field name
COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "civil_id": ("civil id", "civil_id", "civilid", "national id", "id"),
    "company_name": ("company", "company name", "company_name", "employer"),
    "worker_name": ("worker name", "worker_name", "name", "employee name"),
    "work_date": ("work date", "work_date", "date", "day"),
    "location": ("location", "site", "working location"),
    "incharge": ("incharge", "in-charge", "site incharge", "supervisor"),
    "permit_issuer": ("permit issuer", "permit_issuer", "ptw", "issuer"),
    "today_activity": ("today activity", "today_activity", "activity", "main activity", "work activity"),
    "worker_time_from": ("time from", "worker_time_from", "from", "start time", "from time"),
    "worker_time_to": ("time to", "worker_time_to", "to", "end time", "to time"),
    "worker_hours": ("worker hours", "worker_hours", "hours", "hrs"),
    "worker_shift": ("worker shift", "worker_shift", "shift"),
    "leave_reason": ("leave reason", "leave_reason", "reason"),
    "item_tag": ("item tag", "item_tag", "tag"),
}


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _build_header_map(header_row: list[Any]) -> dict[int, str]:
    """Map column index -> canonical field name."""
    col_to_field: dict[int, str] = {}
    alias_lookup: dict[str, str] = {}
    for field, aliases in COLUMN_ALIASES.items():
        for a in aliases:
            alias_lookup[_norm(a)] = field

    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        key = _norm(str(cell))
        if not key:
            continue
        if key in alias_lookup:
            col_to_field[idx] = alias_lookup[key]
    return col_to_field


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d") if val.time() == datetime.min.time() else val.isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    if isinstance(val, int):
        return str(val)
    return str(val).strip()


def _parse_work_date(val: Any) -> str | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    try:
        return date.fromisoformat(s[:10]).isoformat()
    except ValueError:
        pass
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _parse_hhmm(value: str):
    if not value:
        return None
    parts = value.split(":")
    if len(parts) != 2:
        return None
    try:
        hh = int(parts[0])
        mm = int(parts[1])
    except ValueError:
        return None
    if hh < 0 or hh > 23 or mm < 0 or mm > 59:
        return None
    return hh * 60 + mm


def _normalize_time_cell(val: Any) -> str:
    """Excel may store time as datetime or fraction; normalize to HH:MM."""
    if val is None or val == "":
        return ""
    if isinstance(val, datetime):
        return f"{val.hour:02d}:{val.minute:02d}"
    if isinstance(val, (int, float)):
        frac = float(val) % 1.0
        if frac < 0:
            return ""
        total_min = int(round(frac * 24 * 60))
        h = total_min // 60
        m = total_min % 60
        return f"{h:02d}:{m:02d}"
    s = str(val).strip()
    if re.match(r"^\d{1,2}:\d{2}$", s):
        parts = s.split(":")
        return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
    return s


def row_to_payload(row: list[Any], col_to_field: dict[int, str]) -> dict[str, Any]:
    raw: dict[str, Any] = {}
    for idx, cell in enumerate(row):
        field = col_to_field.get(idx)
        if not field:
            continue
        if field == "work_date":
            raw[field] = _parse_work_date(cell) or _cell_str(cell)
        elif field in ("worker_time_from", "worker_time_to"):
            raw[field] = _normalize_time_cell(cell)
        elif field == "worker_hours":
            v = cell
            if v is None or v == "":
                raw[field] = None
            else:
                try:
                    raw[field] = float(v)
                except (TypeError, ValueError):
                    raw[field] = str(v).strip()
        else:
            raw[field] = _cell_str(cell)
    return raw


def validate_and_build_document(db, payload: dict[str, Any]) -> tuple[dict | None, str | None]:
    """
    Same rules as POST /api/work-entries, but work_date comes from payload (required).
    """
    civil_id = (payload.get("civil_id") or "").strip()
    company_name = (payload.get("company_name") or "").strip()
    worker_name = (payload.get("worker_name") or "").strip()
    work_date_raw = payload.get("work_date")
    work_date = _parse_work_date(work_date_raw) or ""
    if not work_date and work_date_raw not in (None, ""):
        return None, f"invalid work_date: {work_date_raw!r}"

    location = normalize_work_location_name((payload.get("location") or "").strip())
    incharge = (payload.get("incharge") or "").strip()
    permit_issuer = (payload.get("permit_issuer") or "").strip()
    today_activity = (payload.get("today_activity") or "").strip()
    worker_time_from = (payload.get("worker_time_from") or "").strip()
    worker_time_to = (payload.get("worker_time_to") or "").strip()
    worker_shift = (payload.get("worker_shift") or "Day").strip()
    leave_reason = (payload.get("leave_reason") or "").strip()
    item_tag = (payload.get("item_tag") or "").strip() or None

    allowed_worker_shifts = {"Day", "Night", "Request for leave"}
    if worker_shift not in allowed_worker_shifts:
        return None, "worker_shift must be Day, Night, or Request for leave"
    if worker_shift == "Request for leave" and not leave_reason:
        return None, "leave_reason is required for leave request"

    if not civil_id or not company_name or not worker_name or not work_date:
        return None, "missing civil_id, company_name, worker_name, or work_date"

    worker_hours = payload.get("worker_hours")
    worker_hours_value = None
    if worker_hours is not None and worker_hours != "":
        try:
            worker_hours_value = float(worker_hours)
        except (TypeError, ValueError):
            return None, "worker_hours must be a valid number"
        if worker_hours_value < 0 or worker_hours_value > 12:
            return None, "work hours must be between 0 and 12"

    if worker_time_from or worker_time_to:
        start_min = _parse_hhmm(worker_time_from)
        end_min = _parse_hhmm(worker_time_to)
        if start_min is None or end_min is None:
            return None, "work time must be in HH:MM format"
        if end_min <= start_min:
            return None, "to time must be after from time"
        diff_hours = (end_min - start_min) / 60.0
        if diff_hours > 12:
            return None, "work hours must be between 0 and 12"
        worker_hours_value = diff_hours

    if worker_shift != "Request for leave" and worker_hours_value is None:
        return None, "work hours are required (From/To time or worker_hours)"
    if worker_shift == "Request for leave":
        worker_hours_value = 0.0
        worker_time_from = ""
        worker_time_to = ""

    location_value = location or None
    incharge_value = incharge or None
    permit_issuer_value = permit_issuer or None

    if not db.companies.find_one({"name": company_name, "is_active": True}):
        return None, "invalid company_name"

    if location_value and not db.locations.find_one({"name": location_value, "is_active": True}):
        return None, "invalid location"

    if incharge_value:
        inc_doc = db.incharge.find_one({"name": incharge_value, "is_active": True})
        if not inc_doc:
            return None, "invalid incharge"
        inc_locations = inc_doc.get("locations")
        if (
            location_value
            and isinstance(inc_locations, list)
            and inc_locations
            and location_value not in inc_locations
        ):
            return None, "incharge does not match selected location"

    if location_value in {"Head Office", "Laydown Office"}:
        permit_issuer_value = None
    if permit_issuer_value and not db.permit_issuer.find_one({"name": permit_issuer_value, "is_active": True}):
        return None, "invalid permit_issuer"

    worker = db.worker_details.find_one({"civil_id": civil_id, "company_name": company_name, "is_active": True})
    if not worker and civil_id != DEVELOPER_MASTER_CIVIL_ID:
        return None, "worker profile not found"

    if worker_time_from and worker_time_to:
        duplicate_timing = db.work_entries.find_one(
            {
                "civil_id": civil_id,
                "work_date": work_date,
                "worker_time_from": worker_time_from,
                "worker_time_to": worker_time_to,
                "status": {"$nin": ["cancelled"]},
            }
        )
        if duplicate_timing:
            return None, "duplicate_timing"

    doc = {
        "civil_id": civil_id,
        "company_name": company_name,
        "worker_name": worker_name,
        "work_date": work_date,
        "location": location_value,
        "incharge": incharge_value,
        "permit_issuer": permit_issuer_value,
        "worker_hours": worker_hours_value,
        "worker_time_from": worker_time_from or None,
        "worker_time_to": worker_time_to or None,
        "worker_shift": worker_shift,
        "leave_reason": leave_reason if worker_shift == "Request for leave" else None,
        "hours": None,
        "attendance_status": None,
        "today_activity": today_activity,
        "item_tag": item_tag,
        "status": "submitted",
    }
    return doc, None


def parse_excel_workbook(content: bytes, sheet_index: int = 0) -> tuple[list[dict[str, Any]], list[str]]:
    """
    Parse .xlsx bytes into list of row dicts. Returns (rows, errors for sheet/header).
    """
    errors: list[str] = []
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheets = wb.sheetnames
        if sheet_index >= len(sheets):
            return [], [f"sheet index {sheet_index} not found"]
        ws = wb[sheets[sheet_index]]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return [], ["empty sheet"]
        col_to_field = _build_header_map(list(header_row))
        required_any = {"civil_id", "company_name", "worker_name", "work_date"}
        missing = required_any - set(col_to_field.values())
        if missing:
            errors.append(
                "missing required columns (match by header name): "
                + ", ".join(sorted(missing))
                + ". See excel_import_worker_entries.COLUMN_ALIASES"
            )
        result: list[dict[str, Any]] = []
        for rnum, row in enumerate(rows_iter, start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            payload = row_to_payload(list(row), col_to_field)
            if not payload.get("civil_id") and not any(payload.values()):
                continue
            payload["_row_number"] = rnum
            result.append(payload)
    finally:
        wb.close()
    return result, errors


def import_excel_rows(
    db,
    content: bytes,
    *,
    sheet_index: int = 0,
    dry_run: bool = False,
    progress: bool = False,
) -> dict[str, Any]:
    """
    Parse Excel and insert each valid row. Returns summary dict.
    """
    rows, parse_errors = parse_excel_workbook(content, sheet_index=sheet_index)
    inserted_ids: list[str] = []
    row_errors: list[dict[str, Any]] = []
    nrows = len(rows)
    step = progress_step(nrows)
    if progress:
        progress_message(
            f"Flat import (sheet index {sheet_index}): {nrows} rows — "
            f"{'dry-run' if dry_run else 'inserting'}…"
        )

    for j, payload in enumerate(rows, 1):
        rnum = payload.pop("_row_number", None)
        doc, err = validate_and_build_document(db, payload)
        if err:
            row_errors.append({"row": rnum, "error": err, "civil_id": payload.get("civil_id")})
            if progress and (j % step == 0 or j == nrows):
                progress_line(j, nrows, "rows ")
            continue
        if dry_run:
            inserted_ids.append("(dry-run)")
        else:
            ins = db.work_entries.insert_one(doc)
            inserted_ids.append(str(ins.inserted_id))
        if progress and (j % step == 0 or j == nrows):
            progress_line(j, nrows, "rows ")
    if progress and nrows:
        progress_done()

    return {
        "ok": True,
        "parse_errors": parse_errors,
        "rows_read": len(rows),
        "inserted": len(inserted_ids) if not dry_run else 0,
        "would_insert": len(inserted_ids) if dry_run else len(inserted_ids),
        "row_errors": row_errors,
        "inserted_ids": inserted_ids if not dry_run else [],
    }

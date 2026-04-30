"""
Direct import: 25-EG-OM-017-Mnp Tracking-March 2026-Updated.xlsx  →  MongoDB

What this does:
  1. Auto-creates any missing companies, locations and worker profiles in master data
  2. Imports all Manpower / Sub Contractor / Driver rows → work_entries
  3. Imports all Equipment rows → equipment_entries

Usage (from backend folder, with venv active):
    python scripts/import_march_2026.py            # live import
    python scripts/import_march_2026.py --dry-run  # validate only, no writes
"""

from __future__ import annotations

import os
import re
import sys
import warnings
from collections import defaultdict
from datetime import date, datetime
from typing import Any

warnings.filterwarnings("ignore", message="Data Validation extension is not supported")

_BACKEND_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _BACKEND_ROOT not in sys.path:
    sys.path.insert(0, _BACKEND_ROOT)

try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(_BACKEND_ROOT, ".env"))
except Exception:
    pass

from openpyxl import load_workbook  # noqa: E402
from db import get_database          # noqa: E402

# ─── CONFIG ──────────────────────────────────────────────────────────────────
XLSX_PATH = os.path.join(
    _BACKEND_ROOT, "..", "25-EG-OM-017-Mnp Tracking-March 2026-Updated.xlsx"
)
SHEET_NAME = "March 2026"
DRY_RUN = "--dry-run" in sys.argv

# Civil ID for equipment entries (developer / import account)
IMPORT_OPERATOR_CIVIL_ID = "141228922"

# Map Excel "Area" column → equipment_details.location  (for equipment matching)
# When multiple possible locations, we fall back to all-location search anyway.
EQUIP_LOCATION_MAP: dict[str, str] = {
    "yard 2":              "YARD 2",
    "yard 3":              "YARD 3",
    "parking area":        "CONTRACTOR PARKING AREA",
    "laydown":             "TSF",          # most laydown equipment lives under TSF in DB
    "tsf":                 "TSF",
    "site office":         "ENCO - INDUSTRIAL AREA",
    "common area":         "YARD 3",
    "reclaimer substation":"YARD 3",
}

MANPOWER_TYPES = frozenset(
    {"manpower", "man", "sub contactor", "subcontractor", "sub contractor", "driver"}
)

# Excel header text → internal field  (normalised lowercase)
HEADER_ALIASES: dict[str, str] = {
    "date":          "work_date",
    "company":       "company_name",
    "name surname":  "worker_name",
    "column1":       "civil_id",
    "id #":          "civil_id",
    "civil id":      "civil_id",
    "civil_id":      "civil_id",
    "designation":   "designation",
    "man/equ":       "man_equ",
    "area":          "location",
    "item tag":      "item_tag",
    "shift":         "worker_shift",
    "manhours":      "worker_hours",
    "main activity": "main_activity",
    "sub act":       "sub_act",
    "sub-sub act":   "sub_sub_act",
    "note":          "note",
    "filled by":     "filled_by",
}

MAX_COLS = 25   # only first 25 columns contain data


# ─── HELPERS ─────────────────────────────────────────────────────────────────

def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _parse_date(v: Any) -> str | None:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v or "").strip()
    try:
        return date.fromisoformat(s[:10]).isoformat()
    except Exception:
        return None


def _parse_civil(v: Any) -> str:
    if v is None:
        return ""
    s = _cell_str(v).strip()
    return "" if s.upper() in ("#N/A", "N/A", "-", "") else s


def _token_jaccard(a: str, b: str) -> float:
    ta = set(re.findall(r"[A-Z0-9]+", (a or "").upper()))
    tb = set(re.findall(r"[A-Z0-9]+", (b or "").upper()))
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)


def _infer_category(designation: str) -> str:
    try:
        from app import _infer_worker_category
        return _infer_worker_category(designation)
    except Exception:
        return "Direct"


def _make_activity(m: dict, shift_raw: str) -> str:
    parts: list[str] = []
    for k in ("main_activity", "sub_act", "sub_sub_act", "note"):
        v = (m.get(k) or "").strip()
        if v and v != "_":
            parts.append(v)
    return " / ".join(parts) or ("Idle" if shift_raw == "idle" else "Work")


# ─── EXCEL PARSING ───────────────────────────────────────────────────────────

def parse_sheet(ws) -> list[dict[str, Any]]:
    header_row = next(
        ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=MAX_COLS, values_only=True),
        None,
    )
    header_map: dict[int, str] = {}
    for idx, cell in enumerate(header_row or []):
        if cell is None:
            continue
        k = _norm(str(cell))
        if k in HEADER_ALIASES:
            header_map[idx] = HEADER_ALIASES[k]

    rows: list[dict[str, Any]] = []
    for ridx, row in enumerate(
        ws.iter_rows(min_row=2, min_col=1, max_col=MAX_COLS, values_only=True), start=2
    ):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        m: dict[str, Any] = {"_row": ridx}
        for idx, cell in enumerate(row):
            field = header_map.get(idx)
            if not field:
                continue
            if field == "work_date":
                m[field] = _parse_date(cell) or _cell_str(cell)
            elif field == "worker_hours":
                try:
                    m[field] = float(cell) if cell is not None and cell != "" else None
                except (TypeError, ValueError):
                    m[field] = None
            elif field == "civil_id":
                m[field] = _parse_civil(cell)
            else:
                m[field] = _cell_str(cell)
        if not m.get("worker_name") and not m.get("company_name"):
            continue
        rows.append(m)
    return rows


# ─── MAIN ────────────────────────────────────────────────────────────────────

def main() -> None:
    print(f"Connecting to database...")
    db = get_database()
    print("  Connected.")

    # ── Build in-memory caches (single round-trip each) ──────────────────────
    print("Loading master data caches...")

    company_cache: dict[str, str] = {}      # norm_name → canonical_name
    for c in db.companies.find({"is_active": True}, {"name": 1}):
        company_cache[_norm(c["name"])] = c["name"]

    location_cache: dict[str, str] = {}     # norm_name → canonical_name
    for l in db.locations.find({"is_active": True}, {"name": 1}):
        location_cache[_norm(l["name"])] = l["name"]

    worker_by_cid: dict[str, dict] = {}     # civil_id → worker doc
    worker_by_name: dict[str, dict] = {}    # norm(name)|norm(company) → worker doc
    for w in db.worker_details.find(
        {"is_active": True},
        {"civil_id": 1, "name": 1, "company_name": 1, "designation": 1},
    ):
        cid = w.get("civil_id")
        if cid:
            worker_by_cid[cid] = w
        key = _norm(w.get("name") or "") + "|" + _norm(w.get("company_name") or "")
        worker_by_name[key] = w

    eq_cache = list(db.equipment_details.find({"is_active": True}))

    # Existing work_entries for March 2026 — load civil_id+date+shift to skip dups
    existing_wp: set[tuple[str, str, str]] = set()
    for e in db.work_entries.find(
        {"work_date": {"$gte": "2026-03-01", "$lte": "2026-03-31"}},
        {"civil_id": 1, "work_date": 1, "worker_shift": 1},
    ):
        existing_wp.add((e.get("civil_id", ""), e.get("work_date", ""), e.get("worker_shift", "")))

    # Existing equipment_entries for March 2026
    existing_eq: set[tuple[str, str]] = set()
    for e in db.equipment_entries.find(
        {"work_date": {"$gte": "2026-03-01", "$lte": "2026-03-31"}},
        {"equipment_id": 1, "work_date": 1},
    ):
        existing_eq.add((e.get("equipment_id", ""), e.get("work_date", "")))

    print(
        f"  {len(company_cache)} companies, {len(location_cache)} locations, "
        f"{len(worker_by_cid)} workers, {len(eq_cache)} equipment items cached."
    )
    print(
        f"  {len(existing_wp)} March work entries already in DB, "
        f"{len(existing_eq)} March equipment entries already in DB."
    )

    # ── Helper closures (use caches, no per-row DB hits) ─────────────────────

    def _ensure_company(name: str) -> str:
        name = name.strip()
        if not name:
            return ""
        key = _norm(name)
        if key in company_cache:
            return company_cache[key]
        if not DRY_RUN:
            try:
                db.companies.insert_one({"name": name, "is_active": True})
            except Exception:
                pass
        company_cache[key] = name
        return name

    def _ensure_location(name: str) -> str:
        name = name.strip()
        if not name:
            return ""
        key = _norm(name)
        if key in location_cache:
            return location_cache[key]
        if not DRY_RUN:
            try:
                db.locations.insert_one({"name": name, "is_active": True})
            except Exception:
                pass
        location_cache[key] = name
        return name

    def _get_worker(civil_id: str, name: str, company: str, desig: str) -> dict | None:
        if civil_id and civil_id in worker_by_cid:
            return worker_by_cid[civil_id]
        key = _norm(name) + "|" + _norm(company)
        if key in worker_by_name:
            return worker_by_name[key]
        if civil_id:
            # Create & cache
            if not DRY_RUN:
                try:
                    db.worker_details.insert_one({
                        "civil_id": civil_id,
                        "name": name.strip(),
                        "designation": desig.strip(),
                        "category": _infer_category(desig),
                        "company_name": company,
                        "email": None,
                        "is_active": True,
                    })
                except Exception:
                    pass
                w = db.worker_details.find_one({"civil_id": civil_id, "is_active": True})
            else:
                # Simulate a worker for dry-run counting
                w = {"civil_id": civil_id, "name": name, "company_name": company}
            if w:
                worker_by_cid[civil_id] = w
                worker_by_name[_norm(w.get("name") or "") + "|" + _norm(w.get("company_name") or "")] = w
            return w
        return None

    def _find_eq(raw_name: str, excel_loc: str) -> dict | None:
        eq_loc = EQUIP_LOCATION_MAP.get(_norm(excel_loc or ""))
        plate_m = re.search(
            r"\b([A-Z]{1,4}\d{3,8}[A-Z]{0,4}|\d{3,8}[A-Z]{2,4})\s*$",
            raw_name.upper(),
        )
        plate_hint = plate_m.group(1) if plate_m else None

        # Pass 1: search within mapped location; Pass 2: all locations
        for cands in (
            [e for e in eq_cache if eq_loc and e.get("location") == eq_loc],
            eq_cache,
        ):
            if not cands:
                continue
            if plate_hint:
                for eq in cands:
                    if (eq.get("plate_number") or "").upper() == plate_hint:
                        return eq
            clean = re.sub(r"\s*\(RENTAL\)\s*", " ", raw_name, flags=re.I)
            clean = re.sub(
                r"\s*-\s*(ZL|Liebherr|Sany|Zoomlion|01|02|03|04|1|2|3|4)\s*$", "", clean, flags=re.I
            ).strip().upper()
            best, best_score = None, 0.0
            for eq in cands:
                en = re.sub(r"\s*\(RENTAL\)\s*", " ", eq.get("name") or "", flags=re.I).strip().upper()
                score = _token_jaccard(clean, en)
                if score > best_score:
                    best_score, best = score, eq
            if best and best_score >= 0.30:
                return best
        return None

    # ── Parse Excel ──────────────────────────────────────────────────────────
    print(f"\n{'[DRY RUN] ' if DRY_RUN else ''}Reading Excel...")
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: sheet {SHEET_NAME!r} not found. Available: {wb.sheetnames}")
        wb.close()
        sys.exit(1)
    rows = parse_sheet(wb[SHEET_NAME])
    wb.close()

    mp_rows = [r for r in rows if _norm(r.get("man_equ") or "") in MANPOWER_TYPES]
    eq_rows = [r for r in rows if _norm(r.get("man_equ") or "") == "equipment"]
    print(f"  {len(rows)} total rows  |  {len(mp_rows)} manpower  |  {len(eq_rows)} equipment")

    # ── Step 1: companies ─────────────────────────────────────────────────────
    print("\n[1] Ensuring companies...")
    comp_new: list[str] = []
    for m in rows:
        c = (m.get("company_name") or "").strip()
        if c and _norm(c) not in company_cache:
            _ensure_company(c)
            comp_new.append(c)
    print(f"  {len(comp_new)} new: {comp_new[:8]}")

    # ── Step 2: locations ─────────────────────────────────────────────────────
    print("\n[2] Ensuring locations...")
    loc_new: list[str] = []
    for m in mp_rows + eq_rows:
        loc = (m.get("location") or "").strip()
        if loc and _norm(loc) not in location_cache:
            _ensure_location(loc)
            loc_new.append(loc)
    print(f"  {len(loc_new)} new: {loc_new}")

    # ── Step 3: workers ───────────────────────────────────────────────────────
    print("\n[3] Ensuring workers...")
    workers_new = 0
    seen_cids: set[str] = set()
    for m in mp_rows:
        civil_id = m.get("civil_id", "")
        name = (m.get("worker_name") or "").strip()
        company = _ensure_company((m.get("company_name") or "").strip())
        desig = (m.get("designation") or "").strip()
        if not civil_id or civil_id in seen_cids:
            continue
        seen_cids.add(civil_id)
        if civil_id not in worker_by_cid:
            _get_worker(civil_id, name, company, desig)
            workers_new += 1
    print(f"  {workers_new} new workers created.")

    # ── Step 4: manpower work entries ─────────────────────────────────────────
    print("\n[4] Importing manpower entries...")
    mp_docs: list[dict] = []
    mp_dup = mp_err = 0
    mp_error_list: list[str] = []

    for m in mp_rows:
        rnum = m.get("_row")
        civil_id = m.get("civil_id", "")
        name = (m.get("worker_name") or "").strip()
        company = _ensure_company((m.get("company_name") or "").strip())
        desig = (m.get("designation") or "").strip()
        work_date = m.get("work_date")

        if not work_date or not name:
            mp_err += 1
            mp_error_list.append(f"Row {rnum}: missing date or name")
            continue

        w = _get_worker(civil_id, name, company, desig)
        if not w:
            mp_err += 1
            mp_error_list.append(f"Row {rnum}: worker not found — {name!r} ({company})")
            continue

        # Shift
        shift_raw = (m.get("worker_shift") or "").strip().lower()
        if shift_raw == "night":
            worker_shift = "Night"
        elif "leave" in shift_raw:
            worker_shift = "Request for leave"
        else:
            worker_shift = "Day"

        # Duplicate check (in-memory)
        dup_key = (w["civil_id"], work_date, worker_shift)
        if dup_key in existing_wp:
            mp_dup += 1
            continue
        existing_wp.add(dup_key)   # prevent re-insert within same run

        # Hours
        wh = m.get("worker_hours")
        try:
            hours_val: float | None = float(wh) if wh is not None and wh != "" else None
        except (TypeError, ValueError):
            hours_val = None
        if worker_shift == "Request for leave":
            hours_val = 0.0
        elif hours_val is None:
            hours_val = 8.0
        elif hours_val > 12:
            hours_val = 12.0

        activity = _make_activity(m, shift_raw)
        loc_canonical = location_cache.get(_norm((m.get("location") or "").strip()))

        mp_docs.append({
            "civil_id": w["civil_id"],
            "company_name": w.get("company_name") or company,
            "worker_name": w.get("name") or name,
            "work_date": work_date,
            "location": loc_canonical,
            "incharge": None,
            "permit_issuer": None,
            "worker_hours": hours_val,
            "worker_time_from": None,
            "worker_time_to": None,
            "worker_shift": worker_shift,
            "leave_reason": activity if worker_shift == "Request for leave" else None,
            "hours": None,
            "attendance_status": None,
            "today_activity": activity,
            "item_tag": (m.get("item_tag") or None) or None,
            "status": "submitted",
        })

    if not DRY_RUN and mp_docs:
        batch_size = 100
        for i in range(0, len(mp_docs), batch_size):
            db.work_entries.insert_many(mp_docs[i:i + batch_size], ordered=False)
            print(f"    Batch {i // batch_size + 1}: inserted {min(i + batch_size, len(mp_docs))}/{len(mp_docs)}")
    print(
        f"  Inserted: {len(mp_docs)}  Duplicates skipped: {mp_dup}  Errors: {mp_err}"
    )
    if mp_error_list:
        print(f"  First errors (up to 10):")
        for e in mp_error_list[:10]:
            print(f"    {e}")

    # ── Step 5: equipment entries ─────────────────────────────────────────────
    print("\n[5] Importing equipment entries...")

    # Group by (date, norm_equipment_name, norm_location)
    eq_groups: dict[tuple[str, str, str], list[dict]] = defaultdict(list)
    for m in eq_rows:
        wd = m.get("work_date")
        if not wd:
            continue
        key = (wd, _norm(m.get("worker_name") or ""), _norm(m.get("location") or ""))
        eq_groups[key].append(m)

    eq_docs: list[dict] = []
    eq_err = eq_dup = 0
    eq_error_list: list[str] = []

    for (wd, _en, _el), group in eq_groups.items():
        first = group[0]
        raw_name = (first.get("worker_name") or "").strip()
        excel_loc = (first.get("location") or "").strip()

        eq = _find_eq(raw_name, excel_loc)
        if not eq:
            eq_err += 1
            eq_error_list.append(
                f"Row {first.get('_row')}: no match for {raw_name!r} @ {excel_loc!r}"
            )
            continue

        eid = str(eq["_id"])
        dup_key = (eid, wd)
        if dup_key in existing_eq:
            eq_dup += 1
            continue
        existing_eq.add(dup_key)

        # Merge hours + activities
        total_h = 0.0
        acts: list[str] = []
        for x in group:
            wh = x.get("worker_hours")
            try:
                total_h += float(wh) if wh is not None and wh != "" else 0.0
            except (TypeError, ValueError):
                pass
            a = _make_activity(x, "")
            if a and a != "Work":
                acts.append(a)
        if total_h <= 0:
            total_h = 8.0
        total_h = min(total_h, 24.0)

        end_min = 6 * 60 + int(round(total_h * 60))
        eh = min(end_min // 60, 23)
        em = end_min % 60 if end_min // 60 <= 23 else 59
        tf, tt = "06:00", f"{eh:02d}:{em:02d}"

        loc_canonical = _ensure_location(excel_loc)

        eq_docs.append({
            "civil_id": IMPORT_OPERATOR_CIVIL_ID,
            "company_name": "",
            "operator_name": "MNP Import — March 2026",
            "equipment_id": eid,
            "equipment_name": eq.get("name", ""),
            "plate_number": eq.get("plate_number", ""),
            "equipment_type": eq.get("equipment_type", ""),
            "equipment_location": eq.get("location", ""),
            "ownership": eq.get("ownership", ""),
            "work_date": wd,
            "location": loc_canonical,
            "equipment_status": "Working",
            "time_from": tf,
            "time_to": tt,
            "hours": round(total_h, 2),
            "activity": " | ".join(dict.fromkeys(acts)) or "Equipment work",
            "supply_rate": eq.get("supply_rate") or {},
            "contract_rate": eq.get("contract_rate") or {},
            "rental_amount": None,
            "status": "submitted",
            "created_at": datetime.now().isoformat(),
        })

    if not DRY_RUN and eq_docs:
        batch_size = 100
        for i in range(0, len(eq_docs), batch_size):
            db.equipment_entries.insert_many(eq_docs[i:i + batch_size], ordered=False)
            print(f"    Batch {i // batch_size + 1}: inserted {min(i + batch_size, len(eq_docs))}/{len(eq_docs)}")
    print(
        f"  Inserted: {len(eq_docs)}  Duplicates skipped: {eq_dup}  Errors: {eq_err}"
    )
    if eq_error_list:
        print(f"  Equipment errors (up to 15):")
        for e in eq_error_list[:15]:
            print(f"    {e}")

    # ── Summary ───────────────────────────────────────────────────────────────
    print("\n================= SUMMARY =================")
    print(f"  Manpower entries inserted : {len(mp_docs)}")
    print(f"  Manpower duplicates skip  : {mp_dup}")
    print(f"  Manpower errors           : {mp_err}")
    print(f"  Equipment entries inserted: {len(eq_docs)}")
    print(f"  Equipment duplicates skip : {eq_dup}")
    print(f"  Equipment errors          : {eq_err}")
    if DRY_RUN:
        print("  *** DRY RUN --- no writes were made ***")
    print("============================================")


if __name__ == "__main__":
    main()

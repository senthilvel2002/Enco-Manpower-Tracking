"""
Import worker entries from Enco MNP tracking Excel (e.g. *Mnp Tracking*.xlsx).

Sheet layout (row is auto-detected, usually 1 or 2):
  Sl. No | Date | Company | Name Surname | Civil ID | Designation | Man/Equ | Area | … | Manhours | Main Activity | …

- Man/Equ = Equipment → equipment_entries (matched by plate/name + location; merged if duplicate same day).
- Manpower, Sub Contractor, Driver → work_entries.
- Civil ID column (or Excel "Column1") resolves the worker; else company + name; optional create_missing_workers.
- Shift "Idle" is stored as worker_shift Day with hours from Manhours; activity text prefers Main/Sub/Note.
"""

from __future__ import annotations

import io
import re
from collections import defaultdict
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook
from pymongo.errors import DuplicateKeyError

from excel_import_progress import progress_done, progress_line, progress_message, progress_step
from excel_import_worker_entries import (
    DEVELOPER_MASTER_CIVIL_ID,
    _cell_str,
    _parse_work_date,
    validate_and_build_document,
)
from location_normalize import normalize_work_location_name

# Sheets to skip when using --all-sheets
MNP_SKIP_SHEETS = frozenset(
    {
        "Cumulative",
        "Tasks 1",
        "Manpower 1",
        "Tasks",
        "Manpower",
    }
)
# When importing every tab, skip the monthly roll-up to avoid duplicating the same rows as daily tabs.
MNP_SKIP_ALL_SHEETS = MNP_SKIP_SHEETS | {"January 2026"}

# When auto-picking a tab, skip obvious non-data sheets before scanning for headers.
_MNP_DEFAULT_SHEET_SKIP = frozenset(
    {"Cumulative", "Tasks 1", "Manpower 1", "Tasks", "Manpower", "ENCOGROUP", "Sheet1", "April 2026"}
)

# Monthly roll-up tab names: "Feb 2026", "February 2026", "Jan 2026", … (not day tabs like "02-Feb-26").
_DEFAULT_MNP_MONTHLY_TAB_RE = re.compile(
    r"(?i)^\s*(?:jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)\s+20\d{2}\s*$"
)

# Header cell text -> field (normalized matching)
MNP_ALIASES: dict[str, tuple[str, ...]] = {
    "work_date": ("date",),
    "company_name": ("company",),
    "worker_name": ("name surname", "name", "employee name"),
    "civil_id": ("civil id", "civil_id", "id #", "national id", "column1"),
    "designation": ("designation",),
    "man_equ": ("man/equ", "man equ", "man or equ"),
    "location": ("area", "location", "site"),
    "item_tag": ("item tag", "item_tag"),
    "worker_shift": ("shift",),
    "worker_hours": ("manhours", "man hours", "manhour"),
    "main_activity": ("main activity", "main_activity"),
    "sub_act": ("sub act", "sub_act"),
    "sub_sub_act": ("sub-sub act", "sub sub act", "sub_sub"),
    "note": ("note",),
    "filled_by": ("filled by", "filled_by"),
}


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _build_mnp_header_map(header_row: list[Any]) -> dict[int, str]:
    col_to_field: dict[int, str] = {}
    alias_lookup: dict[str, str] = {}
    for field, aliases in MNP_ALIASES.items():
        for a in aliases:
            alias_lookup[_norm(a)] = field
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        key = _norm(str(cell))
        if not key:
            continue
        if key in alias_lookup:
            col_to_field[idx] = alias_lookup[key]
    return col_to_field


def _row_looks_like_mnp_header(row: list[Any]) -> bool:
    norms = [_norm(str(c)) if c is not None else "" for c in row]
    joined = " ".join(norms)
    has_date = "date" in norms
    has_company = "company" in norms
    has_name = "name" in joined and "surname" in joined
    return has_date and has_company and has_name


def _find_header_row(ws) -> tuple[int | None, list[Any] | None]:
    for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        row_list = list(row)
        if _row_looks_like_mnp_header(row_list):
            return ridx, row_list
    return None, None


def _default_mnp_ws_name(wb) -> str:
    """
    When no sheet is specified: prefer January 2026 (legacy), then any «Month YYYY» roll-up tab
    (e.g. Feb 2026), else the first sheet that has an MNP header row.

    February workbooks often have «Cumulative» first — that tab has no MNP header, so picking
    sheetnames[0] yielded 0 rows.
    """
    names = list(wb.sheetnames)
    if not names:
        return ""
    if "January 2026" in names:
        return "January 2026"
    for n in names:
        if n and _DEFAULT_MNP_MONTHLY_TAB_RE.match(n.strip()):
            return n
    for n in names:
        if n.strip() in _MNP_DEFAULT_SHEET_SKIP:
            continue
        try:
            ws_try = wb[n]
        except Exception:
            continue
        _hr, header_row = _find_header_row(ws_try)
        if header_row:
            return n
    return names[0]


def _parse_civil_id_cell(cell: Any) -> str:
    if cell is None or cell == "":
        return ""
    if isinstance(cell, str) and cell.strip().upper() in ("#N/A", "N/A", "-"):
        return ""
    s = _cell_str(cell)
    if s.upper() in ("#N/A", "N/A", "-"):
        return ""
    return s.strip()


def _mnp_row_to_payload(row: list[Any], col_to_field: dict[int, str]) -> dict[str, Any]:
    raw: dict[str, Any] = {}
    for idx, cell in enumerate(row):
        field = col_to_field.get(idx)
        if not field:
            continue
        if field == "work_date":
            raw[field] = _parse_work_date(cell) or _cell_str(cell)
        elif field == "worker_hours":
            if cell is None or cell == "":
                raw[field] = None
            else:
                try:
                    raw[field] = float(cell)
                except (TypeError, ValueError):
                    raw[field] = _cell_str(cell)
        elif field == "civil_id":
            raw[field] = _parse_civil_id_cell(cell)
        else:
            raw[field] = _cell_str(cell)
    return raw


def _combine_activity(m: dict[str, Any]) -> str:
    parts = []
    for k in ("main_activity", "sub_act", "sub_sub_act", "note"):
        v = (m.get(k) or "").strip()
        if v:
            parts.append(v)
    return " / ".join(parts)


def _map_shift(raw_shift: str) -> str:
    s = (raw_shift or "").strip()
    if not s:
        return "Day"
    sl = s.lower()
    if sl == "night":
        return "Night"
    if sl == "idle":
        return "Day"
    if "leave" in sl:
        return "Request for leave"
    return "Day"


def _find_worker_by_company_name(db, company_raw: str, name_raw: str) -> dict | None:
    company_raw = (company_raw or "").strip()
    name_raw = (name_raw or "").strip()
    if not company_raw or not name_raw:
        return None
    return db.worker_details.find_one(
        {
            "is_active": True,
            "company_name": {"$regex": "^" + re.escape(company_raw) + "$", "$options": "i"},
            "name": {"$regex": "^" + re.escape(name_raw) + "$", "$options": "i"},
        }
    )


def _resolve_location_canonical(db, raw: str) -> str | None:
    raw = normalize_work_location_name((raw or "").strip())
    if not raw:
        return None
    doc = db.locations.find_one(
        {"name": {"$regex": "^" + re.escape(raw) + "$", "$options": "i"}, "is_active": True},
        {"name": 1},
    )
    return (doc["name"] if doc else None) or None


def _resolve_company_canonical(db, raw: str) -> str | None:
    raw = (raw or "").strip()
    if not raw:
        return None
    doc = db.companies.find_one(
        {"name": {"$regex": "^" + re.escape(raw) + "$", "$options": "i"}, "is_active": True},
        {"name": 1},
    )
    return (doc["name"] if doc else None) or None


def _infer_worker_category_safe(designation: str) -> str:
    try:
        from app import _infer_worker_category

        return _infer_worker_category(designation)
    except Exception:
        return "Direct"


def _create_worker_if_missing(
    db,
    civil_id: str,
    company_raw: str,
    worker_name: str,
    designation: str,
) -> tuple[dict | None, str | None]:
    company_name = _resolve_company_canonical(db, company_raw)
    if not company_name:
        return None, f"company not in master list: {company_raw!r}"
    if not (designation or "").strip():
        return None, "designation required to create worker"
    try:
        db.worker_details.insert_one(
            {
                "civil_id": civil_id,
                "name": (worker_name or "").strip(),
                "designation": designation.strip(),
                "category": _infer_worker_category_safe(designation),
                "company_name": company_name,
                "email": None,
                "is_active": True,
            }
        )
    except DuplicateKeyError:
        pass
    w = db.worker_details.find_one({"civil_id": civil_id, "is_active": True})
    if not w:
        return None, "could not create or load worker"
    return w, None


def _resolve_worker_for_mnp(
    db,
    m: dict[str, Any],
    *,
    create_missing_workers: bool,
    placeholder_for_no_civil_id: bool = False,
    dry_run: bool = False,
) -> tuple[dict | None, str | None, bool]:
    """
    Returns (worker_doc, error, created_new).

    When placeholder_for_no_civil_id=True and a row has no Civil ID, returns a
    synthetic placeholder worker dict with civil_id="0000" so the entry can still
    be stored.  Downstream code checks for the "_is_placeholder" key and bypasses
    strict DB validation (company / worker_details lookup).
    """
    civil_raw = _parse_civil_id_cell(m.get("civil_id"))
    company_raw = (m.get("company_name") or "").strip()
    name_raw = (m.get("worker_name") or "").strip()
    designation = (m.get("designation") or "").strip()

    if civil_raw:
        w = db.worker_details.find_one({"civil_id": civil_raw, "is_active": True})
        if w:
            return w, None, False
        try:
            from app import _resolve_worker_civil_id

            resolved, _err = _resolve_worker_civil_id(civil_raw, db)
            if resolved:
                w = db.worker_details.find_one({"civil_id": resolved, "is_active": True})
                if w:
                    return w, None, False
        except Exception:
            pass
        if create_missing_workers:
            if dry_run:
                return None, "dry-run: would create worker (civil_id not in DB yet)", False
            w, err = _create_worker_if_missing(db, civil_raw, company_raw, name_raw, designation)
            if w:
                return w, None, True
            return None, err or "could not create worker", False
        return None, "worker not found for civil_id", False

    # No civil_id in Excel row
    w = _find_worker_by_company_name(db, company_raw, name_raw)
    if w:
        return w, None, False

    if placeholder_for_no_civil_id:
        # Return a synthetic placeholder — entry stored with civil_id = "0000"
        return {
            "civil_id": "0000",
            "name": name_raw,
            "company_name": company_raw,
            "designation": designation,
            "is_active": True,
            "_is_placeholder": True,
        }, None, False

    return None, "worker not found (company + name surname); add Civil ID or enable create missing workers", False


def _normalize_eq_label(s: str) -> str:
    s = re.sub(r"\s+", " ", (s or "").strip().upper())
    s = re.sub(r"\(RENTAL\)", "", s, flags=re.I)
    return s.strip(" -")


def _extract_plate_hint(name_surname: str) -> str | None:
    s = (name_surname or "").strip()
    if not s:
        return None
    m = re.search(
        r"\b(\d{3,6}[A-Z]{2,4}|[A-Z]{1,3}\d{3,8}[A-Z]{0,4})\s*$",
        s,
        re.I,
    )
    if m:
        return m.group(1).upper()
    m2 = re.search(r"-([A-Z0-9]{2,10})\s*$", s, re.I)
    if m2:
        return m2.group(1).upper()
    return None


def _token_jaccard(a: str, b: str) -> float:
    ta = set(re.findall(r"[A-Z0-9]+", (a or "").upper()))
    tb = set(re.findall(r"[A-Z0-9]+", (b or "").upper()))
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)


def _find_equipment_for_mnp_import(db, name_surname: str, area_raw: str) -> tuple[dict | None, str | None]:
    loc = _resolve_location_canonical(db, area_raw or "")
    if not loc:
        return None, f"unknown location: {(area_raw or '').strip()!r}"

    plate_hint = _extract_plate_hint(name_surname)
    q_base = {
        "is_active": True,
        "location": {"$regex": "^" + re.escape(loc) + "$", "$options": "i"},
    }
    if plate_hint:
        eq = db.equipment_details.find_one(
            {
                **q_base,
                "plate_number": {"$regex": "^" + re.escape(plate_hint) + "$", "$options": "i"},
            }
        )
        if eq:
            return eq, None

    label = _normalize_eq_label(name_surname)
    cands = list(db.equipment_details.find(q_base))
    if not cands:
        return None, f"no equipment registered at {loc}"

    best, best_score = None, 0.0
    for eq in cands:
        en = _normalize_eq_label(eq.get("name") or "")
        pn = (eq.get("plate_number") or "").strip().upper()
        score = 0.0
        if pn and re.search(re.escape(pn), name_surname.upper()):
            score = 0.92
        elif en and (en in label or label in en):
            score = 0.88
        else:
            score = _token_jaccard(label, en)
            if _token_jaccard(label, _normalize_eq_label(eq.get("equipment_type") or "")) > score:
                score = _token_jaccard(label, _normalize_eq_label(eq.get("equipment_type") or ""))
        if score > best_score:
            best_score, best = score, eq
    if best and best_score >= 0.28:
        return best, None
    return None, f"no equipment match for {name_surname!r} at {loc}"


def _time_range_for_hours(hours: float) -> tuple[str, str]:
    h = max(0.0, min(float(hours or 0), 24.0))
    start_h, start_m = 6, 0
    total_min = int(round(h * 60))
    end_min = start_h * 60 + start_m + total_min
    eh = min(end_min // 60, 23)
    em = end_min % 60 if end_min // 60 <= 23 else 59
    return f"{start_h:02d}:{start_m:02d}", f"{eh:02d}:{em:02d}"


def _parse_mnp_sheet(ws) -> tuple[list[dict[str, Any]], list[str]]:
    errors: list[str] = []
    hr, header_row = _find_header_row(ws)
    if not header_row:
        return [], ["could not find header row (expected Date, Company, Name Surname, …)"]

    col_to_field = _build_mnp_header_map(header_row)
    required = {"work_date", "company_name", "worker_name", "man_equ"}
    missing = required - set(col_to_field.values())
    if missing:
        errors.append("missing columns: " + ", ".join(sorted(missing)))

    data_start = (hr or 1) + 1
    rows_out: list[dict[str, Any]] = []
    for ridx, row in enumerate(
        ws.iter_rows(min_row=data_start, values_only=True),
        start=data_start,
    ):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        m = _mnp_row_to_payload(list(row), col_to_field)
        if not m.get("worker_name") and not m.get("company_name"):
            continue
        m["_row_number"] = ridx
        rows_out.append(m)
    return rows_out, errors


def _build_placeholder_work_entry(payload: dict[str, Any]) -> dict[str, Any]:
    """
    Build a work_entry document directly for rows with no Civil ID (civil_id = "0000").
    Bypasses strict company / location / worker_details DB validation so that historical
    entries from external companies can still be stored and assigned a real Civil ID later.
    """
    from datetime import datetime as _dt

    worker_shift = payload.get("worker_shift") or "Day"
    return {
        "civil_id": "0000",
        "company_name": payload.get("company_name") or "",
        "worker_name": payload.get("worker_name") or "",
        "work_date": payload.get("work_date") or "",
        "location": payload.get("location") or None,
        "incharge": None,
        "permit_issuer": None,
        "worker_hours": payload.get("worker_hours"),
        "worker_time_from": None,
        "worker_time_to": None,
        "worker_shift": worker_shift,
        "leave_reason": payload.get("leave_reason") if worker_shift == "Request for leave" else None,
        "hours": None,
        "attendance_status": None,
        "today_activity": payload.get("today_activity") or "",
        "item_tag": payload.get("item_tag"),
        "status": "submitted",
        "created_at": _dt.utcnow().isoformat(),
    }


def mnp_row_to_work_entry_payload(
    db,
    m: dict[str, Any],
    *,
    create_missing_workers: bool = False,
    placeholder_for_no_civil_id: bool = False,
    dry_run: bool = False,
) -> tuple[dict[str, Any] | None, str | None, bool]:
    """Turn one MNP row dict into payload for validate_and_build_document. Third value: new worker created.

    When the resolved worker has _is_placeholder=True the returned payload carries
    _bypass_validation=True so the caller inserts via _build_placeholder_work_entry
    instead of validate_and_build_document.
    """
    man_equ = (m.get("man_equ") or "").strip().lower()
    if man_equ == "equipment":
        return None, "skip_equipment", False
    allowed = (
        "manpower",
        "man",
        "sub contactor",
        "subcontractor",
        "sub contractor",
        "driver",
    )
    if man_equ and man_equ not in allowed:
        return None, f"skip_man_equ:{man_equ or '?'}", False

    worker, werr, created = _resolve_worker_for_mnp(
        db,
        m,
        create_missing_workers=create_missing_workers,
        placeholder_for_no_civil_id=placeholder_for_no_civil_id,
        dry_run=dry_run,
    )
    if not worker:
        return None, werr or "worker not found", False

    is_placeholder = bool(worker.get("_is_placeholder"))
    civil_id = (worker.get("civil_id") or "").strip()
    company_name = (worker.get("company_name") or "").strip()
    worker_name = (worker.get("name") or "").strip()

    work_date = _parse_work_date(m.get("work_date"))
    if not work_date:
        return None, "invalid work_date", False

    shift_raw = m.get("worker_shift") or ""
    worker_shift = _map_shift(shift_raw)
    leave_reason = ""
    if worker_shift == "Request for leave":
        leave_reason = _combine_activity(m) or (m.get("note") or "Leave")

    wh = m.get("worker_hours")
    try:
        worker_hours_val = float(wh) if wh is not None and wh != "" else None
    except (TypeError, ValueError):
        worker_hours_val = None

    if worker_shift == "Request for leave":
        worker_hours_val = 0.0
    elif worker_hours_val is None:
        worker_hours_val = 0.0 if (shift_raw or "").strip().lower() == "idle" else None

    if worker_shift != "Request for leave" and (worker_hours_val is None or worker_hours_val < 0):
        return None, "invalid manhours", False

    if worker_hours_val is not None and worker_hours_val > 12:
        return None, "manhours must be <= 12", False

    today_activity = _combine_activity(m)
    if not today_activity and (shift_raw or "").strip().lower() == "idle":
        today_activity = "Idle"
    if worker_shift != "Request for leave" and not (today_activity or "").strip():
        return None, "missing activity (main/sub/note)", False

    loc_raw = normalize_work_location_name((m.get("location") or "").strip())

    if is_placeholder:
        # Skip DB validation — use canonicalized location; entry stored with civil_id="0000"
        payload = {
            "civil_id": "0000",
            "company_name": company_name,
            "worker_name": worker_name,
            "work_date": work_date,
            "location": loc_raw or "",
            "incharge": "",
            "permit_issuer": "",
            "today_activity": today_activity,
            "worker_time_from": "",
            "worker_time_to": "",
            "worker_hours": worker_hours_val,
            "worker_shift": worker_shift,
            "leave_reason": leave_reason,
            "item_tag": (m.get("item_tag") or "").strip() or None,
            "_bypass_validation": True,
        }
        return payload, None, created

    location_resolved = _resolve_location_canonical(db, loc_raw) if loc_raw else ""
    if loc_raw and not location_resolved:
        return None, f"location not in master data: {loc_raw!r}", False

    item_tag = (m.get("item_tag") or "").strip() or None

    payload = {
        "civil_id": civil_id,
        "company_name": company_name,
        "worker_name": worker_name,
        "work_date": work_date,
        "location": location_resolved or "",
        "incharge": "",
        "permit_issuer": "",
        "today_activity": today_activity,
        "worker_time_from": "",
        "worker_time_to": "",
        "worker_hours": worker_hours_val,
        "worker_shift": worker_shift,
        "leave_reason": leave_reason,
        "item_tag": item_tag,
    }
    return payload, None, created


def _split_mnp_rows(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    manpower_rows: list[dict[str, Any]] = []
    equipment_rows: list[dict[str, Any]] = []
    for raw in rows:
        m = dict(raw)
        me = (m.get("man_equ") or "").strip().lower()
        if me == "equipment":
            equipment_rows.append(m)
        elif me in ("manpower", "man", "sub contactor", "subcontractor", "sub contractor", "driver"):
            manpower_rows.append(m)
    return manpower_rows, equipment_rows


def _merge_equipment_groups(
    db, equipment_rows: list[dict[str, Any]]
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    groups: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    errors: list[dict[str, Any]] = []
    for m in equipment_rows:
        rnum = m.get("_row_number")
        wd = _parse_work_date(m.get("work_date"))
        if not wd:
            errors.append({"row": rnum, "error": "invalid work_date", "kind": "equipment"})
            continue
        key = (wd, _norm(m.get("worker_name") or ""), _norm(m.get("location") or ""))
        groups[key].append(m)

    merged_out: list[dict[str, Any]] = []
    for _key, group in groups.items():
        first = group[0]
        rnums = [x.get("_row_number") for x in group]
        eq, err = _find_equipment_for_mnp_import(db, first.get("worker_name") or "", first.get("location") or "")
        if err or not eq:
            for rn in rnums:
                errors.append({"row": rn, "error": err or "equipment not found", "kind": "equipment"})
            continue
        total_h = 0.0
        acts: list[str] = []
        for x in group:
            wh = x.get("worker_hours")
            try:
                total_h += float(wh) if wh is not None and wh != "" else 0.0
            except (TypeError, ValueError):
                pass
            a = _combine_activity(x)
            if a:
                acts.append(a)
        if total_h <= 0:
            total_h = 8.0
        loc = _resolve_location_canonical(db, (first.get("location") or "").strip())
        if not loc:
            for rn in rnums:
                errors.append({"row": rn, "error": "unknown location", "kind": "equipment"})
            continue
        merged_out.append(
            {
                "_equipment_doc": eq,
                "work_date": _parse_work_date(first.get("work_date")),
                "hours": min(total_h, 24.0),
                "activity": " / ".join(acts) if acts else "Equipment work",
                "location_canonical": loc,
                "filled_by": (first.get("filled_by") or "").strip(),
                "_source_rows": rnums,
            }
        )
    return merged_out, errors


def _insert_equipment_mnp_batch(
    db,
    merged_payloads: list[dict[str, Any]],
    *,
    operator_civil_id: str,
    dry_run: bool,
) -> tuple[int, list[dict[str, Any]]]:
    inserted = 0
    errs: list[dict[str, Any]] = []
    op_civil = (operator_civil_id or "").strip() or DEVELOPER_MASTER_CIVIL_ID
    if not db.worker_details.find_one({"civil_id": op_civil, "is_active": True}) and op_civil != DEVELOPER_MASTER_CIVIL_ID:
        for mp in merged_payloads:
            for rn in mp.get("_source_rows") or []:
                errs.append({"row": rn, "error": "equipment_operator_civil_id not found", "kind": "equipment"})
        return 0, errs

    for mp in merged_payloads:
        eq = mp["_equipment_doc"]
        eid = str(eq["_id"])
        wd = mp["work_date"]
        existing = db.equipment_entries.find_one(
            {"equipment_id": eid, "work_date": wd, "status": {"$nin": ["cancelled"]}}
        )
        if existing:
            for rn in mp.get("_source_rows") or []:
                errs.append(
                    {
                        "row": rn,
                        "error": "duplicate equipment entry for date (already in DB)",
                        "kind": "equipment",
                    }
                )
            continue
        tf, tt = _time_range_for_hours(float(mp["hours"]))
        doc = {
            "civil_id": op_civil,
            "company_name": "",
            "operator_name": (mp.get("filled_by") or "").strip() or "MNP import",
            "equipment_id": eid,
            "equipment_name": eq.get("name", ""),
            "plate_number": eq.get("plate_number", ""),
            "equipment_type": eq.get("equipment_type", ""),
            "equipment_location": eq.get("location", ""),
            "ownership": eq.get("ownership", ""),
            "work_date": wd,
            "location": mp["location_canonical"],
            "equipment_status": "Working",
            "time_from": tf,
            "time_to": tt,
            "hours": round(float(mp["hours"]), 2),
            "activity": mp["activity"],
            "supply_rate": eq.get("supply_rate") or {},
            "contract_rate": eq.get("contract_rate") or {},
            "rental_amount": None,
            "status": "submitted",
            "created_at": datetime.utcnow().isoformat(),
        }
        if dry_run:
            inserted += 1
        else:
            db.equipment_entries.insert_one(doc)
            inserted += 1
    return inserted, errs


def _run_one_sheet_mnp_import(
    db,
    rows: list[dict[str, Any]],
    *,
    sheet_label: str,
    create_missing_workers: bool,
    placeholder_for_no_civil_id: bool,
    import_equipment: bool,
    equipment_operator_civil_id: str | None,
    dry_run: bool,
    progress: bool,
) -> dict[str, Any]:
    manpower_rows, equipment_rows = _split_mnp_rows(rows)
    row_errors: list[dict[str, Any]] = []
    inserted_ids: list[str] = []
    workers_created = 0
    placeholder_inserted = 0
    n_man = len(manpower_rows)
    step = progress_step(max(n_man, 1))

    if progress and n_man:
        progress_message(f"    {n_man} manpower rows — {'dry-run' if dry_run else 'inserting'}…")

    for j, raw in enumerate(manpower_rows, 1):
        m = dict(raw)
        rnum = m.pop("_row_number", None)
        payload, err, created = mnp_row_to_work_entry_payload(
            db,
            m,
            create_missing_workers=create_missing_workers,
            placeholder_for_no_civil_id=placeholder_for_no_civil_id,
            dry_run=dry_run,
        )
        if created:
            workers_created += 1
        if err and err.startswith("skip_"):
            if progress and (j % step == 0 or j == n_man):
                progress_line(j, n_man, (sheet_label.strip()[:22] + " ") if sheet_label else "")
            continue
        if err:
            row_errors.append({"sheet": sheet_label, "row": rnum, "error": err})
            if progress and (j % step == 0 or j == n_man):
                progress_line(j, n_man, (sheet_label.strip()[:22] + " ") if sheet_label else "")
            continue

        # Placeholder entries (no Civil ID) bypass strict DB validation
        if payload.get("_bypass_validation"):
            clean_payload = {k: v for k, v in payload.items() if not k.startswith("_")}
            doc = _build_placeholder_work_entry(clean_payload)
            if dry_run:
                inserted_ids.append("(dry-run-placeholder)")
            else:
                db.work_entries.insert_one(doc)
                inserted_ids.append("(placeholder)")
            placeholder_inserted += 1
            if progress and (j % step == 0 or j == n_man):
                progress_line(j, n_man, (sheet_label.strip()[:22] + " ") if sheet_label else "")
            continue

        doc, verr = validate_and_build_document(db, payload)
        if verr:
            row_errors.append(
                {"sheet": sheet_label, "row": rnum, "error": verr, "civil_id": payload.get("civil_id")}
            )
            if progress and (j % step == 0 or j == n_man):
                progress_line(j, n_man, (sheet_label.strip()[:22] + " ") if sheet_label else "")
            continue
        if dry_run:
            inserted_ids.append("(dry-run)")
        else:
            ins = db.work_entries.insert_one(doc)
            inserted_ids.append(str(ins.inserted_id))
        if progress and (j % step == 0 or j == n_man):
            progress_line(j, n_man, (sheet_label.strip()[:22] + " ") if sheet_label else "")
    if progress and n_man:
        progress_done()

    equipment_inserted = 0
    if import_equipment and equipment_rows:
        merged, merge_errs = _merge_equipment_groups(db, equipment_rows)
        for e in merge_errs:
            row_errors.append({**e, "sheet": sheet_label})
        eq_ins, eq_batch_errs = _insert_equipment_mnp_batch(
            db,
            merged,
            operator_civil_id=equipment_operator_civil_id or "",
            dry_run=dry_run,
        )
        equipment_inserted = eq_ins
        for e in eq_batch_errs:
            row_errors.append({**e, "sheet": sheet_label})

    return {
        "inserted_ids": inserted_ids,
        "row_errors": row_errors,
        "workers_created": workers_created,
        "placeholder_inserted": placeholder_inserted,
        "equipment_inserted": equipment_inserted,
        "manpower_row_count": n_man,
        "equipment_row_count": len(equipment_rows),
    }


def import_mnp_tracking_excel(
    db,
    content: bytes,
    *,
    sheet_name: str | None = None,
    sheet_index: int | None = None,
    all_sheets: bool = False,
    dry_run: bool = False,
    progress: bool = False,
    create_missing_workers: bool = False,
    placeholder_for_no_civil_id: bool = False,
    import_equipment: bool = True,
    equipment_operator_civil_id: str | None = None,
) -> dict[str, Any]:
    """
    Import MNP tracking workbook.

    - sheet_name: exact sheet name (e.g. "January 2026", "March 2026")
    - sheet_index: 0-based index if sheet_name not set
    - all_sheets: every sheet except Cumulative / Tasks / Manpower / monthly roll-ups in MNP_SKIP_ALL_SHEETS
    - create_missing_workers: insert worker_details when civil ID present but not in DB (company must exist in master data)
    - placeholder_for_no_civil_id: store manpower rows with no Civil ID using civil_id="0000" (can be updated later)
    - import_equipment: insert equipment_entries for Man/Equ = Equipment (merged duplicate same day)
    - equipment_operator_civil_id: worker civil ID for equipment entries (defaults to developer master)
    """
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        if all_sheets:
            total_inserted = 0
            total_equipment_inserted = 0
            total_workers_created = 0
            total_errors: list[dict[str, Any]] = []
            parse_errs: list[str] = []
            eligible = [n for n in wb.sheetnames if n not in MNP_SKIP_ALL_SHEETS]
            if progress:
                progress_message(
                    f"MNP import (all sheets): {len(eligible)} sheets (skipping roll-ups + Cumulative + Tasks/Manpower)"
                )
            for si, name in enumerate(eligible, start=1):
                ws = wb[name]
                if progress:
                    progress_message(f"[{si}/{len(eligible)}] {name.strip()} — reading rows…")
                rows, perr = _parse_mnp_sheet(ws)
                parse_errs.extend([f"[{name}] {e}" for e in perr])
                nrows = len(rows)
                if progress and nrows:
                    progress_message(f"    {nrows} data rows — validating / {'dry-run' if dry_run else 'inserting'}…")
                out = _run_one_sheet_mnp_import(
                    db,
                    rows,
                    sheet_label=name,
                    create_missing_workers=create_missing_workers,
                    placeholder_for_no_civil_id=placeholder_for_no_civil_id,
                    import_equipment=import_equipment,
                    equipment_operator_civil_id=equipment_operator_civil_id,
                    dry_run=dry_run,
                    progress=progress,
                )
                total_inserted += len(out["inserted_ids"])
                total_equipment_inserted += out["equipment_inserted"]
                total_workers_created += out["workers_created"]
                total_errors.extend(out["row_errors"])
            return {
                "ok": True,
                "format": "mnp_tracking",
                "mode": "all_sheets",
                "parse_errors": parse_errs,
                "inserted": total_inserted if not dry_run else 0,
                "would_insert": total_inserted,
                "equipment_inserted": total_equipment_inserted if not dry_run else 0,
                "would_equipment_insert": total_equipment_inserted if dry_run else total_equipment_inserted,
                "workers_created": total_workers_created if not dry_run else 0,
                "row_errors": total_errors,
            }

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return {
                    "ok": False,
                    "error": f"sheet not found: {sheet_name!r}; available: {wb.sheetnames}",
                }
            ws = wb[sheet_name]
        elif sheet_index is not None:
            if sheet_index >= len(wb.sheetnames):
                return {"ok": False, "error": f"sheet index {sheet_index} out of range"}
            ws = wb[wb.sheetnames[sheet_index]]
        else:
            ws = wb[_default_mnp_ws_name(wb)]

        sheet_title = getattr(ws, "title", "?")
        rows, parse_errors = _parse_mnp_sheet(ws)
        if progress:
            progress_message(
                f"Sheet «{sheet_title.strip()}»: {len(rows)} data rows — "
                f"{'dry-run (no writes)' if dry_run else 'inserting to database'}…"
            )

        out = _run_one_sheet_mnp_import(
            db,
            rows,
            sheet_label=sheet_title,
            create_missing_workers=create_missing_workers,
            placeholder_for_no_civil_id=placeholder_for_no_civil_id,
            import_equipment=import_equipment,
            equipment_operator_civil_id=equipment_operator_civil_id,
            dry_run=dry_run,
            progress=progress,
        )
        inserted_ids = out["inserted_ids"]
        row_errors = out["row_errors"]

        return {
            "ok": True,
            "format": "mnp_tracking",
            "parse_errors": parse_errors,
            "rows_read": len(rows),
            "inserted": len(inserted_ids) if not dry_run else 0,
            "would_insert": len(inserted_ids),
            "equipment_inserted": out["equipment_inserted"] if not dry_run else 0,
            "would_equipment_insert": out["equipment_inserted"],
            "workers_created": out["workers_created"] if not dry_run else 0,
            "placeholder_inserted": out["placeholder_inserted"] if not dry_run else 0,
            "row_errors": row_errors,
            "inserted_ids": inserted_ids if not dry_run else [],
            "manpower_rows": out["manpower_row_count"],
            "equipment_rows": out["equipment_row_count"],
        }
    finally:
        wb.close()

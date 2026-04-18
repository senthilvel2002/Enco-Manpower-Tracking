import os
from io import BytesIO
from datetime import date, datetime, timedelta
import smtplib
from email.message import EmailMessage

from bson import ObjectId
from flask import Flask, jsonify, request, send_file, send_from_directory
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pymongo.errors import DuplicateKeyError, PyMongoError

from config import Config
from db import DatabaseUnavailable, get_database, ping_database

import re
import base64

try:
    from dotenv import load_dotenv

    load_dotenv()
except Exception:
    pass

app = Flask(__name__)

_cors_origins = [
    "http://localhost:3000",
    "http://127.0.0.1:3000",
    "http://192.168.8.188:3000",
]
_extra_cors = (os.getenv("CORS_ORIGINS") or "").strip()
if _extra_cors:
    for part in _extra_cors.split(","):
        p = part.strip()
        if p and p not in _cors_origins:
            _cors_origins.append(p)

CORS(
    app,
    resources={r"/api/*": {"origins": _cors_origins}},
)
DEVELOPER_MASTER_CIVIL_ID = "141228922"
SITE_INCHARGE_ALLOWED_CIVIL_IDS = {
    "116858297",
    "116915582",
    "122045765",
    "141228922",
}
MANAGEMENT_ALLOWED_CIVIL_IDS = {
    "122045765",
    "127387473",
    "141228922",
    "121598538",
    "121668462",
    "140696356",
    "141246094",
}
SITE_INCHARGE_LOGIN_PASSWORD = "0000"
MANAGEMENT_LOGIN_PASSWORD = "25017"
INDIRECT_LABOUR_DESIGNATIONS = {
    "project manager",
    "site manager",
    "hse manager",
    "hse engineer",
    "hse officer",
    "lead mechanical engineer",
    "mechanical engineer",
    "quality manager",
    "lifting engineer",
    "material manager",
    "mechanical supervisor",
    "mechanical superintendent",
    "hse coordinator",
    "junior hse officer",
    "scheduler/planning engineer",
    "welding supervisor",
    "electrical superintendant",
    "electrical supervisor",
    "mechanic supervisor",
    "qc mechanical",
    "qc mechanical engineer",
    "store-keeper",
    "draftsman engineer",
    "permit receiver",
    "surveyor",
    "admin manager",
    "admin coordinator",
    "admin assistant",
    "junior pro",
    "qs engineer",
    "process engineer",
    "it engineer",
    "driver",
    "scaffolding supervisor",
    "qc mechanical lead",
    "painting inspector",
    "scheduling engineer",
    "qc electrical engineer",
    "hr manager",
    "hr officer",
    "site coordinator / timekeeper",
    "document controller",
    "qc document controller",
    "welding inspector",
    "purchasing manager",
    "accountant",
    "logistics supervisor",
    "forklift operator",
    "crane operator",
    "camp boss/coordinator",
    "cnc operator assistant",
    "electrical engineer",
}


@app.errorhandler(DatabaseUnavailable)
def _handle_db_unavailable(exc):
    return jsonify({"ok": False, "error": "database unavailable", "details": str(exc)}), 503


@app.errorhandler(PyMongoError)
def _handle_pymongo_error(exc):
    return jsonify({"ok": False, "error": "database unavailable", "details": str(exc)}), 503


def _serialize_doc(document):
    if not document:
        return document
    document["_id"] = str(document["_id"])
    return document


def _coalesce_entry_hours(entry):
    """Worker daily submit stores time in worker_hours; official hours may be null until approval."""
    if entry.get("hours") is not None:
        return
    wh = entry.get("worker_hours")
    if wh is not None:
        entry["hours"] = wh


# Stored on worker_details.profile_picture as a data URL (data:image/*;base64,...)
PROFILE_PICTURE_MAX_BYTES = 600 * 1024


def _validate_profile_picture_data_url(data_url: str):
    """Returns (ok, error_message_or_None, normalized_data_url_or_None)."""
    if not data_url or not isinstance(data_url, str):
        return False, "profile_picture must be a non-empty string or null", None
    data_url = data_url.strip()
    if not data_url.startswith("data:image/"):
        return False, "profile_picture must be a data URL (image/*)", None
    if "," not in data_url:
        return False, "invalid data URL", None
    _header, b64 = data_url.split(",", 1)
    try:
        raw = base64.b64decode(b64, validate=True)
    except Exception:
        return False, "invalid base64 in profile_picture", None
    if len(raw) > PROFILE_PICTURE_MAX_BYTES:
        return False, "image too large (max 600KB)", None
    return True, None, data_url


def _infer_worker_category(designation: str):
    key = (designation or "").strip().lower()
    return "Indirect" if key in INDIRECT_LABOUR_DESIGNATIONS else "Direct"


def _resolve_worker_civil_id(raw: str, db):
    """
    Login helper: accept full Civil ID, or exactly 4 digits matching the end of one active worker.
    Returns (full_civil_id_or_None, error_key_or_None) where error_key is
    'empty' | 'not_found' | 'ambiguous'.
    """
    s = (raw or "").strip()
    if not s:
        return None, "empty"
    if db.worker_details.find_one({"civil_id": s, "is_active": True}, {"_id": 1}):
        return s, None
    if re.fullmatch(r"\d{4}", s):
        rx = re.compile(re.escape(s) + r"$")
        candidates = list(
            db.worker_details.find({"civil_id": rx, "is_active": True}, {"civil_id": 1}).limit(25)
        )
        if len(candidates) == 0:
            return None, "not_found"
        if len(candidates) > 1:
            return None, "ambiguous"
        return candidates[0].get("civil_id"), None
    return None, "not_found"


def _resolve_civil_id_in_set(raw: str, allowed: set):
    """
    Match full Civil ID or last 4 digits uniquely against a fixed allow-list (management / site incharge).
    """
    s = (raw or "").strip()
    if not s:
        return None, "empty"
    if s in allowed:
        return s, None
    if re.fullmatch(r"\d{4}", s):
        matches = [cid for cid in allowed if (cid or "").endswith(s)]
        if len(matches) == 1:
            return matches[0], None
        if len(matches) == 0:
            return None, "not_found"
        return None, "ambiguous"
    return None, "not_found"


def _profile_picture_map(db, civil_ids):
    """Map civil_id -> profile_picture data URL (or empty string if none)."""
    uniq = list({cid for cid in civil_ids if cid})
    if not uniq:
        return {}
    out = {}
    for w in db.worker_details.find(
        {"civil_id": {"$in": uniq}, "is_active": True},
        {"civil_id": 1, "profile_picture": 1},
    ):
        cid = w.get("civil_id")
        if cid:
            out[cid] = (w.get("profile_picture") or "").strip()
    return out


def _enrich_entries_profile_pictures(db, entries):
    """Attach profile_picture from worker_details to each entry (by civil_id)."""
    civil_ids = [e.get("civil_id") for e in entries if e.get("civil_id")]
    pic_map = _profile_picture_map(db, civil_ids)
    for e in entries:
        cid = e.get("civil_id")
        e["profile_picture"] = pic_map.get(cid, "") if cid else ""


def _worker_profile_map(db, civil_ids):
    """Map civil_id -> display fields from worker_details for tooltips."""
    uniq = list({cid for cid in civil_ids if cid})
    if not uniq:
        return {}
    out = {}
    for w in db.worker_details.find(
        {"civil_id": {"$in": uniq}, "is_active": True},
        {
            "civil_id": 1,
            "name": 1,
            "designation": 1,
            "category": 1,
            "company_name": 1,
            "email": 1,
        },
    ):
        cid = w.get("civil_id")
        if not cid:
            continue
        designation = w.get("designation") or ""
        category = w.get("category") or _infer_worker_category(designation)
        out[cid] = {
            "name": (w.get("name") or "").strip(),
            "civil_id": cid,
            "company_name": (w.get("company_name") or "").strip(),
            "designation": designation,
            "category": category,
            "email": (w.get("email") or "").strip(),
        }
    return out


def _enrich_entries_worker_profiles(db, entries):
    """Attach worker_profile dict for UI tooltips (name, company, email, …)."""
    civil_ids = [e.get("civil_id") for e in entries if e.get("civil_id")]
    pmap = _worker_profile_map(db, civil_ids)
    for e in entries:
        cid = e.get("civil_id")
        e["worker_profile"] = pmap.get(cid) if cid else None


def _ensure_management_password(payload):
    password = (payload.get("management_password") or "").strip()
    if not password:
        return False, jsonify({"ok": False, "error": "management_password is required"})
    if password != MANAGEMENT_LOGIN_PASSWORD:
        return False, jsonify({"ok": False, "error": "invalid management password"})
    return True, None


def _sync_worker_categories(db):
    for w in db.worker_details.find(
        {"is_active": True},
        {"civil_id": 1, "designation": 1, "category": 1},
    ):
        cid = w.get("civil_id")
        if not cid:
            continue
        expected = _infer_worker_category(w.get("designation") or "")
        if (w.get("category") or "") != expected:
            db.worker_details.update_one({"civil_id": cid}, {"$set": {"category": expected}})


def _normalize_period_filter(period_raw):
    """Map UI labels (daily, weekly, …) to internal keys (day, week, …)."""
    p = (period_raw or "all").strip().lower()
    return {
        "daily": "day",
        "weekly": "week",
        "monthly": "month",
        "quarterly": "quarter",
        "yearly": "year",
    }.get(p, p)


def _build_entries_query(args):
    query = {}
    location = (args.get("location") or "").strip()
    incharge = (args.get("incharge") or "").strip()
    approval_status = (args.get("approval_status") or "").strip().lower()
    period = _normalize_period_filter(args.get("period") or "all")
    anchor_date_raw = (args.get("anchor_date") or "").strip()

    if location and location.lower() != "all":
        query["location"] = location
    if incharge and incharge.lower() != "all":
        query["incharge"] = incharge
    if approval_status and approval_status != "all":
        if approval_status == "pending":
            query["$or"] = [
                {"approval_status": {"$exists": False}},
                {"approval_status": None},
                {"approval_status": ""},
                {"approval_status": "pending"},
            ]
        elif approval_status in {"approved", "rejected"}:
            query["approval_status"] = approval_status

    if period == "all":
        return query

    try:
        anchor = datetime.strptime(anchor_date_raw, "%Y-%m-%d").date() if anchor_date_raw else date.today()
    except ValueError:
        anchor = date.today()

    start = anchor
    end = anchor

    if period == "day":
        start = anchor
        end = anchor
    elif period == "week":
        start = anchor - timedelta(days=anchor.weekday())
        end = start + timedelta(days=6)
    elif period == "month":
        start = anchor.replace(day=1)
        if start.month == 12:
            next_month = start.replace(year=start.year + 1, month=1, day=1)
        else:
            next_month = start.replace(month=start.month + 1, day=1)
        end = next_month - timedelta(days=1)
    elif period == "quarter":
        quarter_start_month = ((anchor.month - 1) // 3) * 3 + 1
        start = anchor.replace(month=quarter_start_month, day=1)
        if quarter_start_month == 10:
            next_quarter = start.replace(year=start.year + 1, month=1, day=1)
        else:
            next_quarter = start.replace(month=quarter_start_month + 3, day=1)
        end = next_quarter - timedelta(days=1)
    elif period == "year":
        start = anchor.replace(month=1, day=1)
        end = anchor.replace(month=12, day=31)
    else:
        return query

    query["work_date"] = {"$gte": start.isoformat(), "$lte": end.isoformat()}
    return query


def _get_anchor_date(args):
    anchor_date_raw = (args.get("anchor_date") or "").strip()
    try:
        return datetime.strptime(anchor_date_raw, "%Y-%m-%d").date() if anchor_date_raw else date.today()
    except ValueError:
        return date.today()


def _fetch_management_export_rows(db, query):
    rows = list(
        db.work_entries.find(
            query,
            {
                "_id": 0,
                "civil_id": 1,
                "worker_name": 1,
                "company_name": 1,
                "location": 1,
                "incharge": 1,
                "work_date": 1,
                "hours": 1,
                "worker_hours": 1,
                "attendance_status": 1,
                "today_activity": 1,
                "approval_status": 1,
                "rejection_reason": 1,
                "worker_shift": 1,
                "leave_reason": 1,
            },
        ).sort("work_date", -1)
    )

    civil_ids = [r.get("civil_id") for r in rows if r.get("civil_id")]
    designation_map = {}
    category_map = {}
    if civil_ids:
        for w in db.worker_details.find(
            {"civil_id": {"$in": civil_ids}, "is_active": True},
            {"civil_id": 1, "designation": 1, "category": 1},
        ):
            cid = w.get("civil_id")
            designation = w.get("designation") or ""
            category = w.get("category") or _infer_worker_category(designation)
            designation_map[cid] = designation
            category_map[cid] = category
            if not w.get("category"):
                db.worker_details.update_one({"civil_id": cid}, {"$set": {"category": category}})

    export_rows = []
    for idx, item in enumerate(rows, start=1):
        shift = item.get("worker_shift") or item.get("attendance_status") or ""
        if shift == "Request for leave":
            shift = "Leave"
        export_rows.append(
            {
                "si_no": idx,
                "approved_by": item.get("incharge", "") or "",
                "date": item.get("work_date", "") or "",
                "company": item.get("company_name", "") or "",
                "name": item.get("worker_name", "") or "",
                "designation": designation_map.get(item.get("civil_id"), ""),
                "category": category_map.get(item.get("civil_id"), "Direct"),
                "area": item.get("location", "") or "",
                "shift": shift,
                "manhours": (
                    item.get("hours")
                    if item.get("hours") is not None
                    else (item.get("worker_hours") if item.get("worker_hours") is not None else "")
                ),
                "main_activity": item.get("today_activity", "") or "",
                "approval_status": item.get("approval_status", "") or "pending",
                "rejection_reason": item.get("rejection_reason", "") or "",
                "leave_reason": item.get("leave_reason", "") or "",
            }
        )
    return export_rows


def _build_management_export_workbook(export_rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Recent Entries"

    headers = [
        "S.I No",
        "Approved By",
        "Date",
        "Company",
        "Name Surname",
        "Designation",
        "Man/Equ",
        "Category",
        "Area",
        "Reference",
        "Item Tag",
        "Shift",
        "Manhours",
        "Main Activity",
        "Approval Status",
        "Rejection Reason",
        "Leave Reason",
    ]
    sheet.append(headers)

    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for column in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=column)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    for row in export_rows:
        sheet.append(
            [
                row["si_no"],
                row["approved_by"],
                row["date"],
                row["company"],
                row["name"],
                row["designation"],
                "Manpower",
                row["category"],
                row["area"],
                "",
                "",
                row["shift"],
                row["manhours"],
                row["main_activity"],
                row["approval_status"],
                row["rejection_reason"],
                row["leave_reason"],
            ]
        )

    max_row = sheet.max_row
    max_col = len(headers)
    for row_idx in range(2, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.alignment = center_alignment
            cell.border = thin_border

    widths = [8, 18, 12, 20, 22, 16, 10, 10, 16, 12, 12, 10, 10, 22, 14, 24, 24]
    for idx, width in enumerate(widths, start=1):
        letter = chr(64 + idx) if idx <= 26 else None
        if letter:
            sheet.column_dimensions[letter].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


@app.get("/api/health")
def health():
    db_connected = False
    db_name = None
    error = None

    try:
        db = get_database()
        ping_database()
        db_connected = True
        db_name = db.name
    except Exception as exc:
        error = str(exc)

    return jsonify(
        {
            "app": "ok",
            "db_connected": db_connected,
            "db_name": db_name,
            "error": error,
        }
    )


@app.post("/api/setup/init-db")
def init_db():
    try:
        db = get_database()

        # Initialize all required collections and indexes.
        db.worker_auth.create_index([("civil_id", 1), ("company_name", 1)], unique=True)
        db.admin_auth.create_index([("civil_id", 1), ("company_name", 1)], unique=True)
        db.worker_details.create_index([("civil_id", 1)], unique=True)
        db.work_entries.create_index([("civil_id", 1), ("work_date", 1)])
        db.companies.create_index([("name", 1)], unique=True)
        db.locations.create_index([("name", 1)], unique=True)
        db.incharge.create_index([("name", 1)], unique=True)
        db.permit_issuer.create_index([("name", 1)], unique=True)

        # Daily fields belong on work_entries only; strip legacy keys from worker_details.
        db.worker_details.update_many(
            {},
            {"$unset": {"working_location": "", "permit_issuer": "", "incharge": ""}},
        )

        # Hide locations not used anymore.
        db.locations.update_one({"name": "Sohar Plant"}, {"$set": {"is_active": False}})

        # Site incharge tied to work locations (names must match locations collection).
        # `locations` is stored as an array so one incharge can serve multiple locations.
        incharge_by_locations = [
            {"name": "Mahmoud Mahmoud Moussa", "locations": ["Car parking"]},
            {"name": "Rabih El Rifai", "locations": ["Yard 2", "Yard 3"]},
            {"name": "Peter Farah", "locations": ["TSF Laydown", "Laydown Office", "Laydown"]},
        ]
        for row in incharge_by_locations:
            db.incharge.update_one(
                {"name": row["name"]},
                {
                    "$set": {
                        "name": row["name"],
                        "locations": row["locations"],
                        "is_active": True,
                    },
                    "$unset": {"location": ""},
                },
                upsert=True,
            )

        # Head Office incharge list (loaded from worker_details by civil_id).
        head_office_incharge_ids = ["10837796", "123541357", "10974059"]
        # Clean any old placeholder rows created earlier.
        db.incharge.delete_many(
            {"civil_id": {"$in": head_office_incharge_ids}, "name": {"$regex": r"^Civil ID\s+"}}
        )
        head_office_workers = list(
            db.worker_details.find(
                {"civil_id": {"$in": head_office_incharge_ids}, "is_active": True},
                {"civil_id": 1, "name": 1},
            )
        )
        head_office_name_by_id = {w.get("civil_id"): (w.get("name") or "") for w in head_office_workers}
        for cid in head_office_incharge_ids:
            name = (head_office_name_by_id.get(cid) or "").strip()
            if not name:
                continue
            # Prefer updating by civil_id so rerunning init refreshes the name.
            db.incharge.update_one(
                {"civil_id": cid},
                {
                    "$set": {
                        "civil_id": cid,
                        "name": name,
                        "locations": ["Head Office"],
                        "is_active": True,
                    },
                    "$unset": {"location": ""},
                },
                upsert=True,
            )

        # Laydown Office incharge list (loaded from worker_details by civil_id).
        laydown_office_incharge_ids = [
            "121668462",
            "121598538",
            "127387473",
            "141381507",
            "118833981",
            "122370101",
            "140841472",
        ]
        db.incharge.delete_many(
            {"civil_id": {"$in": laydown_office_incharge_ids}, "name": {"$regex": r"^Civil ID\s+"}}
        )
        laydown_office_workers = list(
            db.worker_details.find(
                {"civil_id": {"$in": laydown_office_incharge_ids}, "is_active": True},
                {"civil_id": 1, "name": 1},
            )
        )
        laydown_office_name_by_id = {w.get("civil_id"): (w.get("name") or "") for w in laydown_office_workers}
        for cid in laydown_office_incharge_ids:
            name = (laydown_office_name_by_id.get(cid) or "").strip()
            if not name:
                continue
            db.incharge.update_one(
                {"civil_id": cid},
                {
                    "$set": {
                        "civil_id": cid,
                        "name": name,
                        "locations": ["Laydown Office"],
                        "is_active": True,
                    },
                    "$unset": {"location": ""},
                },
                upsert=True,
            )

        permit_issuer_names = [
            "Atif Rasheed",
            "Tamoor Ali Khan",
            "Mohammad Shoaib Akhtar",
            "Ghulam Qadir Muhammad Yousaf",
        ]
        for pname in permit_issuer_names:
            db.permit_issuer.update_one(
                {"name": pname},
                {"$set": {"name": pname, "is_active": True}},
                upsert=True,
            )

        result = db.system_meta.update_one(
            {"key": "db_initialized"},
            {
                "$set": {
                    "key": "db_initialized",
                    "value": True,
                    "collections": [
                        "worker_auth",
                        "admin_auth",
                        "worker_details",
                        "work_entries",
                        "companies",
                        "locations",
                        "incharge",
                        "permit_issuer",
                    ],
                }
            },
            upsert=True,
        )
        return jsonify(
            {
                "ok": True,
                "db_name": db.name,
                "collection": "system_meta",
                "upserted_id": str(result.upserted_id) if result.upserted_id else None,
            }
        )
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.post("/api/master-data/location")
def create_location():
    payload = request.get_json(force=True) or {}
    name = (payload.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "name is required"}), 400

    db = get_database()
    try:
        result = db.locations.insert_one({"name": name, "is_active": True})
        return jsonify({"ok": True, "id": str(result.inserted_id), "name": name})
    except DuplicateKeyError:
        return jsonify({"ok": False, "error": "location already exists"}), 409


@app.post("/api/master-data/company")
def create_company():
    payload = request.get_json(force=True) or {}
    name = (payload.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "name is required"}), 400

    db = get_database()
    try:
        result = db.companies.insert_one({"name": name, "is_active": True})
        return jsonify({"ok": True, "id": str(result.inserted_id), "name": name})
    except DuplicateKeyError:
        return jsonify({"ok": False, "error": "company already exists"}), 409


@app.post("/api/master-data/incharge")
def create_incharge():
    payload = request.get_json(force=True) or {}
    name = (payload.get("name") or "").strip()
    locations = payload.get("locations")
    location = (payload.get("location") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "name is required"}), 400
    if locations is None:
        locations = [location] if location else []
    if not isinstance(locations, list):
        return jsonify({"ok": False, "error": "locations must be a list of location names"}), 400
    locations = [(str(x) or "").strip() for x in locations]
    locations = [x for x in locations if x]
    if not locations:
        return jsonify({"ok": False, "error": "locations is required"}), 400

    db = get_database()
    for loc in locations:
        if not db.locations.find_one({"name": loc, "is_active": True}):
            return jsonify({"ok": False, "error": f"location not found in locations: {loc}"}), 400

    try:
        result = db.incharge.insert_one({"name": name, "locations": locations, "is_active": True})
        return jsonify({"ok": True, "id": str(result.inserted_id), "name": name, "locations": locations})
    except DuplicateKeyError:
        return jsonify({"ok": False, "error": "incharge already exists"}), 409


@app.post("/api/master-data/permit-issuer")
def create_permit_issuer():
    payload = request.get_json(force=True) or {}
    name = (payload.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "name is required"}), 400

    db = get_database()
    try:
        result = db.permit_issuer.insert_one({"name": name, "is_active": True})
        return jsonify({"ok": True, "id": str(result.inserted_id), "name": name})
    except DuplicateKeyError:
        return jsonify({"ok": False, "error": "permit issuer already exists"}), 409


@app.get("/api/master-data/options")
def get_master_data_options():
    db = get_database()
    companies = [_serialize_doc(item) for item in db.companies.find({"is_active": True})]
    locations = [_serialize_doc(item) for item in db.locations.find({"is_active": True})]
    location_filter = (request.args.get("location") or "").strip()
    incharge_query = {"is_active": True}
    if location_filter:
        incharge_query["locations"] = location_filter
    incharges = [_serialize_doc(item) for item in db.incharge.find(incharge_query)]
    permit_issuers = [_serialize_doc(item) for item in db.permit_issuer.find({"is_active": True})]

    return jsonify(
        {
            "ok": True,
            "companies": companies,
            "locations": locations,
            "incharge": incharges,
            "permit_issuer": permit_issuers,
        }
    )


@app.post("/api/workers")
def create_worker():
    payload = request.get_json(force=True) or {}
    civil_id = (payload.get("civil_id") or "").strip()
    worker_name = (payload.get("name") or "").strip()
    designation = (payload.get("designation") or "").strip()
    company_name = (payload.get("company_name") or "").strip()
    email = (payload.get("email") or "").strip()

    required_fields = {
        "civil_id": civil_id,
        "name": worker_name,
        "designation": designation,
        "company_name": company_name,
    }
    missing = [key for key, value in required_fields.items() if not value]
    if missing:
        return jsonify({"ok": False, "error": f"missing required fields: {', '.join(missing)}"}), 400

    db = get_database()

    if not db.companies.find_one({"name": company_name, "is_active": True}):
        return jsonify({"ok": False, "error": "company_name not found in companies"}), 400

    try:
        result = db.worker_details.insert_one(
            {
                "civil_id": civil_id,
                "name": worker_name,
                "designation": designation,
                "category": _infer_worker_category(designation),
                "company_name": company_name,
                "email": email or None,
                "is_active": True,
            }
        )
        return jsonify({"ok": True, "id": str(result.inserted_id)})
    except DuplicateKeyError:
        return jsonify({"ok": False, "error": "worker with this civil_id already exists"}), 409


@app.get("/api/workers")
def list_workers():
    company_name = (request.args.get("company_name") or "").strip()
    include_inactive = (request.args.get("include_inactive") or "").strip().lower() in {"1", "true", "yes"}
    query = {} if include_inactive else {"is_active": True}
    if company_name:
        query["company_name"] = company_name

    db = get_database()
    workers = [
        _serialize_doc(item)
        for item in db.worker_details.find(
            query,
            {
                "name": 1,
                "civil_id": 1,
                "company_name": 1,
                "designation": 1,
                "category": 1,
                "email": 1,
                "is_active": 1,
                "profile_picture": 1,
            },
        )
    ]
    return jsonify({"ok": True, "workers": workers})


@app.get("/api/workers/search")
def search_workers():
    q = (request.args.get("q") or "").strip()
    if not q:
        return jsonify({"ok": True, "workers": []})

    db = get_database()
    include_inactive = (request.args.get("include_inactive") or "").strip().lower() in {"1", "true", "yes"}
    query = {} if include_inactive else {"is_active": True}
    # If q is numeric treat as civil id search; else name search.
    if re.fullmatch(r"\d+", q):
        query["civil_id"] = q
    else:
        query["name"] = {"$regex": re.escape(q), "$options": "i"}

    workers = [
        _serialize_doc(item)
        for item in db.worker_details.find(
            query,
            {
                "civil_id": 1,
                "name": 1,
                "designation": 1,
                "company_name": 1,
                "category": 1,
                "email": 1,
                "is_active": 1,
                "profile_picture": 1,
            },
        ).sort("_id", -1)
    ]
    return jsonify({"ok": True, "workers": workers})


@app.patch("/api/workers/<civil_id>")
def update_worker(civil_id):
    payload = request.get_json(force=True) or {}
    ok, err = _ensure_management_password(payload)
    if not ok:
        return err, 401
    name = (payload.get("name") or "").strip()
    designation = (payload.get("designation") or "").strip()
    company_name = (payload.get("company_name") or "").strip()
    email = (payload.get("email") or "").strip()
    category_raw = (payload.get("category") or "").strip()
    is_active_raw = payload.get("is_active", None)
    new_civil_id = (payload.get("new_civil_id") or "").strip()

    required_fields = {"name": name, "designation": designation, "company_name": company_name}
    missing = [k for k, v in required_fields.items() if not v]
    if missing:
        return jsonify({"ok": False, "error": f"missing required fields: {', '.join(missing)}"}), 400

    db = get_database()
    if not db.companies.find_one({"name": company_name, "is_active": True}):
        return jsonify({"ok": False, "error": "company_name not found in companies"}), 400

    existing = db.worker_details.find_one({"civil_id": civil_id}, {"_id": 1, "civil_id": 1})
    if not existing:
        return jsonify({"ok": False, "error": "worker not found"}), 404

    if category_raw and category_raw not in {"Direct", "Indirect"}:
        return jsonify({"ok": False, "error": "category must be Direct or Indirect"}), 400
    category_value = category_raw or _infer_worker_category(designation)

    is_active_value = None
    if is_active_raw is not None:
        if isinstance(is_active_raw, bool):
            is_active_value = is_active_raw
        else:
            text = str(is_active_raw).strip().lower()
            if text in {"1", "true", "yes", "active"}:
                is_active_value = True
            elif text in {"0", "false", "no", "inactive"}:
                is_active_value = False
            else:
                return jsonify({"ok": False, "error": "is_active must be true or false"}), 400

    target_civil_id = new_civil_id or civil_id
    if target_civil_id != civil_id:
        dup = db.worker_details.find_one({"civil_id": target_civil_id}, {"_id": 1})
        if dup:
            return jsonify({"ok": False, "error": "new civil_id already exists"}), 409

    db.worker_details.update_one(
        {"civil_id": civil_id},
        {
            "$set": {
                "civil_id": target_civil_id,
                "name": name,
                "designation": designation,
                "category": category_value,
                "company_name": company_name,
                "email": email or None,
                **({"is_active": is_active_value} if is_active_value is not None else {}),
            }
        },
    )
    if target_civil_id != civil_id:
        db.work_entries.update_many({"civil_id": civil_id}, {"$set": {"civil_id": target_civil_id}})
        db.worker_auth.update_many({"civil_id": civil_id}, {"$set": {"civil_id": target_civil_id}})
    return jsonify({"ok": True, "civil_id": target_civil_id})


@app.post("/api/auth/worker-register")
def register_worker_auth():
    payload = request.get_json(force=True) or {}
    civil_id = (payload.get("civil_id") or "").strip()
    company_name = (payload.get("company_name") or "").strip()

    if not civil_id or not company_name:
        return jsonify({"ok": False, "error": "civil_id and company_name are required"}), 400

    db = get_database()
    try:
        result = db.worker_auth.insert_one(
            {
                "civil_id": civil_id,
                "company_name": company_name,
                "is_active": True,
            }
        )
        return jsonify({"ok": True, "id": str(result.inserted_id)})
    except DuplicateKeyError:
        return jsonify({"ok": False, "error": "worker auth already exists"}), 409


@app.post("/api/auth/admin-register")
def register_admin_auth():
    payload = request.get_json(force=True) or {}
    civil_id = (payload.get("civil_id") or "").strip()
    company_name = (payload.get("company_name") or "").strip()
    view_type = (payload.get("view_type") or "management").strip().lower()

    if not civil_id or not company_name:
        return jsonify({"ok": False, "error": "civil_id and company_name are required"}), 400
    if view_type not in {"management", "site_incharge"}:
        return jsonify({"ok": False, "error": "view_type must be management or site_incharge"}), 400

    db = get_database()
    try:
        result = db.admin_auth.insert_one(
            {
                "civil_id": civil_id,
                "company_name": company_name,
                "view_type": view_type,
                "is_active": True,
            }
        )
        return jsonify({"ok": True, "id": str(result.inserted_id)})
    except DuplicateKeyError:
        return jsonify({"ok": False, "error": "admin auth already exists"}), 409


@app.post("/api/auth/worker-login")
def worker_login():
    payload = request.get_json(force=True) or {}
    raw = (payload.get("civil_id") or "").strip()

    if not raw:
        return jsonify({"ok": False, "error": "civil_id is required"}), 400

    db = get_database()
    resolved, err = _resolve_worker_civil_id(raw, db)
    if err == "ambiguous":
        return jsonify(
            {
                "ok": False,
                "error": "multiple workers share these last 4 digits; enter your full Civil ID",
            }
        ), 409
    if resolved is None:
        return jsonify({"ok": False, "error": "invalid worker credentials"}), 401

    worker_profile = db.worker_details.find_one({"civil_id": resolved, "is_active": True})
    if not worker_profile:
        return jsonify({"ok": False, "error": "invalid worker credentials"}), 401

    company_name = (worker_profile or {}).get("company_name") or ""
    return jsonify(
        {
            "ok": True,
            "role": "worker",
            "civil_id": resolved,
            "company_name": company_name,
            "worker_profile": _serialize_doc(worker_profile),
        }
    )


@app.get("/api/worker/self")
def worker_self_profile():
    """Worker portal: load profile by civil_id (same trust model as worker-login)."""
    civil_id = (request.args.get("civil_id") or "").strip()
    if not civil_id:
        return jsonify({"ok": False, "error": "civil_id is required"}), 400
    db = get_database()
    worker = db.worker_details.find_one({"civil_id": civil_id, "is_active": True})
    if not worker:
        return jsonify({"ok": False, "error": "worker not found"}), 404
    return jsonify({"ok": True, "worker_profile": _serialize_doc(worker)})


@app.patch("/api/worker/self/profile-picture")
def worker_patch_profile_picture():
    """Set or clear profile photo on worker_details (keyed by civil_id)."""
    payload = request.get_json(force=True) or {}
    civil_id = (payload.get("civil_id") or "").strip()
    raw_pic = payload.get("profile_picture", "__missing__")

    if not civil_id:
        return jsonify({"ok": False, "error": "civil_id is required"}), 400

    db = get_database()
    worker = db.worker_details.find_one({"civil_id": civil_id, "is_active": True})
    if not worker:
        return jsonify({"ok": False, "error": "worker not found"}), 404

    if raw_pic is None:
        db.worker_details.update_one({"civil_id": civil_id}, {"$unset": {"profile_picture": ""}})
        return jsonify({"ok": True, "profile_picture": None})

    if raw_pic == "__missing__":
        return jsonify({"ok": False, "error": "profile_picture is required (or null to remove)"}), 400

    if not isinstance(raw_pic, str) or not raw_pic.strip():
        return jsonify({"ok": False, "error": "profile_picture must be a data URL or null"}), 400

    ok, err, normalized = _validate_profile_picture_data_url(raw_pic)
    if not ok:
        return jsonify({"ok": False, "error": err}), 400

    db.worker_details.update_one(
        {"civil_id": civil_id},
        {"$set": {"profile_picture": normalized}},
    )
    return jsonify({"ok": True, "profile_picture": normalized})


@app.patch("/api/equipment/self/picture")
def worker_update_equipment_picture():
    """Set or clear equipment photo on equipment_details (worker portal; civil_id + equipment_id)."""
    payload = request.get_json(force=True) or {}
    civil_id = (payload.get("civil_id") or "").strip()
    equipment_id = (payload.get("equipment_id") or "").strip()
    raw_pic = payload.get("equipment_picture", "__missing__")

    if not civil_id or not equipment_id:
        return jsonify({"ok": False, "error": "civil_id and equipment_id are required"}), 400

    db = get_database()
    if not db.worker_details.find_one({"civil_id": civil_id, "is_active": True}):
        return jsonify({"ok": False, "error": "worker not found"}), 404
    try:
        oid = ObjectId(equipment_id)
    except Exception:
        return jsonify({"ok": False, "error": "invalid equipment id"}), 400
    if not db.equipment_details.find_one({"_id": oid, "is_active": True}):
        return jsonify({"ok": False, "error": "equipment not found"}), 404

    if raw_pic is None:
        db.equipment_details.update_one({"_id": oid}, {"$unset": {"equipment_picture": ""}})
        return jsonify({"ok": True, "equipment_picture": None})

    if raw_pic == "__missing__":
        return jsonify({"ok": False, "error": "equipment_picture is required (or null to remove)"}), 400

    if not isinstance(raw_pic, str) or not raw_pic.strip():
        return jsonify({"ok": False, "error": "equipment_picture must be a data URL or null"}), 400

    ok, err, normalized = _validate_profile_picture_data_url(raw_pic)
    if not ok:
        return jsonify({"ok": False, "error": err}), 400

    db.equipment_details.update_one({"_id": oid}, {"$set": {"equipment_picture": normalized}})
    return jsonify({"ok": True, "equipment_picture": normalized})


@app.get("/api/worker/my-entries")
def worker_my_entries():
    """Recent work entries for the signed-in worker (civil_id in query)."""
    civil_id = (request.args.get("civil_id") or "").strip()
    if not civil_id:
        return jsonify({"ok": False, "error": "civil_id is required"}), 400
    db = get_database()
    if not db.worker_details.find_one({"civil_id": civil_id, "is_active": True}):
        return jsonify({"ok": False, "error": "worker not found"}), 404
    limit = request.args.get("limit") or "40"
    try:
        n = max(1, min(int(limit), 100))
    except ValueError:
        n = 40
    cur = (
        db.work_entries.find(
            {"civil_id": civil_id, "status": {"$nin": ["cancelled"]}},
            {
                "work_date": 1,
                "today_activity": 1,
                "location": 1,
                "worker_hours": 1,
                "worker_shift": 1,
                "worker_time_from": 1,
                "worker_time_to": 1,
            },
        )
        .sort([("work_date", -1), ("_id", -1)])
        .limit(n)
    )
    entries = [_serialize_doc(item) for item in cur]
    return jsonify({"ok": True, "entries": entries})


@app.post("/api/auth/admin-login")
def admin_login():
    payload = request.get_json(force=True) or {}
    raw_civil = (payload.get("civil_id") or "").strip()
    view_type = (payload.get("view_type") or "").strip().lower()
    password = str(payload.get("password") or "").strip()

    if not raw_civil:
        return jsonify({"ok": False, "error": "civil_id is required"}), 400
    if view_type not in {"management", "site_incharge"}:
        return jsonify({"ok": False, "error": "select valid view type"}), 400

    db = get_database()
    allowed_ids = (
        SITE_INCHARGE_ALLOWED_CIVIL_IDS if view_type == "site_incharge" else MANAGEMENT_ALLOWED_CIVIL_IDS
    )
    resolved, err = _resolve_civil_id_in_set(raw_civil, allowed_ids)
    if err == "ambiguous":
        return jsonify(
            {
                "ok": False,
                "error": "multiple accounts match these last 4 digits; enter full Civil ID",
            }
        ), 409
    if resolved is None:
        return jsonify({"ok": False, "error": "invalid admin credentials"}), 401
    civil_id = resolved
    if view_type == "site_incharge" and password != SITE_INCHARGE_LOGIN_PASSWORD:
        return jsonify({"ok": False, "error": "invalid admin credentials"}), 401
    if view_type == "management" and password != MANAGEMENT_LOGIN_PASSWORD:
        return jsonify({"ok": False, "error": "invalid admin credentials"}), 401

    worker_profile = db.worker_details.find_one(
        {"civil_id": civil_id, "is_active": True},
        {"company_name": 1, "name": 1},
    )
    if not worker_profile:
        return jsonify({"ok": False, "error": "worker profile not found for this civil id"}), 404

    company_name = (worker_profile or {}).get("company_name", "")
    admin_name = (worker_profile or {}).get("name", "")
    return jsonify(
        {
            "ok": True,
            "role": "admin",
            "view_type": view_type,
            "civil_id": civil_id,
            "company_name": company_name,
            "admin_name": admin_name,
        }
    )


@app.get("/api/management/dashboard")
def management_dashboard():
    db = get_database()
    _sync_worker_categories(db)
    query = _build_entries_query(request.args)
    anchor = _get_anchor_date(request.args)
    day_date = anchor.isoformat()

    total_workers = db.worker_details.count_documents({"is_active": True})
    total_entries = db.work_entries.count_documents(query)
    today_entries = db.work_entries.count_documents(
        {**query, "work_date": date.today().isoformat()}
    )
    logged_civil_ids = db.work_entries.distinct("civil_id", {"work_date": day_date})
    logged_civil_ids = [cid for cid in logged_civil_ids if cid]
    logged_workers = len(logged_civil_ids)
    not_logged_workers = max(0, total_workers - logged_workers)

    # Direct vs Indirect breakdown
    all_active = list(db.worker_details.find(
        {"is_active": True}, {"civil_id": 1, "category": 1, "designation": 1}
    ))
    def _cat(w):
        return w.get("category") or _infer_worker_category(w.get("designation") or "")
    direct_total   = sum(1 for w in all_active if _cat(w) == "Direct")
    indirect_total = sum(1 for w in all_active if _cat(w) == "Indirect")
    cat_map = {w["civil_id"]: _cat(w) for w in all_active if w.get("civil_id")}
    direct_logged   = sum(1 for cid in logged_civil_ids if cat_map.get(cid) == "Direct")
    indirect_logged = sum(1 for cid in logged_civil_ids if cat_map.get(cid) == "Indirect")

    by_location = list(
        db.work_entries.aggregate(
            [
                {"$match": query},
                {
                    "$group": {
                        "_id": "$location",
                        "count": {"$sum": 1},
                        "total_hours": {
                            "$sum": {
                                "$ifNull": ["$hours", {"$ifNull": ["$worker_hours", 0]}]
                            }
                        },
                    }
                },
                {"$sort": {"count": -1}},
            ]
        )
    )
    by_incharge = list(
        db.work_entries.aggregate(
            [
                {"$match": query},
                {"$group": {"_id": "$incharge", "count": {"$sum": 1}}},
                {"$sort": {"count": -1}},
            ]
        )
    )

    recent_entries = [
        _serialize_doc(item)
        for item in db.work_entries.find(
            query,
            {
                "_id": 1,
                "civil_id": 1,
                "worker_name": 1,
                "company_name": 1,
                "location": 1,
                "incharge": 1,
                "work_date": 1,
                "hours": 1,
                "worker_hours": 1,
                "attendance_status": 1,
                "today_activity": 1,
                "status": 1,
                "approval_status": 1,
                "rejection_reason": 1,
                "worker_shift": 1,
            },
        )
        .sort("_id", -1)
    ]
    for e in recent_entries:
        _coalesce_entry_hours(e)

    # Enrich designation/category from worker_details for display/export.
    civil_ids = [e.get("civil_id") for e in recent_entries if e.get("civil_id")]
    designation_map = {}
    category_map = {}
    picture_map = {}
    if civil_ids:
        for w in db.worker_details.find(
            {"civil_id": {"$in": civil_ids}, "is_active": True},
            {
                "civil_id": 1,
                "designation": 1,
                "category": 1,
                "profile_picture": 1,
            },
        ):
            cid = w.get("civil_id")
            designation = w.get("designation") or ""
            category = w.get("category") or _infer_worker_category(designation)
            designation_map[cid] = designation
            category_map[cid] = category
            picture_map[cid] = (w.get("profile_picture") or "").strip()
            if not w.get("category"):
                db.worker_details.update_one({"civil_id": cid}, {"$set": {"category": category}})
    for e in recent_entries:
        cid = e.get("civil_id")
        e["designation"] = designation_map.get(cid, "")
        e["category"] = category_map.get(cid, "Direct")
        e["profile_picture"] = picture_map.get(cid, "") if cid else ""
    _enrich_entries_worker_profiles(db, recent_entries)

    return jsonify(
        {
            "ok": True,
            "summary": {
                "total_workers": total_workers,
                "total_entries": total_entries,
                "today_entries": today_entries,
                "logged_workers": logged_workers,
                "not_logged_workers": not_logged_workers,
                "direct_workers": direct_total,
                "indirect_workers": indirect_total,
                "direct_logged": direct_logged,
                "indirect_logged": indirect_logged,
                "summary_date": day_date,
            },
            "site_incharge_view": [
                {"name": item.get("_id") or "N/A", "entries": item.get("count", 0)}
                for item in by_incharge
            ],
            "location_analytics": [
                {
                    "name": item.get("_id") or "N/A",
                    "entries": item.get("count", 0),
                    "hours": round(item.get("total_hours") or 0, 2),
                }
                for item in by_location
            ],
            "recent_entries": recent_entries,
        }
    )


@app.get("/api/management/not-logged-workers")
def management_not_logged_workers():
    db = get_database()
    _sync_worker_categories(db)
    anchor = _get_anchor_date(request.args)
    day_date = anchor.isoformat()

    logged_civil_ids = db.work_entries.distinct("civil_id", {"work_date": day_date})
    logged_civil_ids = [cid for cid in logged_civil_ids if cid]

    query = {"is_active": True}
    if logged_civil_ids:
        query["civil_id"] = {"$nin": logged_civil_ids}

    workers = [
        _serialize_doc(item)
        for item in db.worker_details.find(
            query,
            {
                "civil_id": 1,
                "name": 1,
                "designation": 1,
                "company_name": 1,
                "category": 1,
                "email": 1,
                "profile_picture": 1,
            },
        ).sort("name", 1)
    ]

    return jsonify(
        {
            "ok": True,
            "date": day_date,
            "count": len(workers),
            "workers": workers,
        }
    )


@app.post("/api/management/not-logged-workers/send-reminder")
def management_send_not_logged_reminder():
    db = get_database()
    _sync_worker_categories(db)
    anchor = _get_anchor_date(request.get_json(force=True) or {})
    day_date = anchor.isoformat()

    logged_civil_ids = db.work_entries.distinct("civil_id", {"work_date": day_date})
    logged_civil_ids = [cid for cid in logged_civil_ids if cid]
    query = {"is_active": True}
    if logged_civil_ids:
        query["civil_id"] = {"$nin": logged_civil_ids}

    workers = list(
        db.worker_details.find(
            query,
            {"civil_id": 1, "name": 1, "email": 1},
        )
    )
    to_emails = []
    for w in workers:
        email = (w.get("email") or "").strip()
        if email and "@" in email:
            to_emails.append(email)
    to_emails = sorted(set(to_emails))

    smtp_user = os.getenv("OUTLOOK_SMTP_USER", "").strip()
    smtp_pass = os.getenv("OUTLOOK_SMTP_PASSWORD", "").strip()
    smtp_host = os.getenv("OUTLOOK_SMTP_HOST", "smtp.office365.com").strip()
    smtp_port = int(os.getenv("OUTLOOK_SMTP_PORT", "587"))
    from_email = os.getenv("OUTLOOK_FROM_EMAIL", smtp_user).strip()
    subject = (
        os.getenv("OUTLOOK_REMINDER_SUBJECT", "").strip()
        or f"Reminder: Please submit daily login ({day_date})"
    )

    if not smtp_user or not smtp_pass or not from_email:
        return jsonify({"ok": False, "error": "email settings missing for SMTP sender"}), 500

    if not to_emails:
        return jsonify(
            {
                "ok": True,
                "date": day_date,
                "sent": 0,
                "skipped": len(workers),
                "message": "No worker email IDs available yet",
            }
        )

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = from_email
    msg["To"] = ", ".join(to_emails)
    msg.set_content(
        "Dear Team,\n\n"
        "This is a reminder to complete your daily manpower login for today.\n\n"
        "Thank you."
    )

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"failed to send reminder: {str(exc)}"}), 500

    return jsonify(
        {
            "ok": True,
            "date": day_date,
            "sent": len(to_emails),
            "skipped": max(0, len(workers) - len(to_emails)),
            "recipients": to_emails,
        }
    )


@app.get("/api/management/filter-entries")
def management_filter_entries():
    db = get_database()
    _sync_worker_categories(db)
    query = _build_entries_query(request.args)

    entries = [
        _serialize_doc(item)
        for item in db.work_entries.find(
            query,
            {
                "_id": 1,
                "worker_name": 1,
                "civil_id": 1,
                "company_name": 1,
                "location": 1,
                "incharge": 1,
                "work_date": 1,
                "hours": 1,
                "worker_hours": 1,
                "status": 1,
                "approval_status": 1,
                "rejection_reason": 1,
                "worker_shift": 1,
                "category": 1,
            },
        ).sort("_id", -1)
    ]
    for e in entries:
        _coalesce_entry_hours(e)
    _enrich_entries_profile_pictures(db, entries)
    _enrich_entries_worker_profiles(db, entries)
    return jsonify({"ok": True, "entries": entries})


@app.get("/api/management/export-excel")
def management_export_excel():
    db = get_database()
    _sync_worker_categories(db)
    query = _build_entries_query(request.args)
    export_rows = _fetch_management_export_rows(db, query)
    buffer = _build_management_export_workbook(export_rows)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"manpower_entries_{date.today().isoformat()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/api/management/export-excel/email")
def management_export_excel_email():
    payload = request.get_json(force=True) or {}
    query = _build_entries_query(payload)
    db = get_database()
    _sync_worker_categories(db)
    export_rows = _fetch_management_export_rows(db, query)
    buffer = _build_management_export_workbook(export_rows)

    smtp_user = os.getenv("OUTLOOK_SMTP_USER", "").strip()
    smtp_pass = os.getenv("OUTLOOK_SMTP_PASSWORD", "").strip()
    smtp_host = os.getenv("OUTLOOK_SMTP_HOST", "smtp.office365.com").strip()
    smtp_port = int(os.getenv("OUTLOOK_SMTP_PORT", "587"))
    from_email = os.getenv("OUTLOOK_FROM_EMAIL", smtp_user).strip()
    to_emails_raw = os.getenv("OUTLOOK_TO_EMAILS", "").strip()
    to_emails = [x.strip() for x in to_emails_raw.split(",") if x.strip()]
    email_subject = (
        os.getenv("OUTLOOK_EMAIL_SUBJECT", "").strip()
        or f"Manpower Tracking Excel - {date.today().isoformat()}"
    )

    if not smtp_user or not smtp_pass or not from_email or not to_emails:
        return (
            jsonify(
                {
                    "ok": False,
                    "error": (
                        "email settings missing. Set OUTLOOK_SMTP_USER, OUTLOOK_SMTP_PASSWORD, "
                        "OUTLOOK_FROM_EMAIL, OUTLOOK_TO_EMAILS"
                    ),
                }
            ),
            500,
        )

    msg = EmailMessage()
    msg["Subject"] = email_subject
    msg["From"] = from_email
    msg["To"] = ", ".join(to_emails)
    msg.set_content(
        "Dear Enco,\n"
        "Please find the Manpower Tracking excel upto this date.\n\n"
        "thank you"
    )
    msg.add_attachment(
        buffer.getvalue(),
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"manpower_entries_{date.today().isoformat()}.xlsx",
    )

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"failed to send email: {str(exc)}"}), 500

    return jsonify({"ok": True, "sent_to": to_emails, "rows": len(export_rows)})


@app.get("/api/site-incharge/today-entries")
def site_incharge_today_entries():
    db = get_database()
    requested_date = (request.args.get("date") or "").strip()
    target_date = requested_date or date.today().isoformat()
    location = (request.args.get("location") or "").strip()
    permit_issuer = (request.args.get("permit_issuer") or "").strip()

    query = {}
    # Default behavior shows all entries; pass ?date=YYYY-MM-DD to filter a specific day.
    if requested_date:
        query["work_date"] = target_date
    if location:
        query["location"] = location
    if permit_issuer:
        query["permit_issuer"] = permit_issuer

    entries = [
        _serialize_doc(item)
        for item in db.work_entries.find(
            query,
            {
                "_id": 1,
                "worker_name": 1,
                "civil_id": 1,
                "company_name": 1,
                "location": 1,
                "incharge": 1,
                "permit_issuer": 1,
                "work_date": 1,
                "hours": 1,
                "worker_hours": 1,
                "worker_time_from": 1,
                "worker_time_to": 1,
                "today_activity": 1,
                "worker_shift": 1,
                "leave_reason": 1,
                "attendance_status": 1,
                "approval_status": 1,
                "status": 1,
            },
        ).sort("_id", -1)
    ]
    # Enrich designation from worker_details for display (not stored in work_entries).
    civil_ids = [e.get("civil_id") for e in entries if e.get("civil_id")]
    designation_map = {}
    picture_map = {}
    if civil_ids:
        for w in db.worker_details.find(
            {"civil_id": {"$in": civil_ids}, "is_active": True},
            {"civil_id": 1, "designation": 1, "profile_picture": 1},
        ):
            cid = w.get("civil_id")
            designation_map[cid] = w.get("designation") or ""
            picture_map[cid] = (w.get("profile_picture") or "").strip()
    for e in entries:
        cid = e.get("civil_id")
        e["designation"] = designation_map.get(cid, "")
        e["profile_picture"] = picture_map.get(cid, "") if cid else ""
    _enrich_entries_worker_profiles(db, entries)

    return jsonify(
        {
            "ok": True,
            "entries": entries,
            "date": target_date,
            "filters": {
                "date": requested_date or "",
                "location": location,
                "permit_issuer": permit_issuer,
            },
        }
    )


@app.get("/api/site-incharge/dashboard")
def site_incharge_dashboard():
    # Compatibility alias for clients calling /dashboard.
    return site_incharge_today_entries()


@app.get("/api/site-incharge/filter-entries")
def site_incharge_filter_entries():
    # Compatibility alias for clients calling /filter-entries.
    return site_incharge_today_entries()


@app.patch("/api/site-incharge/entries/<entry_id>/hours")
def site_incharge_update_hours(entry_id):
    payload = request.get_json(force=True) or {}
    hours = payload.get("hours")
    attendance_status = (payload.get("attendance_status") or "").strip().lower()

    allowed_status = {"idle", "non idle", ""}
    if attendance_status not in allowed_status:
        return jsonify({"ok": False, "error": "attendance_status must be Idle or Non idle"}), 400

    attendance_status_value = attendance_status.capitalize() if attendance_status else None

    try:
        hours_value = float(hours)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "hours must be a valid number"}), 400

    if hours_value < 0 or hours_value > 24:
        return jsonify({"ok": False, "error": "hours must be between 0 and 24"}), 400

    db = get_database()
    try:
        oid = ObjectId(entry_id)
    except Exception:
        return jsonify({"ok": False, "error": "invalid entry id"}), 400

    existing_entry = db.work_entries.find_one({"_id": oid}, {"hours": 1, "status": 1})
    if not existing_entry:
        return jsonify({"ok": False, "error": "entry not found"}), 404
    if existing_entry.get("hours") is not None:
        return jsonify({"ok": False, "error": "hours already submitted for this entry"}), 409

    result = db.work_entries.update_one(
        {"_id": oid},
        {
            "$set": {
                "hours": hours_value,
                "attendance_status": attendance_status_value,
                "approval_status": "approved",
                "status": "hours_updated",
            }
        },
    )

    return jsonify(
        {
            "ok": True,
            "entry_id": entry_id,
            "hours": hours_value,
            "attendance_status": attendance_status_value,
            "approval_status": "approved",
        }
    )


@app.patch("/api/site-incharge/entries/<entry_id>/reject")
def site_incharge_reject_entry(entry_id):
    payload = request.get_json(force=True) or {}
    rejection_reason = (payload.get("rejection_reason") or "").strip()
    if not rejection_reason:
        return jsonify({"ok": False, "error": "rejection_reason is required"}), 400

    db = get_database()
    try:
        oid = ObjectId(entry_id)
    except Exception:
        return jsonify({"ok": False, "error": "invalid entry id"}), 400

    existing_entry = db.work_entries.find_one({"_id": oid}, {"approval_status": 1, "hours": 1})
    if not existing_entry:
        return jsonify({"ok": False, "error": "entry not found"}), 404
    if existing_entry.get("hours") is not None or existing_entry.get("approval_status") == "approved":
        return jsonify({"ok": False, "error": "already approved"}), 409

    db.work_entries.update_one(
        {"_id": oid},
        {
            "$set": {
                "approval_status": "rejected",
                "status": "rejected",
                "rejection_reason": rejection_reason,
            }
        },
    )
    return jsonify(
        {
            "ok": True,
            "entry_id": entry_id,
            "approval_status": "rejected",
            "rejection_reason": rejection_reason,
        }
    )


@app.post("/api/work-entries")
def create_work_entry():
    payload = request.get_json(force=True) or {}
    civil_id = (payload.get("civil_id") or "").strip()
    company_name = (payload.get("company_name") or "").strip()
    worker_name = (payload.get("worker_name") or "").strip()
    work_date = date.today().isoformat()
    location = (payload.get("location") or "").strip()
    incharge = (payload.get("incharge") or "").strip()
    permit_issuer = (payload.get("permit_issuer") or "").strip()
    today_activity = (payload.get("today_activity") or "").strip()
    worker_hours = payload.get("worker_hours")
    worker_time_from = (payload.get("worker_time_from") or "").strip()
    worker_time_to = (payload.get("worker_time_to") or "").strip()
    worker_shift = (payload.get("worker_shift") or "Day").strip()
    leave_reason = (payload.get("leave_reason") or "").strip()
    allowed_worker_shifts = {"Day", "Night", "Request for leave"}
    if worker_shift not in allowed_worker_shifts:
        return jsonify({"ok": False, "error": "worker_shift must be Day, Night, or Request for leave"}), 400
    if worker_shift == "Request for leave" and not leave_reason:
        return jsonify({"ok": False, "error": "leave_reason is required for leave request"}), 400


    required_fields = {
        "civil_id": civil_id,
        "company_name": company_name,
        "worker_name": worker_name,
        "work_date": work_date,
    }
    missing = [key for key, value in required_fields.items() if not value]
    if missing:
        return jsonify({"ok": False, "error": f"missing required fields: {', '.join(missing)}"}), 400

    def _parse_hhmm(value: str):
        if not value:
            return None
        parts = value.split(":")
        if len(parts) != 2:
            return None
        try:
            hh = int(parts[0])
            mm = int(parts[1])
        except ValueError:
            return None
        if hh < 0 or hh > 23 or mm < 0 or mm > 59:
            return None
        return hh * 60 + mm

    worker_hours_value = None
    if worker_hours is not None and worker_hours != "":
        try:
            worker_hours_value = float(worker_hours)
        except (TypeError, ValueError):
            return jsonify({"ok": False, "error": "worker_hours must be a valid number"}), 400
        if worker_hours_value < 0 or worker_hours_value > 12:
            return jsonify({"ok": False, "error": "work hours must be between 0 and 12"}), 400

    # Prefer time range if provided (From/To). If one is provided, both are required.
    if worker_time_from or worker_time_to:
        start_min = _parse_hhmm(worker_time_from)
        end_min = _parse_hhmm(worker_time_to)
        if start_min is None or end_min is None:
            return jsonify({"ok": False, "error": "work time must be in HH:MM format"}), 400
        if end_min <= start_min:
            return jsonify({"ok": False, "error": "to time must be after from time"}), 400
        diff_hours = (end_min - start_min) / 60.0
        if diff_hours > 12:
            return jsonify({"ok": False, "error": "work hours must be between 0 and 12"}), 400
        worker_hours_value = diff_hours

    if worker_shift != "Request for leave" and worker_hours_value is None:
        return jsonify({"ok": False, "error": "work hours are required (select From/To time)"}), 400
    if worker_shift == "Request for leave":
        worker_hours_value = 0.0
        worker_time_from = ""
        worker_time_to = ""

    # Optional daily fields: allow Null/empty from UI and store as None.
    location_value = location or None
    incharge_value = incharge or None
    permit_issuer_value = permit_issuer or None

    db = get_database()

    if not db.companies.find_one({"name": company_name, "is_active": True}):
        return jsonify({"ok": False, "error": "invalid company_name"}), 400

    if location_value and not db.locations.find_one({"name": location_value, "is_active": True}):
        return jsonify({"ok": False, "error": "invalid location"}), 400

    if incharge_value:
        inc_doc = db.incharge.find_one({"name": incharge_value, "is_active": True})
        if not inc_doc:
            return jsonify({"ok": False, "error": "invalid incharge"}), 400
        inc_locations = inc_doc.get("locations")
        if (
            location_value
            and isinstance(inc_locations, list)
            and inc_locations
            and location_value not in inc_locations
        ):
            return jsonify({"ok": False, "error": "incharge does not match selected location"}), 400

    # For Head Office / Laydown Office permit issuer is not required (force null).
    if location_value in {"Head Office", "Laydown Office"}:
        permit_issuer_value = None
    if permit_issuer_value and not db.permit_issuer.find_one({"name": permit_issuer_value, "is_active": True}):
        return jsonify({"ok": False, "error": "invalid permit_issuer"}), 400

    worker = db.worker_details.find_one(
        {"civil_id": civil_id, "company_name": company_name, "is_active": True}
    )
    if not worker and civil_id != DEVELOPER_MASTER_CIVIL_ID:
        return jsonify({"ok": False, "error": "worker profile not found"}), 404

    # Block duplicate timing: same civil_id + same date + same from/to time
    if worker_time_from and worker_time_to:
        duplicate_timing = db.work_entries.find_one({
            "civil_id": civil_id,
            "work_date": work_date,
            "worker_time_from": worker_time_from,
            "worker_time_to": worker_time_to,
            "status": {"$nin": ["cancelled"]},
        })
        if duplicate_timing:
            return jsonify({
                "ok": False,
                "error": "duplicate_timing",
                "message": f"You already have an entry from {worker_time_from} to {worker_time_to} today. Please choose a different time.",
            }), 409

    item_tag = (payload.get("item_tag") or "").strip() or None

    result = db.work_entries.insert_one(
        {
            "civil_id": civil_id,
            "company_name": company_name,
            "worker_name": worker_name,
            "work_date": work_date,
            "location": location_value,
            "incharge": incharge_value,
            "permit_issuer": permit_issuer_value,
            "worker_hours": worker_hours_value,
            "worker_time_from": worker_time_from or None,
            "worker_time_to": worker_time_to or None,
            "worker_shift": worker_shift,
            "leave_reason": leave_reason if worker_shift == "Request for leave" else None,
            "hours": None,
            "attendance_status": None,
            "today_activity": today_activity,
            "item_tag": item_tag,
            "status": "submitted",
        }
    )
    return jsonify({"ok": True, "id": str(result.inserted_id)})


@app.get("/api/work-entries/today-status")
def work_entry_today_status():
    """Check whether a worker already has an entry for today."""
    civil_id = (request.args.get("civil_id") or "").strip()
    if not civil_id:
        return jsonify({"ok": False, "error": "civil_id required"}), 400
    work_date = date.today().isoformat()
    db = get_database()
    existing = db.work_entries.find_one(
        {"civil_id": civil_id, "work_date": work_date, "status": {"$nin": ["cancelled"]}},
        {"worker_time_from": 1, "worker_time_to": 1, "worker_shift": 1,
         "location": 1, "today_activity": 1, "worker_hours": 1},
    )
    if existing:
        existing["_id"] = str(existing["_id"])
        return jsonify({"ok": True, "submitted": True, "entry": existing})
    return jsonify({"ok": True, "submitted": False})


@app.patch("/api/work-entries/<entry_id>")
def update_work_entry(entry_id):
    payload = request.get_json(force=True) or {}
    ok, err = _ensure_management_password(payload)
    if not ok:
        return err, 401
    db = get_database()
    try:
        oid = ObjectId(entry_id)
    except Exception:
        return jsonify({"ok": False, "error": "invalid entry id"}), 400

    existing = db.work_entries.find_one({"_id": oid})
    if not existing:
        return jsonify({"ok": False, "error": "entry not found"}), 404

    allowed_status = {"Day", "Idle", "Absent", "", None}
    attendance_status = payload.get("attendance_status", existing.get("attendance_status"))
    if attendance_status not in allowed_status:
        return jsonify({"ok": False, "error": "attendance_status must be Day, Idle, or Absent"}), 400

    hours_raw = payload.get("hours", existing.get("hours"))
    hours_value = None
    if hours_raw is not None and hours_raw != "":
        try:
            hours_value = float(hours_raw)
        except (TypeError, ValueError):
            return jsonify({"ok": False, "error": "hours must be a valid number"}), 400
        if hours_value < 0 or hours_value > 24:
            return jsonify({"ok": False, "error": "hours must be between 0 and 24"}), 400

    location_value = (payload.get("location", existing.get("location")) or "").strip() or None
    incharge_value = (payload.get("incharge", existing.get("incharge")) or "").strip() or None
    today_activity_value = (payload.get("today_activity", existing.get("today_activity")) or "").strip()
    work_date_value = (payload.get("work_date", existing.get("work_date")) or "").strip()
    if not work_date_value:
        work_date_value = existing.get("work_date") or date.today().isoformat()

    if location_value and not db.locations.find_one({"name": location_value, "is_active": True}):
        return jsonify({"ok": False, "error": "invalid location"}), 400

    if incharge_value:
        inc_doc = db.incharge.find_one({"name": incharge_value, "is_active": True})
        if not inc_doc:
            return jsonify({"ok": False, "error": "invalid incharge"}), 400
        inc_locations = inc_doc.get("locations")
        if (
            location_value
            and isinstance(inc_locations, list)
            and inc_locations
            and location_value not in inc_locations
        ):
            return jsonify({"ok": False, "error": "incharge does not match selected location"}), 400

    approval_status_value = (payload.get("approval_status", existing.get("approval_status")) or "").strip().lower()
    if approval_status_value not in {"", "pending", "approved", "rejected"}:
        return jsonify({"ok": False, "error": "approval_status must be pending, approved, rejected, or empty"}), 400
    rejection_reason_value = (payload.get("rejection_reason", existing.get("rejection_reason")) or "").strip() or None
    if approval_status_value == "rejected" and not rejection_reason_value:
        return jsonify({"ok": False, "error": "rejection_reason is required when approval_status is rejected"}), 400
    if approval_status_value != "rejected":
        rejection_reason_value = None

    db.work_entries.update_one(
        {"_id": oid},
        {
            "$set": {
                "work_date": work_date_value,
                "location": location_value,
                "incharge": incharge_value,
                "today_activity": today_activity_value,
                "attendance_status": attendance_status or None,
                "hours": hours_value,
                "approval_status": approval_status_value or None,
                "rejection_reason": rejection_reason_value,
                "status": "edited_by_management",
            }
        },
    )
    return jsonify({"ok": True, "entry_id": entry_id})


@app.delete("/api/work-entries/<entry_id>")
def delete_work_entry(entry_id):
    db = get_database()
    try:
        oid = ObjectId(entry_id)
    except Exception:
        return jsonify({"ok": False, "error": "invalid entry id"}), 400

    result = db.work_entries.delete_one({"_id": oid})
    if result.deleted_count == 0:
        return jsonify({"ok": False, "error": "entry not found"}), 404
    return jsonify({"ok": True, "entry_id": entry_id})


# ─────────────────────────────────────────────────────────────────────────────
# SITE-INCHARGE – equipment entry approval
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/site-incharge/equipment-entries")
def site_incharge_equipment_entries():
    db = get_database()
    requested_date = (request.args.get("date") or "").strip()
    location = (request.args.get("location") or "").strip()
    eq_type = (request.args.get("type") or "").strip()

    query = {}
    if requested_date:
        query["work_date"] = requested_date
    if location:
        query["location"] = location
    if eq_type:
        query["equipment_type"] = eq_type

    entries = [
        _serialize_doc(item)
        for item in db.equipment_entries.find(query).sort("_id", -1)
    ]
    _enrich_entries_profile_pictures(db, entries)
    return jsonify({"ok": True, "entries": entries})


@app.patch("/api/site-incharge/equipment-entries/<entry_id>/approve")
def site_incharge_approve_equipment(entry_id):
    db = get_database()
    try:
        oid = ObjectId(entry_id)
    except Exception:
        return jsonify({"ok": False, "error": "invalid entry id"}), 400

    existing = db.equipment_entries.find_one({"_id": oid})
    if not existing:
        return jsonify({"ok": False, "error": "entry not found"}), 404
    if existing.get("approval_status") == "approved":
        return jsonify({"ok": False, "error": "already approved"}), 409

    payload = request.get_json(force=True) or {}
    approved_by = (payload.get("approved_by") or "").strip()

    db.equipment_entries.update_one(
        {"_id": oid},
        {"$set": {
            "approval_status": "approved",
            "approved_by": approved_by or None,
            "status": "approved",
        }},
    )
    return jsonify({"ok": True, "entry_id": entry_id})


@app.patch("/api/site-incharge/equipment-entries/<entry_id>/reject")
def site_incharge_reject_equipment(entry_id):
    payload = request.get_json(force=True) or {}
    rejection_reason = (payload.get("rejection_reason") or "").strip()
    if not rejection_reason:
        return jsonify({"ok": False, "error": "rejection_reason is required"}), 400

    db = get_database()
    try:
        oid = ObjectId(entry_id)
    except Exception:
        return jsonify({"ok": False, "error": "invalid entry id"}), 400

    existing = db.equipment_entries.find_one({"_id": oid})
    if not existing:
        return jsonify({"ok": False, "error": "entry not found"}), 404
    if existing.get("approval_status") == "approved":
        return jsonify({"ok": False, "error": "already approved"}), 409

    db.equipment_entries.update_one(
        {"_id": oid},
        {"$set": {
            "approval_status": "rejected",
            "rejection_reason": rejection_reason,
            "status": "rejected",
        }},
    )
    return jsonify({"ok": True, "entry_id": entry_id})


# ─────────────────────────────────────────────────────────────────────────────
# MANAGEMENT – equipment dashboard + edit/delete + Excel
# ─────────────────────────────────────────────────────────────────────────────

def _opt_float_rate(val):
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


@app.post("/api/management/equipment")
def management_create_equipment():
    """Create a new equipment row in equipment_details (management only)."""
    payload = request.get_json(force=True) or {}
    name = (payload.get("name") or "").strip()
    plate_number = (payload.get("plate_number") or "").strip()
    equipment_type = (payload.get("equipment_type") or "").strip()
    location = (payload.get("location") or "").strip()
    ownership_raw = (payload.get("ownership") or "owned").strip().lower()

    if not name or not equipment_type or not location:
        return jsonify({"ok": False, "error": "name, equipment_type, and location are required"}), 400
    if ownership_raw not in {"owned", "rental"}:
        return jsonify({"ok": False, "error": "ownership must be owned or rental"}), 400

    sr = payload.get("supply_rate") or {}
    cr = payload.get("contract_rate") or {}

    def _rate_dict(src):
        return {
            "hourly": _opt_float_rate(src.get("hourly")),
            "daily": _opt_float_rate(src.get("daily")),
            "weekly": _opt_float_rate(src.get("weekly")),
            "monthly": _opt_float_rate(src.get("monthly")),
        }

    supply_rate = _rate_dict(sr)
    contract_rate = _rate_dict(cr)

    doc = {
        "name": name,
        "plate_number": plate_number,
        "equipment_type": equipment_type,
        "location": location,
        "ownership": ownership_raw,
        "supply_rate": supply_rate,
        "contract_rate": contract_rate,
        "is_active": True,
    }
    doc["display_name"] = _equipment_display_name(doc)

    db = get_database()
    try:
        db.equipment_details.create_index(
            [("name", 1), ("plate_number", 1), ("location", 1)], unique=True
        )
    except Exception:
        pass

    if "equipment_picture" in payload:
        ep_raw = payload.get("equipment_picture")
        if ep_raw is not None:
            if not isinstance(ep_raw, str) or not ep_raw.strip():
                return jsonify({"ok": False, "error": "equipment_picture must be a data URL or null"}), 400
            ok, err, normalized = _validate_profile_picture_data_url(ep_raw)
            if not ok:
                return jsonify({"ok": False, "error": err}), 400
            doc["equipment_picture"] = normalized

    try:
        result = db.equipment_details.insert_one(doc)
    except DuplicateKeyError:
        return jsonify({"ok": False, "error": "equipment with same name, plate, and location already exists"}), 409

    doc["_id"] = str(result.inserted_id)
    return jsonify({"ok": True, "equipment": doc})


@app.get("/api/management/progress-updates")
def management_list_progress_updates():
    db = get_database()
    query = {}
    location = (request.args.get("location") or "").strip()
    status = (request.args.get("status") or "").strip().lower()
    period = _normalize_period_filter(request.args.get("period") or "all")
    anchor_date_raw = (request.args.get("anchor_date") or "").strip()
    try:
        anchor = datetime.strptime(anchor_date_raw, "%Y-%m-%d").date() if anchor_date_raw else date.today()
    except ValueError:
        anchor = date.today()

    if location and location.lower() != "all":
        query["location"] = location
    if status and status != "all":
        query["status"] = status

    if period != "all":
        start = anchor
        end = anchor
        if period == "day":
            start = end = anchor
        elif period == "week":
            start = anchor - timedelta(days=anchor.weekday())
            end = start + timedelta(days=6)
        elif period == "month":
            start = anchor.replace(day=1)
            next_month = start.replace(year=start.year + 1, month=1, day=1) if start.month == 12 else start.replace(month=start.month + 1, day=1)
            end = next_month - timedelta(days=1)
        elif period == "quarter":
            qsm = ((anchor.month - 1) // 3) * 3 + 1
            start = anchor.replace(month=qsm, day=1)
            next_q = start.replace(year=start.year + 1, month=1, day=1) if qsm == 10 else start.replace(month=qsm + 3, day=1)
            end = next_q - timedelta(days=1)
        elif period == "year":
            start = anchor.replace(month=1, day=1)
            end = anchor.replace(month=12, day=31)
        query["work_date"] = {"$gte": start.isoformat(), "$lte": end.isoformat()}

    entries = [_serialize_doc(x) for x in db.progress_updates.find(query).sort([("work_date", -1), ("_id", -1)])]
    total = len(entries)
    avg_progress = round(sum(float(e.get("progress_percent") or 0) for e in entries) / total, 2) if total else 0
    completed = len([e for e in entries if (e.get("status") or "").lower() == "completed"])
    in_progress = len([e for e in entries if (e.get("status") or "").lower() == "in-progress"])
    delayed = len([e for e in entries if (e.get("status") or "").lower() == "delayed"])

    by_location = list(
        db.progress_updates.aggregate(
            [
                {"$match": query},
                {"$group": {"_id": "$location", "count": {"$sum": 1}, "avg_progress": {"$avg": "$progress_percent"}}},
                {"$sort": {"count": -1}},
            ]
        )
    )
    return jsonify(
        {
            "ok": True,
            "summary": {
                "total_updates": total,
                "avg_progress_percent": avg_progress,
                "completed": completed,
                "in_progress": in_progress,
                "delayed": delayed,
            },
            "by_location": [
                {"name": i.get("_id") or "N/A", "updates": i.get("count", 0), "avg_progress": round(i.get("avg_progress") or 0, 2)}
                for i in by_location
            ],
            "entries": entries,
        }
    )


@app.post("/api/management/progress-updates")
def management_create_progress_update():
    payload = request.get_json(force=True) or {}
    ok, err = _ensure_management_password(payload)
    if not ok:
        return err, 401

    title = (payload.get("title") or "").strip()
    work_date = (payload.get("work_date") or date.today().isoformat()).strip()
    location = (payload.get("location") or "").strip()
    update_text = (payload.get("update_text") or "").strip()
    status = (payload.get("status") or "in-progress").strip().lower()
    updated_by = (payload.get("updated_by") or "").strip()
    remarks = (payload.get("remarks") or "").strip()
    progress_percent_raw = payload.get("progress_percent")

    if not title or not location or not update_text:
        return jsonify({"ok": False, "error": "title, location, and update_text are required"}), 400
    if status not in {"planned", "in-progress", "completed", "delayed"}:
        return jsonify({"ok": False, "error": "status must be planned, in-progress, completed, or delayed"}), 400
    try:
        progress_percent = float(progress_percent_raw)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "progress_percent must be a number"}), 400
    if progress_percent < 0 or progress_percent > 100:
        return jsonify({"ok": False, "error": "progress_percent must be between 0 and 100"}), 400

    db = get_database()
    result = db.progress_updates.insert_one(
        {
            "title": title,
            "work_date": work_date,
            "location": location,
            "update_text": update_text,
            "status": status,
            "progress_percent": progress_percent,
            "updated_by": updated_by or None,
            "remarks": remarks or None,
            "created_at": datetime.utcnow().isoformat(),
            "updated_at": datetime.utcnow().isoformat(),
        }
    )
    return jsonify({"ok": True, "id": str(result.inserted_id)})


@app.patch("/api/management/progress-updates/<update_id>")
def management_update_progress_update(update_id):
    payload = request.get_json(force=True) or {}
    ok, err = _ensure_management_password(payload)
    if not ok:
        return err, 401
    db = get_database()
    try:
        oid = ObjectId(update_id)
    except Exception:
        return jsonify({"ok": False, "error": "invalid update id"}), 400
    existing = db.progress_updates.find_one({"_id": oid})
    if not existing:
        return jsonify({"ok": False, "error": "progress update not found"}), 404

    title = (payload.get("title", existing.get("title")) or "").strip()
    work_date = (payload.get("work_date", existing.get("work_date")) or "").strip()
    location = (payload.get("location", existing.get("location")) or "").strip()
    update_text = (payload.get("update_text", existing.get("update_text")) or "").strip()
    status = (payload.get("status", existing.get("status")) or "in-progress").strip().lower()
    updated_by = (payload.get("updated_by", existing.get("updated_by")) or "").strip()
    remarks = (payload.get("remarks", existing.get("remarks")) or "").strip()
    progress_percent_raw = payload.get("progress_percent", existing.get("progress_percent"))

    if not title or not location or not update_text:
        return jsonify({"ok": False, "error": "title, location, and update_text are required"}), 400
    if status not in {"planned", "in-progress", "completed", "delayed"}:
        return jsonify({"ok": False, "error": "status must be planned, in-progress, completed, or delayed"}), 400
    try:
        progress_percent = float(progress_percent_raw)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "progress_percent must be a number"}), 400
    if progress_percent < 0 or progress_percent > 100:
        return jsonify({"ok": False, "error": "progress_percent must be between 0 and 100"}), 400

    db.progress_updates.update_one(
        {"_id": oid},
        {
            "$set": {
                "title": title,
                "work_date": work_date,
                "location": location,
                "update_text": update_text,
                "status": status,
                "progress_percent": progress_percent,
                "updated_by": updated_by or None,
                "remarks": remarks or None,
                "updated_at": datetime.utcnow().isoformat(),
            }
        },
    )
    return jsonify({"ok": True, "id": update_id})


@app.delete("/api/management/progress-updates/<update_id>")
def management_delete_progress_update(update_id):
    payload = request.get_json(force=True) or {}
    ok, err = _ensure_management_password(payload)
    if not ok:
        return err, 401
    db = get_database()
    try:
        oid = ObjectId(update_id)
    except Exception:
        return jsonify({"ok": False, "error": "invalid update id"}), 400
    result = db.progress_updates.delete_one({"_id": oid})
    if result.deleted_count == 0:
        return jsonify({"ok": False, "error": "progress update not found"}), 404
    return jsonify({"ok": True, "id": update_id})


@app.patch("/api/management/equipment/<equipment_id>")
def management_update_equipment(equipment_id):
    """Update full equipment master details (management)."""
    payload = request.get_json(force=True) or {}
    ok, err = _ensure_management_password(payload)
    if not ok:
        return err, 401
    db = get_database()
    try:
        oid = ObjectId(equipment_id)
    except Exception:
        return jsonify({"ok": False, "error": "invalid equipment id"}), 400

    existing = db.equipment_details.find_one({"_id": oid})
    if not existing:
        return jsonify({"ok": False, "error": "equipment not found"}), 404

    name = (payload.get("name", existing.get("name")) or "").strip()
    plate_number = (payload.get("plate_number", existing.get("plate_number")) or "").strip()
    equipment_type = (payload.get("equipment_type", existing.get("equipment_type")) or "").strip()
    location = (payload.get("location", existing.get("location")) or "").strip()
    ownership_raw = (payload.get("ownership", existing.get("ownership")) or "owned").strip().lower()
    is_active_raw = payload.get("is_active", existing.get("is_active", True))

    if not name or not equipment_type or not location:
        return jsonify({"ok": False, "error": "name, equipment_type, and location are required"}), 400
    if ownership_raw not in {"owned", "rental"}:
        return jsonify({"ok": False, "error": "ownership must be owned or rental"}), 400

    if isinstance(is_active_raw, bool):
        is_active_value = is_active_raw
    else:
        text = str(is_active_raw).strip().lower()
        if text in {"1", "true", "yes", "active"}:
            is_active_value = True
        elif text in {"0", "false", "no", "inactive"}:
            is_active_value = False
        else:
            return jsonify({"ok": False, "error": "is_active must be true or false"}), 400

    sr_payload = payload.get("supply_rate")
    cr_payload = payload.get("contract_rate")
    supply_rate = existing.get("supply_rate") or {}
    contract_rate = existing.get("contract_rate") or {}

    if isinstance(sr_payload, dict):
        supply_rate = {
            "hourly": _opt_float_rate(sr_payload.get("hourly")),
            "daily": _opt_float_rate(sr_payload.get("daily")),
            "weekly": _opt_float_rate(sr_payload.get("weekly")),
            "monthly": _opt_float_rate(sr_payload.get("monthly")),
        }
    if isinstance(cr_payload, dict):
        contract_rate = {
            "hourly": _opt_float_rate(cr_payload.get("hourly")),
            "daily": _opt_float_rate(cr_payload.get("daily")),
            "weekly": _opt_float_rate(cr_payload.get("weekly")),
            "monthly": _opt_float_rate(cr_payload.get("monthly")),
        }

    dup = db.equipment_details.find_one(
        {
            "_id": {"$ne": oid},
            "name": name,
            "plate_number": plate_number,
            "location": location,
        },
        {"_id": 1},
    )
    if dup:
        return jsonify({"ok": False, "error": "another equipment with same name, plate, and location already exists"}), 409

    updated_doc = {
        "name": name,
        "plate_number": plate_number,
        "equipment_type": equipment_type,
        "location": location,
        "ownership": ownership_raw,
        "supply_rate": supply_rate,
        "contract_rate": contract_rate,
        "is_active": is_active_value,
    }
    updated_doc["display_name"] = _equipment_display_name(updated_doc)

    mongo_update = {"$set": updated_doc}
    if "equipment_picture" in payload:
        ep_raw = payload.get("equipment_picture")
        if ep_raw is None:
            mongo_update["$unset"] = {"equipment_picture": ""}
        else:
            if not isinstance(ep_raw, str) or not ep_raw.strip():
                return jsonify({"ok": False, "error": "equipment_picture must be a data URL or null"}), 400
            ok, err, normalized = _validate_profile_picture_data_url(ep_raw)
            if not ok:
                return jsonify({"ok": False, "error": err}), 400
            updated_doc["equipment_picture"] = normalized

    db.equipment_details.update_one({"_id": oid}, mongo_update)

    # Keep entry snapshots aligned for reporting readability.
    db.equipment_entries.update_many(
        {"equipment_id": equipment_id},
        {
            "$set": {
                "equipment_name": name,
                "plate_number": plate_number,
                "equipment_type": equipment_type,
                "equipment_location": location,
                "ownership": ownership_raw,
                "supply_rate": supply_rate,
                "contract_rate": contract_rate,
            }
        },
    )
    updated_doc["_id"] = equipment_id
    return jsonify({"ok": True, "equipment": updated_doc})


@app.get("/api/management/equipment-dashboard")
def management_equipment_dashboard():
    db = get_database()
    query = {}
    work_date = (request.args.get("date") or "").strip()
    location = (request.args.get("location") or "").strip()
    eq_type = (request.args.get("type") or "").strip()
    period = _normalize_period_filter(request.args.get("period") or "all")
    anchor_date_raw = (request.args.get("anchor_date") or "").strip()

    try:
        anchor = datetime.strptime(anchor_date_raw, "%Y-%m-%d").date() if anchor_date_raw else date.today()
    except ValueError:
        anchor = date.today()

    if work_date:
        query["work_date"] = work_date
    elif period != "all":
        start = anchor
        end = anchor
        if period == "day":
            start = end = anchor
        elif period == "week":
            start = anchor - timedelta(days=anchor.weekday())
            end = start + timedelta(days=6)
        elif period == "month":
            start = anchor.replace(day=1)
            next_m = (start.replace(month=start.month % 12 + 1, day=1)
                      if start.month < 12 else start.replace(year=start.year + 1, month=1, day=1))
            end = next_m - timedelta(days=1)
        elif period == "quarter":
            qsm = ((anchor.month - 1) // 3) * 3 + 1
            start = anchor.replace(month=qsm, day=1)
            next_q = (start.replace(month=qsm + 3, day=1)
                      if qsm <= 9 else start.replace(year=start.year + 1, month=1, day=1))
            end = next_q - timedelta(days=1)
        elif period == "year":
            start = anchor.replace(month=1, day=1)
            end = anchor.replace(month=12, day=31)
        query["work_date"] = {"$gte": start.isoformat(), "$lte": end.isoformat()}

    if location and location != "all":
        query["location"] = location
    if eq_type and eq_type != "all":
        query["equipment_type"] = eq_type

    # ── Core counts ──
    total_equipment = db.equipment_details.count_documents({"is_active": True})
    total_entries = db.equipment_entries.count_documents(query)
    owned_entries = db.equipment_entries.count_documents({**query, "ownership": "owned"})
    rental_entries = db.equipment_entries.count_documents({**query, "ownership": "rental"})

    # ── By type ──
    by_type = list(db.equipment_entries.aggregate([
        {"$match": query},
        {"$group": {
            "_id": "$equipment_type",
            "count": {"$sum": 1},
            "total_hours": {"$sum": "$hours"},
            "rental_cost": {"$sum": {"$cond": [{"$eq": ["$ownership", "rental"]}, {"$ifNull": ["$rental_amount", 0]}, 0]}},
        }},
        {"$sort": {"count": -1}},
    ]))

    # ── By location ──
    by_location = list(db.equipment_entries.aggregate([
        {"$match": query},
        {"$group": {
            "_id": "$location",
            "count": {"$sum": 1},
            "total_hours": {"$sum": "$hours"},
            "rental_cost": {"$sum": {"$cond": [{"$eq": ["$ownership", "rental"]}, {"$ifNull": ["$rental_amount", 0]}, 0]}},
        }},
        {"$sort": {"count": -1}},
    ]))

    # ── Per-equipment breakdown ──
    by_equipment = list(db.equipment_entries.aggregate([
        {"$match": query},
        {"$group": {
            "_id": "$equipment_id",
            "equipment_name": {"$first": "$equipment_name"},
            "plate_number": {"$first": "$plate_number"},
            "equipment_type": {"$first": "$equipment_type"},
            "equipment_location": {"$first": "$equipment_location"},
            "ownership": {"$first": "$ownership"},
            "count": {"$sum": 1},
            "total_hours": {"$sum": "$hours"},
            "rental_cost": {"$sum": {"$cond": [{"$eq": ["$ownership", "rental"]}, {"$ifNull": ["$rental_amount", 0]}, 0]}},
        }},
        {"$sort": {"rental_cost": -1}},
    ]))

    # ── Weekly breakdown (last 8 weeks relative to anchor) ──
    week_start = anchor - timedelta(days=anchor.weekday())
    weekly_data = []
    for i in range(7, -1, -1):
        ws = week_start - timedelta(weeks=i)
        we = ws + timedelta(days=6)
        label = f"W{ws.isocalendar()[1]} ({ws.strftime('%d %b')})"
        agg = list(db.equipment_entries.aggregate([
            {"$match": {**{k: v for k, v in query.items() if k != "work_date"},
                        "work_date": {"$gte": ws.isoformat(), "$lte": we.isoformat()}}},
            {"$group": {
                "_id": None,
                "entries": {"$sum": 1},
                "total_hours": {"$sum": "$hours"},
                "rental_cost": {"$sum": {"$cond": [{"$eq": ["$ownership", "rental"]}, {"$ifNull": ["$rental_amount", 0]}, 0]}},
            }},
        ]))
        row = agg[0] if agg else {}
        weekly_data.append({
            "week": label,
            "week_start": ws.isoformat(),
            "entries": row.get("entries", 0),
            "total_hours": round(row.get("total_hours") or 0, 2),
            "rental_cost": round(row.get("rental_cost") or 0, 2),
        })

    # ── Monthly breakdown (last 6 months) ──
    monthly_data = []
    for i in range(5, -1, -1):
        mo = (anchor.month - i - 1) % 12 + 1
        yr = anchor.year + ((anchor.month - i - 1) // 12)
        ms = date(yr, mo, 1)
        next_mo = date(yr + (mo // 12), mo % 12 + 1, 1) if mo < 12 else date(yr + 1, 1, 1)
        me = next_mo - timedelta(days=1)
        agg = list(db.equipment_entries.aggregate([
            {"$match": {**{k: v for k, v in query.items() if k != "work_date"},
                        "work_date": {"$gte": ms.isoformat(), "$lte": me.isoformat()}}},
            {"$group": {
                "_id": None,
                "entries": {"$sum": 1},
                "total_hours": {"$sum": "$hours"},
                "rental_cost": {"$sum": {"$cond": [{"$eq": ["$ownership", "rental"]}, {"$ifNull": ["$rental_amount", 0]}, 0]}},
            }},
        ]))
        row = agg[0] if agg else {}
        monthly_data.append({
            "month": ms.strftime("%b %Y"),
            "month_start": ms.isoformat(),
            "entries": row.get("entries", 0),
            "total_hours": round(row.get("total_hours") or 0, 2),
            "rental_cost": round(row.get("rental_cost") or 0, 2),
        })

    # ── Total rental cost from entries (rental_amount field) ──
    rental_cost_agg = list(db.equipment_entries.aggregate([
        {"$match": {**query, "ownership": "rental"}},
        {"$group": {"_id": None,
                    "total": {"$sum": {"$ifNull": ["$rental_amount", 0]}},
                    "total_hours": {"$sum": "$hours"}}},
    ]))
    total_rental_cost = rental_cost_agg[0]["total"] if rental_cost_agg else 0
    total_rental_hours = rental_cost_agg[0]["total_hours"] if rental_cost_agg else 0

    # ── All entries (for table) ──
    entries = [
        _serialize_doc(item)
        for item in db.equipment_entries.find(query).sort("work_date", -1)
    ]
    _enrich_entries_profile_pictures(db, entries)

    # ── Equipment master list (all active equipment details) ──
    eq_filter = {"is_active": True}
    eq_type_filter = eq_type if (eq_type and eq_type != "all") else None
    eq_loc_filter = location if (location and location != "all") else None
    if eq_type_filter:
        eq_filter["equipment_type"] = eq_type_filter
    if eq_loc_filter:
        eq_filter["location"] = eq_loc_filter
    equipment_master = [
        _serialize_doc(e)
        for e in db.equipment_details.find(eq_filter).sort([("equipment_type", 1), ("name", 1)])
    ]

    return jsonify({
        "ok": True,
        "summary": {
            "total_equipment": total_equipment,
            "total_entries": total_entries,
            "owned_entries": owned_entries,
            "rental_entries": rental_entries,
            "total_rental_cost": round(total_rental_cost, 2),
            "total_rental_hours": round(total_rental_hours or 0, 2),
            "anchor_date": anchor.isoformat(),
        },
        "by_type": [
            {
                "name": x.get("_id") or "N/A",
                "entries": x.get("count", 0),
                "total_hours": round(x.get("total_hours") or 0, 2),
                "rental_cost": round(x.get("rental_cost") or 0, 2),
            }
            for x in by_type
        ],
        "by_location": [
            {
                "name": x.get("_id") or "N/A",
                "entries": x.get("count", 0),
                "total_hours": round(x.get("total_hours") or 0, 2),
                "rental_cost": round(x.get("rental_cost") or 0, 2),
            }
            for x in by_location
        ],
        "by_equipment": [
            {
                "equipment_id": x.get("_id") or "",
                "equipment_name": x.get("equipment_name") or "",
                "plate_number": x.get("plate_number") or "",
                "equipment_type": x.get("equipment_type") or "",
                "equipment_location": x.get("equipment_location") or "",
                "ownership": x.get("ownership") or "",
                "entries": x.get("count", 0),
                "total_hours": round(x.get("total_hours") or 0, 2),
                "rental_cost": round(x.get("rental_cost") or 0, 2),
            }
            for x in by_equipment
        ],
        "weekly_trend": weekly_data,
        "monthly_trend": monthly_data,
        "equipment_master": equipment_master,
        "entries": entries,
    })


@app.patch("/api/equipment-entries/<entry_id>")
def update_equipment_entry(entry_id):
    db = get_database()
    try:
        oid = ObjectId(entry_id)
    except Exception:
        return jsonify({"ok": False, "error": "invalid entry id"}), 400

    existing = db.equipment_entries.find_one({"_id": oid})
    if not existing:
        return jsonify({"ok": False, "error": "entry not found"}), 404

    payload = request.get_json(force=True) or {}
    ok, err = _ensure_management_password(payload)
    if not ok:
        return err, 401
    update_fields = {}

    for field in ["work_date", "location", "activity", "operator_name", "time_from", "time_to", "equipment_status"]:
        if field in payload:
            update_fields[field] = (payload[field] or "").strip() or None

    if "hours" in payload:
        try:
            h = float(payload["hours"])
            update_fields["hours"] = h
        except (TypeError, ValueError):
            pass

    if "rental_amount" in payload:
        try:
            update_fields["rental_amount"] = float(payload["rental_amount"])
        except (TypeError, ValueError):
            update_fields["rental_amount"] = None

    if "approval_status" in payload:
        approval_status_value = (payload["approval_status"] or "").strip().lower()
        if approval_status_value not in {"", "pending", "approved", "rejected"}:
            return jsonify({"ok": False, "error": "approval_status must be pending, approved, rejected, or empty"}), 400
        update_fields["approval_status"] = approval_status_value or None
        if approval_status_value != "rejected":
            update_fields["rejection_reason"] = None
        if approval_status_value != "approved":
            update_fields["approved_by"] = None

    if "rejection_reason" in payload:
        reason = (payload.get("rejection_reason") or "").strip()
        update_fields["rejection_reason"] = reason or None

    if "approved_by" in payload:
        update_fields["approved_by"] = (payload.get("approved_by") or "").strip() or None

    update_fields["status"] = "edited_by_management"
    db.equipment_entries.update_one({"_id": oid}, {"$set": update_fields})
    return jsonify({"ok": True, "entry_id": entry_id})


@app.delete("/api/equipment-entries/<entry_id>")
def delete_equipment_entry(entry_id):
    db = get_database()
    try:
        oid = ObjectId(entry_id)
    except Exception:
        return jsonify({"ok": False, "error": "invalid entry id"}), 400

    result = db.equipment_entries.delete_one({"_id": oid})
    if result.deleted_count == 0:
        return jsonify({"ok": False, "error": "entry not found"}), 404
    return jsonify({"ok": True, "entry_id": entry_id})


def _build_equipment_export_workbook(entries):
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment Entries"

    headers = [
        "S.I No", "Date", "Operator", "Company", "Equipment Name", "Plate No.",
        "Type", "Ownership", "Location", "Hours", "Supply Rate (Monthly)",
        "Contract Rate (Hourly)", "Rental Amount (OMR)", "Activity",
        "Approval Status", "Approved By", "Rejection Reason",
    ]
    ws.append(headers)

    hdr_font = Font(bold=True, color="000000")
    hdr_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin

    for idx, e in enumerate(entries, start=1):
        supply_monthly = (e.get("supply_rate") or {}).get("monthly") or ""
        contract_hourly = (e.get("contract_rate") or {}).get("hourly") or ""
        ws.append([
            idx,
            e.get("work_date", ""),
            e.get("operator_name", ""),
            e.get("company_name", ""),
            e.get("equipment_name", ""),
            e.get("plate_number", ""),
            e.get("equipment_type", ""),
            "ENCO Owned" if e.get("ownership") == "owned" else "Rental",
            e.get("location", ""),
            e.get("hours", ""),
            supply_monthly,
            contract_hourly,
            e.get("rental_amount", ""),
            e.get("activity", ""),
            e.get("approval_status", "pending"),
            e.get("approved_by", ""),
            e.get("rejection_reason", ""),
        ])

    max_row = ws.max_row
    for r in range(2, max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = center
            cell.border = thin

    col_widths = [6, 12, 20, 18, 24, 14, 14, 12, 22, 8, 18, 18, 16, 24, 14, 18, 22]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.get("/api/management/export-equipment-excel")
def export_equipment_excel():
    db = get_database()
    query = {}
    work_date = (request.args.get("date") or "").strip()
    location = (request.args.get("location") or "").strip()
    eq_type = (request.args.get("type") or "").strip()
    period = _normalize_period_filter(request.args.get("period") or "all")
    anchor_date_raw = (request.args.get("anchor_date") or "").strip()

    try:
        anchor = datetime.strptime(anchor_date_raw, "%Y-%m-%d").date() if anchor_date_raw else date.today()
    except ValueError:
        anchor = date.today()

    if work_date:
        query["work_date"] = work_date
    elif period != "all":
        start = end = anchor
        if period == "day":
            start = end = anchor
        elif period == "week":
            start = anchor - timedelta(days=anchor.weekday())
            end = start + timedelta(days=6)
        elif period == "month":
            start = anchor.replace(day=1)
            next_m = (start.replace(month=start.month + 1, day=1) if start.month < 12
                      else start.replace(year=start.year + 1, month=1, day=1))
            end = next_m - timedelta(days=1)
        elif period == "quarter":
            qsm = ((anchor.month - 1) // 3) * 3 + 1
            start = anchor.replace(month=qsm, day=1)
            next_q = (start.replace(month=qsm + 3, day=1)
                      if qsm <= 9 else start.replace(year=start.year + 1, month=1, day=1))
            end = next_q - timedelta(days=1)
        elif period == "year":
            start = anchor.replace(month=1, day=1)
            end = anchor.replace(month=12, day=31)
        query["work_date"] = {"$gte": start.isoformat(), "$lte": end.isoformat()}

    if location and location != "all":
        query["location"] = location
    if eq_type and eq_type != "all":
        query["equipment_type"] = eq_type

    entries = list(db.equipment_entries.find(query).sort("work_date", -1))
    buf = _build_equipment_export_workbook(entries)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"equipment_entries_{anchor.isoformat()}.xlsx",
    )


# ─────────────────────────────────────────────────────────────────────────────
# EQUIPMENT – helpers
# ─────────────────────────────────────────────────────────────────────────────

EQUIPMENT_DATA = [
    # ENCO - INDUSTRIAL AREA
    {"name": "40T TADANO CRAWLER CRANE", "plate_number": "", "equipment_type": "Crane", "location": "ENCO - INDUSTRIAL AREA", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}},
    {"name": "80T GROVE TELESCOPIC CRANE", "plate_number": "", "equipment_type": "Crane", "location": "ENCO - INDUSTRIAL AREA", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 0, "daily": 0, "weekly": 4880, "monthly": 12200}},
    {"name": "100T SANY TELESCOPIC CRANE", "plate_number": "", "equipment_type": "Crane", "location": "ENCO - INDUSTRIAL AREA", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 0, "daily": 0, "weekly": 4880, "monthly": 12200}},
    {"name": "14M HAULOTE MANLIFT", "plate_number": "", "equipment_type": "Manlift", "location": "ENCO - INDUSTRIAL AREA", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 11, "daily": 95, "weekly": 518, "monthly": 1550}},
    {"name": "41M JLG MANLIFT", "plate_number": "", "equipment_type": "Manlift", "location": "ENCO - INDUSTRIAL AREA", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 21, "daily": 182, "weekly": 1000, "monthly": 3000}},
    {"name": "7T HYSTER FORKLIFT", "plate_number": "", "equipment_type": "Forklift", "location": "ENCO - INDUSTRIAL AREA", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 15, "daily": 135, "weekly": 735, "monthly": 2200}},
    {"name": "22 SEAT BUS", "plate_number": "2746MB", "equipment_type": "Bus", "location": "ENCO - INDUSTRIAL AREA", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}},
    {"name": "FLATBED TRAILER", "plate_number": "", "equipment_type": "Trailer", "location": "ENCO - INDUSTRIAL AREA", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 20, "daily": 175, "weekly": 940, "monthly": 2800}},
    {"name": "KAESER COMPRESSOR", "plate_number": "", "equipment_type": "Compressor", "location": "ENCO - INDUSTRIAL AREA", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}},
    # CONTRACTOR PARKING AREA
    {"name": "70T TELESCOPIC CRANE", "plate_number": "6427MH", "equipment_type": "Crane", "location": "CONTRACTOR PARKING AREA", "ownership": "rental", "supply_rate": {"hourly": 0, "daily": 0, "weekly": 0, "monthly": 2400}, "contract_rate": {"hourly": 0, "daily": 454, "weekly": 2500, "monthly": 7500}},
    {"name": "130T CRAWLER CRANE", "plate_number": "2854BA", "equipment_type": "Crane", "location": "CONTRACTOR PARKING AREA", "ownership": "rental", "supply_rate": {"hourly": 0, "daily": 0, "weekly": 0, "monthly": 6800}, "contract_rate": {"hourly": 0, "daily": 0, "weekly": 6730, "monthly": 16380}},
    {"name": "22M HAULOTTE MANLIFT", "plate_number": "2164847", "equipment_type": "Manlift", "location": "CONTRACTOR PARKING AREA", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 15, "daily": 135, "weekly": 740, "monthly": 2200}},
    {"name": "TOWER LIGHT", "plate_number": "2702250313", "equipment_type": "Tower Light", "location": "CONTRACTOR PARKING AREA", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 8.6, "daily": 34, "weekly": 185, "monthly": 550}},
    {"name": "7.2KW TOWER LIGHT", "plate_number": "2702250311", "equipment_type": "Tower Light", "location": "CONTRACTOR PARKING AREA", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 8.6, "daily": 34, "weekly": 185, "monthly": 550}},
    {"name": "PICK-UP", "plate_number": "4265BM", "equipment_type": "Pick-Up", "location": "CONTRACTOR PARKING AREA", "ownership": "rental", "supply_rate": {"hourly": 0, "daily": 0, "weekly": 0, "monthly": 300}, "contract_rate": {"hourly": 0, "daily": 60, "weekly": 480, "monthly": 1200}},
    {"name": "12M SCISSOR LIFT", "plate_number": "20251128-01", "equipment_type": "Scissor Lift", "location": "CONTRACTOR PARKING AREA", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 10, "daily": 85, "weekly": 470, "monthly": 1400}},
    {"name": "12M SCISSOR LIFT", "plate_number": "20251128-02", "equipment_type": "Scissor Lift", "location": "CONTRACTOR PARKING AREA", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 10, "daily": 85, "weekly": 470, "monthly": 1400}},
    {"name": "12M SCISSOR LIFT", "plate_number": "20251128-03", "equipment_type": "Scissor Lift", "location": "CONTRACTOR PARKING AREA", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 10, "daily": 85, "weekly": 470, "monthly": 1400}},
    {"name": "55 SEAT BUS", "plate_number": "", "equipment_type": "Bus", "location": "CONTRACTOR PARKING AREA", "ownership": "rental", "supply_rate": {"hourly": 0, "daily": 0, "weekly": 0, "monthly": 600}, "contract_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}},
    # YARD 2
    {"name": "100T TELESCOPIC CRANE", "plate_number": "3988DS", "equipment_type": "Crane", "location": "YARD 2", "ownership": "rental", "supply_rate": {"hourly": 0, "daily": 0, "weekly": 0, "monthly": 5000}, "contract_rate": {"hourly": 0, "daily": 0, "weekly": 4880, "monthly": 12200}},
    {"name": "150T CRAWLER CRANE", "plate_number": "3285LH", "equipment_type": "Crane", "location": "YARD 2", "ownership": "rental", "supply_rate": {"hourly": 0, "daily": 0, "weekly": 0, "monthly": 7000}, "contract_rate": {"hourly": 0, "daily": 0, "weekly": 7480, "monthly": 18700}},
    {"name": "22M LCMG MANLIFT", "plate_number": "LWJAB202KS3720001", "equipment_type": "Manlift", "location": "YARD 2", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 15, "daily": 135, "weekly": 740, "monthly": 2200}},
    {"name": "28M LCMG MANLIFT", "plate_number": "LWJAB262ES3739913", "equipment_type": "Manlift", "location": "YARD 2", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 15, "daily": 135, "weekly": 740, "monthly": 2200}},
    {"name": "22KVA GENERATOR", "plate_number": "", "equipment_type": "Generator", "location": "YARD 2", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 7.85, "daily": 31, "weekly": 170, "monthly": 500}},
    # YARD 3
    {"name": "70T CRAWLER CRANE", "plate_number": "6901YS", "equipment_type": "Crane", "location": "YARD 3", "ownership": "rental", "supply_rate": {"hourly": 0, "daily": 0, "weekly": 0, "monthly": 3800}, "contract_rate": {"hourly": 0, "daily": 0, "weekly": 3950, "monthly": 10950}},
    {"name": "150T CRAWLER CRANE", "plate_number": "3286LH", "equipment_type": "Crane", "location": "YARD 3", "ownership": "rental", "supply_rate": {"hourly": 0, "daily": 0, "weekly": 0, "monthly": 7000}, "contract_rate": {"hourly": 0, "daily": 0, "weekly": 7480, "monthly": 18700}},
    {"name": "40KVA CUMMINS GENERATOR", "plate_number": "", "equipment_type": "Generator", "location": "YARD 3", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 8.6, "daily": 34, "weekly": 185, "monthly": 550}},
    {"name": "FLATBED TRAILER", "plate_number": "3576MD", "equipment_type": "Trailer", "location": "YARD 3", "ownership": "rental", "supply_rate": {"hourly": 0, "daily": 0, "weekly": 0, "monthly": 1400}, "contract_rate": {"hourly": 20, "daily": 175, "weekly": 940, "monthly": 2800}},
    {"name": "PICK-UP", "plate_number": "2819DA", "equipment_type": "Pick-Up", "location": "YARD 3", "ownership": "rental", "supply_rate": {"hourly": 0, "daily": 0, "weekly": 0, "monthly": 300}, "contract_rate": {"hourly": 0, "daily": 60, "weekly": 480, "monthly": 1200}},
    # TSF
    {"name": "70T TELESCOPIC CRANE", "plate_number": "3155MH", "equipment_type": "Crane", "location": "TSF", "ownership": "rental", "supply_rate": {"hourly": 0, "daily": 0, "weekly": 0, "monthly": 2400}, "contract_rate": {"hourly": 0, "daily": 454, "weekly": 2500, "monthly": 7500}},
    {"name": "70T CRAWLER CRANE", "plate_number": "9068HW", "equipment_type": "Crane", "location": "TSF", "ownership": "rental", "supply_rate": {"hourly": 0, "daily": 0, "weekly": 0, "monthly": 3800}, "contract_rate": {"hourly": 0, "daily": 0, "weekly": 3950, "monthly": 10950}},
    {"name": "100T TELESCOPIC CRANE", "plate_number": "9480ML", "equipment_type": "Crane", "location": "TSF", "ownership": "rental", "supply_rate": {"hourly": 0, "daily": 0, "weekly": 0, "monthly": 6800}, "contract_rate": {"hourly": 0, "daily": 0, "weekly": 6730, "monthly": 16380}},
    {"name": "7T HI-UP", "plate_number": "1077DS", "equipment_type": "Forklift", "location": "TSF", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 15, "daily": 135, "weekly": 720, "monthly": 2150}},
    {"name": "14T FORKLIFT", "plate_number": "2088LD", "equipment_type": "Forklift", "location": "TSF", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 29.25, "daily": 263.5, "weekly": 1445, "monthly": 4335}},
    {"name": "4T JCB TELEHANDLER", "plate_number": "5630YW", "equipment_type": "Telehandler", "location": "TSF", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 15, "daily": 135, "weekly": 735, "monthly": 2200}},
    {"name": "FLATBED TRAILER", "plate_number": "7465BM", "equipment_type": "Trailer", "location": "TSF", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 20, "daily": 175, "weekly": 940, "monthly": 2800}},
    {"name": "RENAULT LOWBED TRAILER", "plate_number": "2939ML", "equipment_type": "Trailer", "location": "TSF", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 40, "daily": 350, "weekly": 1880, "monthly": 5600}},
    {"name": "12M SCISSOR LIFT", "plate_number": "20251128-04", "equipment_type": "Scissor Lift", "location": "TSF", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 10, "daily": 85, "weekly": 470, "monthly": 1400}},
    {"name": "300/330KVA CUMMINS GENERATOR", "plate_number": "", "equipment_type": "Generator", "location": "TSF", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}},
    {"name": "35/40KVA CUMMINS GENERATOR", "plate_number": "", "equipment_type": "Generator", "location": "TSF", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 8.6, "daily": 34, "weekly": 185, "monthly": 550}},
    {"name": "6.5KW TOWER LIGHT LED", "plate_number": "LD1000JN202510071", "equipment_type": "Tower Light", "location": "TSF", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 8.6, "daily": 34, "weekly": 185, "monthly": 550}},
    {"name": "7.2KW TOWER LIGHT", "plate_number": "2702250306", "equipment_type": "Tower Light", "location": "TSF", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 8.6, "daily": 34, "weekly": 185, "monthly": 550}},
    {"name": "7.2KW TOWER LIGHT", "plate_number": "2025121001", "equipment_type": "Tower Light", "location": "TSF", "ownership": "owned", "supply_rate": {"hourly": None, "daily": None, "weekly": None, "monthly": None}, "contract_rate": {"hourly": 8.6, "daily": 34, "weekly": 185, "monthly": 550}},
    {"name": "PICK-UP", "plate_number": "569WA", "equipment_type": "Pick-Up", "location": "TSF", "ownership": "rental", "supply_rate": {"hourly": 0, "daily": 0, "weekly": 0, "monthly": 300}, "contract_rate": {"hourly": 0, "daily": 60, "weekly": 480, "monthly": 1200}},
    {"name": "PICK-UP", "plate_number": "15BM", "equipment_type": "Pick-Up", "location": "TSF", "ownership": "rental", "supply_rate": {"hourly": 0, "daily": 0, "weekly": 0, "monthly": 300}, "contract_rate": {"hourly": 0, "daily": 60, "weekly": 480, "monthly": 1200}},
]


def _equipment_display_name(eq):
    """Return the display name shown in dropdowns (name + plate if any)."""
    plate = (eq.get("plate_number") or "").strip()
    name = (eq.get("name") or "").strip()
    return f"{name} [{plate}]" if plate else name


# ─────────────────────────────────────────────────────────────────────────────
# EQUIPMENT – routes
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/setup/init-equipment")
def init_equipment():
    """Seed equipment_details collection from built-in data."""
    try:
        db = get_database()
        db.equipment_details.create_index(
            [("name", 1), ("plate_number", 1), ("location", 1)], unique=True
        )
        db.equipment_entries.create_index(
            [("equipment_id", 1), ("work_date", 1)]
        )
        inserted = 0
        updated = 0
        for eq in EQUIPMENT_DATA:
            result = db.equipment_details.update_one(
                {
                    "name": eq["name"],
                    "plate_number": eq["plate_number"],
                    "location": eq["location"],
                },
                {
                    "$set": {
                        **eq,
                        "is_active": True,
                        "display_name": _equipment_display_name(eq),
                    }
                },
                upsert=True,
            )
            if result.upserted_id:
                inserted += 1
            else:
                updated += 1
        return jsonify({"ok": True, "inserted": inserted, "updated": updated})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.get("/api/equipment/types")
def get_equipment_types():
    """Return distinct equipment types."""
    db = get_database()
    types = db.equipment_details.distinct("equipment_type", {"is_active": True})
    return jsonify({"ok": True, "types": sorted(types)})


@app.get("/api/equipment/list")
def get_equipment_list():
    """Return equipment list, optionally filtered by type and/or location."""
    db = get_database()
    query = {"is_active": True}
    eq_type = (request.args.get("type") or "").strip()
    location = (request.args.get("location") or "").strip()
    if eq_type:
        query["equipment_type"] = eq_type
    if location:
        query["location"] = location

    work_date = (request.args.get("date") or date.today().isoformat()).strip()

    # Find equipment already used today
    locked_ids = set()
    for entry in db.equipment_entries.find(
        {"work_date": work_date, "status": {"$nin": ["cancelled"]}},
        {"equipment_id": 1},
    ):
        eid = entry.get("equipment_id")
        if eid:
            locked_ids.add(str(eid))

    equipment = []
    for eq in db.equipment_details.find(query).sort([("equipment_type", 1), ("name", 1)]):
        eq_id = str(eq["_id"])
        eq["_id"] = eq_id
        eq["is_locked"] = eq_id in locked_ids
        equipment.append(eq)

    return jsonify({"ok": True, "equipment": equipment})


@app.get("/api/equipment/<equipment_id>")
def get_equipment_detail(equipment_id):
    """Return a single equipment document by ID."""
    db = get_database()
    try:
        oid = ObjectId(equipment_id)
    except Exception:
        return jsonify({"ok": False, "error": "invalid equipment id"}), 400
    eq = db.equipment_details.find_one({"_id": oid})
    if not eq:
        return jsonify({"ok": False, "error": "equipment not found"}), 404
    eq["_id"] = str(eq["_id"])
    return jsonify({"ok": True, "equipment": eq})


EQUIPMENT_STATUS_CHOICES = [
    "Working", "Under Maintenance", "Transporting",
    "Disassembly", "Assembly", "Inspection", "Expired"
]

@app.post("/api/equipment-entries")
def create_equipment_entry():
    payload = request.get_json(force=True) or {}

    civil_id = (payload.get("civil_id") or "").strip()
    company_name = (payload.get("company_name") or "").strip()
    operator_name = (payload.get("operator_name") or "").strip()
    equipment_id = (payload.get("equipment_id") or "").strip()
    work_date = (payload.get("work_date") or date.today().isoformat()).strip()
    time_from = (payload.get("time_from") or "").strip()
    time_to = (payload.get("time_to") or "").strip()
    location = (payload.get("location") or "").strip()
    activity = (payload.get("activity") or "").strip()
    equipment_status = (payload.get("equipment_status") or "").strip()
    rental_amount = payload.get("rental_amount")

    missing = [k for k, v in {
        "civil_id": civil_id,
        "equipment_id": equipment_id,
        "work_date": work_date,
        "location": location,
        "equipment_status": equipment_status,
    }.items() if not v]
    if missing:
        return jsonify({"ok": False, "error": f"missing required fields: {', '.join(missing)}"}), 400

    def _parse_hhmm(value):
        if not value:
            return None
        parts = value.split(":")
        if len(parts) != 2:
            return None
        try:
            hh, mm = int(parts[0]), int(parts[1])
        except ValueError:
            return None
        if hh < 0 or hh > 23 or mm < 0 or mm > 59:
            return None
        return hh * 60 + mm

    hours_value = None
    if time_from and time_to:
        s = _parse_hhmm(time_from)
        e = _parse_hhmm(time_to)
        if s is None or e is None:
            return jsonify({"ok": False, "error": "time must be in HH:MM format"}), 400
        if e <= s:
            return jsonify({"ok": False, "error": "to time must be after from time"}), 400
        hours_value = round((e - s) / 60.0, 2)

    db = get_database()

    try:
        oid = ObjectId(equipment_id)
    except Exception:
        return jsonify({"ok": False, "error": "invalid equipment_id"}), 400

    eq = db.equipment_details.find_one({"_id": oid, "is_active": True})
    if not eq:
        return jsonify({"ok": False, "error": "equipment not found"}), 404

    # Block duplicate timing: same civil_id + same date + same from/to time
    if time_from and time_to:
        duplicate_timing = db.equipment_entries.find_one({
            "civil_id": civil_id,
            "work_date": work_date,
            "time_from": time_from,
            "time_to": time_to,
            "status": {"$nin": ["cancelled"]},
        })
        if duplicate_timing:
            return jsonify({
                "ok": False,
                "error": "duplicate_timing",
                "message": f"You already have an entry from {time_from} to {time_to} today. Please choose a different time.",
            }), 409

    # Check if equipment is already locked for this date (another operator)
    existing = db.equipment_entries.find_one(
        {"equipment_id": equipment_id, "work_date": work_date, "status": {"$nin": ["cancelled"]}}
    )
    if existing:
        return jsonify({"ok": False, "error": "This equipment is already logged for today by another operator"}), 409

    # Validate civil_id is a recognised worker
    worker = db.worker_details.find_one({"civil_id": civil_id, "is_active": True})
    if not worker and civil_id != DEVELOPER_MASTER_CIVIL_ID:
        return jsonify({"ok": False, "error": "worker profile not found"}), 404

    rental_amount_value = None
    if rental_amount is not None and rental_amount != "":
        try:
            rental_amount_value = float(rental_amount)
        except (TypeError, ValueError):
            pass

    result = db.equipment_entries.insert_one({
        "civil_id": civil_id,
        "company_name": company_name,
        "operator_name": operator_name,
        "equipment_id": equipment_id,
        "equipment_name": eq.get("name", ""),
        "plate_number": eq.get("plate_number", ""),
        "equipment_type": eq.get("equipment_type", ""),
        "equipment_location": eq.get("location", ""),
        "ownership": eq.get("ownership", ""),
        "work_date": work_date,
        "location": location,
        "equipment_status": equipment_status,
        "time_from": time_from or None,
        "time_to": time_to or None,
        "hours": hours_value,
        "activity": activity,
        "supply_rate": eq.get("supply_rate", {}),
        "contract_rate": eq.get("contract_rate", {}),
        "rental_amount": rental_amount_value,
        "status": "submitted",
        "created_at": datetime.utcnow().isoformat(),
    })
    return jsonify({"ok": True, "id": str(result.inserted_id)})


@app.get("/api/equipment-entries/today-status")
def equipment_entry_today_status():
    """Check whether an operator already has an equipment entry for today."""
    civil_id = (request.args.get("civil_id") or "").strip()
    if not civil_id:
        return jsonify({"ok": False, "error": "civil_id required"}), 400
    work_date = date.today().isoformat()
    db = get_database()
    existing = db.equipment_entries.find_one(
        {"civil_id": civil_id, "work_date": work_date, "status": {"$nin": ["cancelled"]}},
        {"time_from": 1, "time_to": 1, "equipment_name": 1, "plate_number": 1,
         "location": 1, "equipment_status": 1, "hours": 1},
    )
    if existing:
        existing["_id"] = str(existing["_id"])
        return jsonify({"ok": True, "submitted": True, "entry": existing})
    return jsonify({"ok": True, "submitted": False})


@app.get("/api/equipment-entries")
def list_equipment_entries():
    """Return equipment entries with optional filters."""
    db = get_database()
    query = {}
    work_date = (request.args.get("date") or "").strip()
    location = (request.args.get("location") or "").strip()
    if work_date:
        query["work_date"] = work_date
    if location:
        query["location"] = location

    entries = []
    for e in db.equipment_entries.find(query).sort("created_at", -1):
        e["_id"] = str(e["_id"])
        entries.append(e)
    _enrich_entries_profile_pictures(db, entries)
    return jsonify({"ok": True, "entries": entries})


# ─────────────────────────────────────────────────────────────────────────────
# PROGRESS TRACKING – constants and helpers
# ─────────────────────────────────────────────────────────────────────────────

PROGRESS_PROCESS_TYPES = [
    "material_received",
    "material_unloading",
    "pre_assembly",
    "erection",
    "alignment_torquing",
    "painting_piping",
    "final_inspection",
]

PROGRESS_PROCESS_LABELS = {
    "material_received": "Material Received",
    "material_unloading": "Material Unloading",
    "pre_assembly": "Pre-Assembly",
    "erection": "Erection",
    "alignment_torquing": "Alignment & Torquing",
    "painting_piping": "Painting & Piping",
    "final_inspection": "Final Inspection",
}

# Baseline data extracted from Progress Update.xlsx (SUMMARY DATA sheet)
# prev = previous week cumulative entries, this = this week additions
PROGRESS_ITEM_TAGS_DATA = [
    {
        "tag": "CT-1010-02", "gross_weight": 855.4132770000027,
        "material_received": 932,
        "prev": {"material_unloading": 217.2134, "pre_assembly": 111.7539, "erection": 103.3856, "alignment_torquing": 40.9461, "painting_piping": 0, "final_inspection": 0},
        "this": {"material_unloading": 256.5870, "pre_assembly": 127.7912, "erection": 88.3952, "alignment_torquing": 38.5494, "painting_piping": 0, "final_inspection": 0},
    },
    {
        "tag": "CT-1010-20", "gross_weight": 714.4082303761917,
        "material_received": 227,
        "prev": {"material_unloading": 0, "pre_assembly": 0, "erection": 0, "alignment_torquing": 0, "painting_piping": 0, "final_inspection": 0},
        "this": {"material_unloading": 0, "pre_assembly": 0, "erection": 0, "alignment_torquing": 0, "painting_piping": 0, "final_inspection": 0},
    },
    {
        "tag": "CT-1010-21", "gross_weight": 247.7419000000002,
        "material_received": 0,
        "prev": {"material_unloading": 0, "pre_assembly": 0, "erection": 0, "alignment_torquing": 0, "painting_piping": 0, "final_inspection": 0},
        "this": {"material_unloading": 0, "pre_assembly": 0, "erection": 0, "alignment_torquing": 0, "painting_piping": 0, "final_inspection": 0},
    },
    {
        "tag": "CT-1010-27", "gross_weight": 106.45699999999998,
        "material_received": 0,
        "prev": {"material_unloading": 0, "pre_assembly": 0, "erection": 0, "alignment_torquing": 0, "painting_piping": 0, "final_inspection": 0},
        "this": {"material_unloading": 0, "pre_assembly": 0, "erection": 0, "alignment_torquing": 0, "painting_piping": 0, "final_inspection": 0},
    },
    {
        "tag": "TR-1010-01", "gross_weight": 147.42659999999984,
        "material_received": 337,
        "prev": {"material_unloading": 0, "pre_assembly": 0, "erection": 0, "alignment_torquing": 0, "painting_piping": 0, "final_inspection": 0},
        "this": {"material_unloading": 0, "pre_assembly": 0, "erection": 0, "alignment_torquing": 0, "painting_piping": 0, "final_inspection": 0},
    },
    {
        "tag": "TR-1010-21", "gross_weight": 1460.99638000001,
        "material_received": 2175,
        "prev": {"material_unloading": 0, "pre_assembly": 0, "erection": 0, "alignment_torquing": 0, "painting_piping": 0, "final_inspection": 0},
        "this": {"material_unloading": 56.3064, "pre_assembly": 46.0240, "erection": 0, "alignment_torquing": 0, "painting_piping": 0, "final_inspection": 0},
    },
    {
        "tag": "TR-1010-26", "gross_weight": 7.558200000000001,
        "material_received": 2932,
        "prev": {"material_unloading": 293.0081, "pre_assembly": 241.3823, "erection": 0, "alignment_torquing": 0, "painting_piping": 0, "final_inspection": 0},
        "this": {"material_unloading": -293.0081, "pre_assembly": -241.3823, "erection": 0, "alignment_torquing": 0, "painting_piping": 0, "final_inspection": 0},
    },
    {
        "tag": "TR-1010-20", "gross_weight": 300.9615000000003,
        "material_received": 396,
        "prev": {"material_unloading": 0, "pre_assembly": 0, "erection": 0, "alignment_torquing": 0, "painting_piping": 0, "final_inspection": 0},
        "this": {"material_unloading": 0, "pre_assembly": 0, "erection": 0, "alignment_torquing": 0, "painting_piping": 0, "final_inspection": 0},
    },
    {
        "tag": "TR-1010-27", "gross_weight": 211.6867,
        "material_received": 96,
        "prev": {"material_unloading": 0, "pre_assembly": 0, "erection": 0, "alignment_torquing": 0, "painting_piping": 0, "final_inspection": 0},
        "this": {"material_unloading": 0, "pre_assembly": 0, "erection": 0, "alignment_torquing": 0, "painting_piping": 0, "final_inspection": 0},
    },
]


# ─────────────────────────────────────────────────────────────────────────────
# PROGRESS TRACKING – routes
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/setup/init-progress")
def init_progress():
    """Seed progress_item_tags and baseline progress_entries from built-in Excel data."""
    try:
        db = get_database()
        db.progress_item_tags.create_index([("tag", 1)], unique=True)
        db.progress_entries.create_index([("item_tag", 1), ("process_type", 1), ("date", 1)])

        inserted_tags = 0
        for tag_data in PROGRESS_ITEM_TAGS_DATA:
            result = db.progress_item_tags.update_one(
                {"tag": tag_data["tag"]},
                {"$setOnInsert": {
                    "tag": tag_data["tag"],
                    "gross_weight": tag_data["gross_weight"],
                    "created_at": datetime.utcnow().isoformat(),
                }},
                upsert=True,
            )
            if result.upserted_id:
                inserted_tags += 1

        # Seed baseline entries – previous week: 2026-04-07 (W15), this week: 2026-04-14 (W16)
        prev_date, prev_week, this_date, this_week, seed_year = "2026-04-07", 15, "2026-04-14", 16, 2026
        inserted_entries = 0

        for tag_data in PROGRESS_ITEM_TAGS_DATA:
            tag = tag_data["tag"]

            def _seed(proc, val, entry_date, wk):
                nonlocal inserted_entries
                # Skip zero or negative values – negative entries are Excel corrections
                # that cancel previous data; seeding them would produce negative bars
                if not val or float(val) <= 0:
                    return
                exists = db.progress_entries.find_one(
                    {"item_tag": tag, "process_type": proc, "date": entry_date, "is_baseline": True}
                )
                if not exists:
                    db.progress_entries.insert_one({
                        "date": entry_date, "week_number": wk, "year": seed_year,
                        "item_tag": tag, "process_type": proc,
                        "tonnage": round(float(val), 4),
                        "description": "Baseline seed data",
                        "entered_by": "System",
                        "is_baseline": True,
                        "created_at": datetime.utcnow().isoformat(),
                    })
                    inserted_entries += 1

            mr = tag_data.get("material_received", 0)
            _seed("material_received", mr, prev_date, prev_week)

            for proc, val in tag_data["prev"].items():
                _seed(proc, val, prev_date, prev_week)
            for proc, val in tag_data["this"].items():
                _seed(proc, val, this_date, this_week)

        return jsonify({"ok": True, "inserted_tags": inserted_tags, "inserted_entries": inserted_entries})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.get("/api/progress/item-tags")
def get_progress_item_tags():
    """Return all item tags for dropdowns."""
    db = get_database()
    tags = [
        {"tag": t["tag"], "gross_weight": round(float(t.get("gross_weight") or 0), 4)}
        for t in db.progress_item_tags.find({}, {"tag": 1, "gross_weight": 1}).sort("tag", 1)
    ]
    return jsonify({"ok": True, "item_tags": tags})


@app.post("/api/management/progress-entries")
def create_progress_entry():
    payload = request.get_json(force=True) or {}
    ok, err = _ensure_management_password(payload)
    if not ok:
        return err, 401

    item_tag = (payload.get("item_tag") or "").strip()
    process_type = (payload.get("process_type") or "").strip()
    entry_date = (payload.get("date") or date.today().isoformat()).strip()
    entered_by = (payload.get("entered_by") or "").strip()
    description = (payload.get("description") or "").strip()

    if not item_tag or not process_type:
        return jsonify({"ok": False, "error": "item_tag and process_type are required"}), 400
    if process_type not in PROGRESS_PROCESS_TYPES:
        return jsonify({"ok": False, "error": f"process_type must be one of: {', '.join(PROGRESS_PROCESS_TYPES)}"}), 400

    try:
        tonnage = float(payload.get("tonnage", 0))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "tonnage must be a number"}), 400

    db = get_database()
    if not db.progress_item_tags.find_one({"tag": item_tag}):
        return jsonify({"ok": False, "error": "item_tag not found"}), 404

    try:
        d = datetime.strptime(entry_date, "%Y-%m-%d").date()
        iso_cal = d.isocalendar()
        week_number = iso_cal[1]
        yr = iso_cal[0]
    except ValueError:
        return jsonify({"ok": False, "error": "date must be YYYY-MM-DD"}), 400

    result = db.progress_entries.insert_one({
        "date": entry_date, "week_number": week_number, "year": yr,
        "item_tag": item_tag, "process_type": process_type,
        "tonnage": tonnage,
        "description": description or None,
        "entered_by": entered_by or None,
        "is_baseline": False,
        "created_at": datetime.utcnow().isoformat(),
    })
    return jsonify({"ok": True, "id": str(result.inserted_id)})


@app.get("/api/management/progress-entries")
def list_progress_entries():
    db = get_database()
    query = {}
    item_tag = (request.args.get("item_tag") or "").strip()
    process_type = (request.args.get("process_type") or "").strip()
    start_date = (request.args.get("start_date") or "").strip()
    end_date = (request.args.get("end_date") or "").strip()
    include_baseline = request.args.get("include_baseline", "true").strip().lower() == "true"

    if item_tag and item_tag != "all":
        query["item_tag"] = item_tag
    if process_type and process_type != "all":
        query["process_type"] = process_type
    if start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        query["date"] = {"$gte": start_date}
    elif end_date:
        query["date"] = {"$lte": end_date}
    if not include_baseline:
        query["is_baseline"] = {"$ne": True}

    entries = [_serialize_doc(e) for e in db.progress_entries.find(query).sort([("date", -1), ("_id", -1)])]
    return jsonify({"ok": True, "entries": entries})


@app.patch("/api/management/progress-entries/<entry_id>")
def update_progress_entry(entry_id):
    payload = request.get_json(force=True) or {}
    ok, err = _ensure_management_password(payload)
    if not ok:
        return err, 401
    db = get_database()
    try:
        oid = ObjectId(entry_id)
    except Exception:
        return jsonify({"ok": False, "error": "invalid entry id"}), 400
    existing = db.progress_entries.find_one({"_id": oid})
    if not existing:
        return jsonify({"ok": False, "error": "progress entry not found"}), 404

    upd = {}
    if "item_tag" in payload:
        upd["item_tag"] = (payload["item_tag"] or "").strip()
    if "process_type" in payload:
        pt = (payload["process_type"] or "").strip()
        if pt and pt not in PROGRESS_PROCESS_TYPES:
            return jsonify({"ok": False, "error": "invalid process_type"}), 400
        upd["process_type"] = pt
    if "date" in payload:
        ed = (payload["date"] or "").strip()
        try:
            d2 = datetime.strptime(ed, "%Y-%m-%d").date()
            ic = d2.isocalendar()
            upd["date"] = ed
            upd["week_number"] = ic[1]
            upd["year"] = ic[0]
        except ValueError:
            return jsonify({"ok": False, "error": "date must be YYYY-MM-DD"}), 400
    if "tonnage" in payload:
        try:
            upd["tonnage"] = float(payload["tonnage"])
        except (TypeError, ValueError):
            return jsonify({"ok": False, "error": "tonnage must be a number"}), 400
    if "description" in payload:
        upd["description"] = (payload["description"] or "").strip() or None
    if "entered_by" in payload:
        upd["entered_by"] = (payload["entered_by"] or "").strip() or None

    upd["updated_at"] = datetime.utcnow().isoformat()
    db.progress_entries.update_one({"_id": oid}, {"$set": upd})
    return jsonify({"ok": True, "id": entry_id})


@app.delete("/api/management/progress-entries/<entry_id>")
def delete_progress_entry(entry_id):
    payload = request.get_json(force=True) or {}
    ok, err = _ensure_management_password(payload)
    if not ok:
        return err, 401
    db = get_database()
    try:
        oid = ObjectId(entry_id)
    except Exception:
        return jsonify({"ok": False, "error": "invalid entry id"}), 400
    result = db.progress_entries.delete_one({"_id": oid})
    if result.deleted_count == 0:
        return jsonify({"ok": False, "error": "progress entry not found"}), 404
    return jsonify({"ok": True, "id": entry_id})


@app.get("/api/management/progress-dashboard")
def progress_dashboard():
    """Full progress analytics: per-tag summary, weekly trend, prev vs this week."""
    db = get_database()

    # Optional date-range filter (YYYY-MM format from frontend month pickers)
    from_month = (request.args.get("from_month") or "").strip()   # e.g. "2026-01"
    to_month   = (request.args.get("to_month")   or "").strip()   # e.g. "2026-04"

    entry_query = {}
    if from_month:
        # Convert "YYYY-MM" → first day of that month
        try:
            fm = datetime.strptime(from_month, "%Y-%m").date()
            entry_query.setdefault("date", {})["$gte"] = fm.isoformat()
        except ValueError:
            pass
    if to_month:
        # Convert "YYYY-MM" → last day of that month
        try:
            tm_d = datetime.strptime(to_month, "%Y-%m").date()
            # last day = first day of next month minus 1
            if tm_d.month == 12:
                last_day = tm_d.replace(year=tm_d.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                last_day = tm_d.replace(month=tm_d.month + 1, day=1) - timedelta(days=1)
            entry_query.setdefault("date", {})["$lte"] = last_day.isoformat()
        except ValueError:
            pass

    item_tags_docs = list(db.progress_item_tags.find({}).sort("tag", 1))
    all_entries = list(db.progress_entries.find(entry_query))

    # Build cumulative tonnage per (tag, process_type)
    totals_map = {}
    for e in all_entries:
        tag = e.get("item_tag", "")
        pt = e.get("process_type", "")
        t = float(e.get("tonnage") or 0)
        if tag not in totals_map:
            totals_map[tag] = {}
        totals_map[tag][pt] = totals_map[tag].get(pt, 0) + t

    # Per-tag summary
    item_tag_summary = []
    for td in item_tags_docs:
        tag = td["tag"]
        gw = float(td.get("gross_weight") or 0)
        tm = totals_map.get(tag, {})
        processes = {}
        for pt in PROGRESS_PROCESS_TYPES:
            tn = round(tm.get(pt, 0), 4)
            pct = round((tn / gw * 100) if gw > 0 else 0, 2)
            processes[pt] = {"tonnage": tn, "percent": pct}
        item_tag_summary.append({"tag": tag, "gross_weight": round(gw, 4), "processes": processes})

    # Grand totals
    total_gw = round(sum(float(td.get("gross_weight") or 0) for td in item_tags_docs), 4)
    grand_totals = {}
    for pt in PROGRESS_PROCESS_TYPES:
        tn = round(sum(totals_map.get(td["tag"], {}).get(pt, 0) for td in item_tags_docs), 4)
        pct = round((tn / total_gw * 100) if total_gw > 0 else 0, 2)
        grand_totals[pt] = {"tonnage": tn, "percent": pct}

    # Weekly trend – 8 weeks ending today
    today = date.today()
    wstart = today - timedelta(days=today.weekday())
    weekly_trend = []
    for i in range(7, -1, -1):
        ws = wstart - timedelta(weeks=i)
        we = ws + timedelta(days=6)
        iso_w = ws.isocalendar()[1]
        label = f"W{iso_w} ({ws.strftime('%d %b')})"
        week_es = [e for e in all_entries if ws.isoformat() <= (e.get("date") or "") <= we.isoformat()]
        by_proc = {pt: round(sum(float(e.get("tonnage") or 0) for e in week_es if e.get("process_type") == pt), 4)
                   for pt in PROGRESS_PROCESS_TYPES}
        weekly_trend.append({"week": label, "week_start": ws.isoformat(), "week_end": we.isoformat(), "by_process": by_proc})

    # Prev week vs this week vs total
    this_ws = wstart
    this_we = wstart + timedelta(days=6)
    prev_ws = wstart - timedelta(weeks=1)
    prev_we = wstart - timedelta(days=1)

    prev_es = [e for e in all_entries if prev_ws.isoformat() <= (e.get("date") or "") <= prev_we.isoformat()]
    this_es = [e for e in all_entries if this_ws.isoformat() <= (e.get("date") or "") <= this_we.isoformat()]

    def _by_proc(entries_list):
        return {pt: round(sum(float(e.get("tonnage") or 0) for e in entries_list if e.get("process_type") == pt), 4)
                for pt in PROGRESS_PROCESS_TYPES}

    return jsonify({
        "ok": True,
        "item_tag_summary": item_tag_summary,
        "grand_totals": grand_totals,
        "total_gross_weight": total_gw,
        "process_types": PROGRESS_PROCESS_TYPES,
        "process_labels": PROGRESS_PROCESS_LABELS,
        "weekly_trend": weekly_trend,
        "prev_vs_this": {
            "prev_week": _by_proc(prev_es),
            "this_week": _by_proc(this_es),
            "total": _by_proc(all_entries),
        },
    })


@app.get("/api/management/equipment-tracking")
def equipment_tracking():
    """Return per-equipment tracking: current status, location, hours breakdown."""
    db = get_database()
    equipment_list = list(db.equipment_details.find({"is_active": True}))

    result = []
    for eq in equipment_list:
        eq_id = str(eq["_id"])

        entries = list(
            db.equipment_entries.find(
                {"equipment_id": eq_id},
                sort=[("work_date", -1), ("created_at", -1)]
            )
        )

        latest = entries[0] if entries else None
        total_hours = round(sum((e.get("hours") or 0) for e in entries), 2)

        hours_by_location: dict = {}
        hours_by_status: dict = {}
        count_by_status: dict = {}
        for e in entries:
            loc = (e.get("location") or "").strip() or "Unknown"
            st  = (e.get("equipment_status") or "").strip() or "Unknown"
            h   = e.get("hours") or 0
            hours_by_location[loc] = round(hours_by_location.get(loc, 0) + h, 2)
            hours_by_status[st]    = round(hours_by_status.get(st, 0) + h, 2)
            count_by_status[st]    = count_by_status.get(st, 0) + 1

        # Last entries for history display (with operator profile photo)
        slice_entries = entries[:20]
        civil_ids_hist = [e.get("civil_id") for e in slice_entries if e.get("civil_id")]
        pic_hist = _profile_picture_map(db, civil_ids_hist)
        history = []
        for e in slice_entries:
            cid = e.get("civil_id")
            history.append({
                "entry_id": str(e["_id"]),
                "work_date": e.get("work_date", ""),
                "equipment_status": e.get("equipment_status") or "",
                "location": e.get("location") or "",
                "hours": e.get("hours"),
                "activity": e.get("activity") or "",
                "operator_name": e.get("operator_name") or "",
                "civil_id": cid or "",
                "profile_picture": pic_hist.get(cid, "") if cid else "",
                "approval_status": e.get("approval_status") or e.get("status") or "",
            })

        result.append({
            "_id": eq_id,
            "name": eq.get("name", ""),
            "plate_number": eq.get("plate_number", ""),
            "equipment_type": eq.get("equipment_type", ""),
            "ownership": eq.get("ownership", ""),
            "base_location": eq.get("location", ""),
            "current_status": (latest.get("equipment_status") or "") if latest else "",
            "current_location": (latest.get("location") or "") if latest else "",
            "last_date": latest.get("work_date", "") if latest else "",
            "total_hours": total_hours,
            "entry_count": len(entries),
            "hours_by_location": hours_by_location,
            "hours_by_status": hours_by_status,
            "count_by_status": count_by_status,
            "history": history,
        })

    # Status summary across all equipment (latest entry per equipment)
    status_summary: dict = {}
    for item in result:
        st = item["current_status"] or "No Entry"
        status_summary[st] = status_summary.get(st, 0) + 1

    return jsonify({"ok": True, "equipment": result, "status_summary": status_summary})


_FRONTEND_BUILD = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "frontend", "build"))


@app.route("/api/health")
def health_check():
    return jsonify({"ok": True, "service": "enco-manpower"})


if os.path.isdir(_FRONTEND_BUILD):

    @app.route("/", defaults={"path": ""})
    @app.route("/<path:path>")
    def _serve_spa(path):
        if path.startswith("api"):
            return jsonify({"ok": False, "error": "not found"}), 404
        safe_root = os.path.realpath(_FRONTEND_BUILD)
        if path:
            candidate = os.path.realpath(os.path.join(_FRONTEND_BUILD, path))
            if candidate.startswith(safe_root) and os.path.isfile(candidate):
                return send_from_directory(_FRONTEND_BUILD, path)
        return send_from_directory(_FRONTEND_BUILD, "index.html")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=Config.FLASK_PORT, debug=Config.FLASK_ENV == "development")
